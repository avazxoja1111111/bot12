import logging
import asyncio
import json
import os
import re
import time
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import random

# Database imports
from database import (
    DatabaseService, init_database, User, Question, TestResult, 
    TestAnswer, Feedback, RegionData
)

from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message, KeyboardButton, ReplyKeyboardMarkup,
    ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup,
    CallbackQuery, BufferedInputFile
)
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage

# Excel va PDF uchun kutubxonalar
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:
    SimpleDocTemplate = None

# 🔑 Bot token va admin ID'lar
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN")
ADMINS = [int(x) for x in os.getenv("ADMIN_IDS", "6578706277,7853664401").split(",")]

# 🔧 Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 🔧 Bot va Dispatcher
storage = MemoryStorage()
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=storage)

# 🔧 Holatlar (FSM States)
class RegistrationStates(StatesGroup):
    waiting_for_child_name = State()
    waiting_for_parent_name = State()
    waiting_for_age_group = State()
    waiting_for_region = State()
    waiting_for_manual_region = State()
    waiting_for_district = State()
    waiting_for_manual_district = State()
    waiting_for_mahalla = State()

class TestStates(StatesGroup):
    taking_test = State()
    test_question = State()

class AdminStates(StatesGroup):
    waiting_for_broadcast = State()
    waiting_for_question_text = State()
    waiting_for_question_file = State()
    waiting_for_new_question = State()
    waiting_for_new_options = State()
    waiting_for_correct_answer = State()
    waiting_for_age_group_selection = State()
    waiting_for_delete_question_id = State()
    waiting_for_bulk_questions = State()

class FeedbackStates(StatesGroup):
    waiting_for_feedback = State()

# 🔧 Ma'lumotlar saqlash
user_data = {}
active_tests = {}
test_timers = {}

# 🔧 Legacy data files (for migration if needed)
DATA_DIR = "data"
USERS_FILE = os.path.join(DATA_DIR, "users.json")
QUESTIONS_FILE = os.path.join(DATA_DIR, "test_questions.json")
REGIONS_FILE = os.path.join(DATA_DIR, "regions.json")

def ensure_data_dir():
    """Ma'lumotlar papkasini yaratish (legacy support)"""
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

def load_data(filename: str, default: Any = None) -> Any:
    """JSON fayldan ma'lumot yuklash (legacy support)"""
    ensure_data_dir()
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return default or {}

def save_data(filename: str, data: Any):
    """JSON faylga ma'lumot saqlash (legacy support)"""
    ensure_data_dir()
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# 🔧 Database helper functions
def load_users_from_db() -> Dict[str, Any]:
    """Load users from database in legacy format for compatibility"""
    users = DatabaseService.get_all_users()
    result = {}
    for user in users:
        result[user.user_id] = {
            "child_name": user.child_name,
            "parent_name": user.parent_name,
            "phone": user.phone,
            "age_group": user.age_group,
            "region": user.region,
            "district": user.district,
            "mahalla": user.mahalla,
            "telegram_username": user.telegram_username,
            "telegram_name": user.telegram_name,
            "registered": user.registered,
            "registration_date": user.registration_date.isoformat(),
            "test_results": [
                {
                    "age_group": result.age_group,
                    "total_questions": result.total_questions,
                    "correct_answers": result.correct_answers,
                    "percentage": result.percentage,
                    "duration": result.duration_seconds,
                    "date": result.completed_at.isoformat()
                }
                for result in user.test_results
            ]
        }
    return result

def load_questions_from_db() -> Dict[str, List[Dict]]:
    """Load questions from database in legacy format"""
    result = {}
    for age_group in ["7-10", "11-14"]:
        questions = DatabaseService.get_questions_by_age_group(age_group)
        result[age_group] = [
            {
                "question": q.question_text,
                "options": q.options,
                "correct": q.correct_answer
            }
            for q in questions
        ]
    return result

def save_user_to_db(user_id: str, user_data: Dict[str, Any]):
    """Save user to database"""
    try:
        existing_user = DatabaseService.get_user(user_id)
        if existing_user:
            DatabaseService.update_user(user_id, user_data)
        else:
            user_data["user_id"] = user_id
            DatabaseService.create_user(user_data)
    except Exception as e:
        logger.error(f"Error saving user to database: {e}")

def save_question_to_db(question_data: Dict[str, Any]) -> bool:
    """Save question to database"""
    try:
        DatabaseService.add_question(question_data)
        return True
    except Exception as e:
        logger.error(f"Error saving question to database: {e}")
        return False

# 🔧 Viloyatlar va tumanlar ma'lumotlarini yuklash/yaratish
def init_regions_data():
    """Viloyatlar va tumanlar ma'lumotlarini boshlang'ich holatga keltirish"""
    # Try to load from database first
    try:
        regions_from_db = DatabaseService.get_regions()
        if regions_from_db:
            return regions_from_db
    except Exception as e:
        logger.warning(f"Could not load regions from database: {e}")
    
    # Fallback to default data
    regions_data = {
        "Toshkent": {
            "districts": ["Chilonzor", "Shayxontohur", "Yunusobod", "Mirzo Ulug'bek", "Yakkasaroy", "Mirobod"],
            "mahallas": {
                "Chilonzor": ["1-mahalla", "2-mahalla", "3-mahalla"],
                "Shayxontohur": ["Markaziy", "Sharqiy", "G'arbiy"],
                "Yunusobod": ["Yangi Yunusobod", "Eski Yunusobod"]
            }
        },
        "Samarqand": {
            "districts": ["Samarqand shahar", "Urgut", "Bulung'ur", "Ishtixon", "Kattaqo'rg'on"],
            "mahallas": {
                "Samarqand shahar": ["Registon", "Siob", "Darvozaboy"],
                "Urgut": ["Markaz", "Yangiobod"],
                "Bulung'ur": ["Markaz", "Qishloq"]
            }
        },
        "Farg'ona": {
            "districts": ["Farg'ona shahar", "Marg'ilon", "Qo'qon", "Beshariq", "Bog'dod"],
            "mahallas": {
                "Farg'ona shahar": ["Markaz", "Yangi shahar"],
                "Marg'ilon": ["Ipak yo'li", "Hunarmandchilik"],
                "Qo'qon": ["Eski shahar", "Yangi shahar"]
            }
        },
        "Andijon": {
            "districts": ["Andijon shahar", "Xo'jaobod", "Asaka", "Baliqchi", "Bo'z"],
            "mahallas": {
                "Andijon shahar": ["Markaz", "Sanoat"],
                "Xo'jaobod": ["Qishloq", "Shahar"],
                "Asaka": ["Markaz"]
            }
        },
        "Namangan": {
            "districts": ["Namangan shahar", "To'raqo'rg'on", "Uchqo'rg'on", "Chust", "Pop"],
            "mahallas": {
                "Namangan shahar": ["Markaz", "Sanoat zone"],
                "To'raqo'rg'on": ["Markaz", "Qishloq"],
                "Chust": ["Hunarmandchilik", "Dehqonchilik"]
            }
        },
        "Qashqadaryo": {
            "districts": ["Qarshi", "Shahrisabz", "Kitob", "Yakkabog'", "Chiroqchi"],
            "mahallas": {
                "Qarshi": ["Markaz", "Sanoat"],
                "Shahrisabz": ["Tarixiy markaz", "Yangi shahar"],
                "Kitob": ["Markaz"]
            }
        },
        "Surxondaryo": {
            "districts": ["Termiz", "Denov", "Qaraqo'l", "Sho'rchi", "Muzrabot"],
            "mahallas": {
                "Termiz": ["Markaz", "Port zone"],
                "Denov": ["Markaz", "Qishloq"],
                "Qaraqo'l": ["Markaz"]
            }
        },
        "Sirdaryo": {
            "districts": ["Guliston", "Yangiyer", "Sirdaryo", "Boyovut", "Mirzaobod"],
            "mahallas": {
                "Guliston": ["Markaz", "Sanoat"],
                "Yangiyer": ["Markaz", "Qishloq"],
                "Sirdaryo": ["Markaz"]
            }
        },
        "Jizzax": {
            "districts": ["Jizzax shahar", "G'allaorol", "Zomin", "Baxtiyor", "Mirzacho'l"],
            "mahallas": {
                "Jizzax shahar": ["Markaz", "Yangi shahar"],
                "G'allaorol": ["Markaz", "Qishloq"],
                "Zomin": ["Tog'li", "Tekislik"]
            }
        },
        "Navoiy": {
            "districts": ["Navoiy shahar", "Zarafshon", "Nurota", "Karmana", "Qiziltepa"],
            "mahallas": {
                "Navoiy shahar": ["Markaz", "Sanoat"],
                "Zarafshon": ["Oltin qazish", "Shahar"],
                "Nurota": ["Tarixiy", "Markaz"]
            }
        },
        "Buxoro": {
            "districts": ["Buxoro shahar", "G'ijduvon", "Kogon", "Olot", "Peshku"],
            "mahallas": {
                "Buxoro shahar": ["Eski shahar", "Yangi shahar"],
                "G'ijduvon": ["Markaz", "Qishloq"],
                "Kogon": ["Temir yo'l", "Markaz"]
            }
        },
        "Xorazm": {
            "districts": ["Urganch", "Xiva", "Bog'ot", "Gurlan", "Qo'shko'pir"],
            "mahallas": {
                "Urganch": ["Markaz", "Sanoat"],
                "Xiva": ["Ichan qal'a", "Tashqi shahar"],
                "Bog'ot": ["Markaz"]
            }
        }
    }
    
    existing_data = load_data(REGIONS_FILE, {})
    if not existing_data:
        save_data(REGIONS_FILE, regions_data)
    return regions_data

def init_test_questions():
    """Test savollarini boshlang'ich holatga keltirish - Empty questions as requested"""
    # Clear all existing questions from database
    try:
        count_7_10 = DatabaseService.clear_questions_by_age_group("7-10")
        count_11_14 = DatabaseService.clear_questions_by_age_group("11-14")
        logger.info(f"Cleared {count_7_10} questions from age group 7-10")
        logger.info(f"Cleared {count_11_14} questions from age group 11-14")
    except Exception as e:
        logger.warning(f"Could not clear questions from database: {e}")
    
    # Return empty structure
    questions_data = {
        "7-10": [],
        "11-14": []
    }
    
    # Clear legacy file as well
    save_data(QUESTIONS_FILE, questions_data)
    return questions_data

# 🔧 Ma'lumotlarni yuklash
def load_all_data():
    """Barcha ma'lumotlarni yuklash"""
    global user_data
    
    # Initialize database
    try:
        init_database()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Database initialization failed: {e}")
    
    # Load data
    try:
        users_data = load_users_from_db()
        if users_data:
            user_data.update(users_data)
        logger.info(f"Loaded {len(user_data)} users from database")
    except Exception as e:
        logger.warning(f"Could not load users from database: {e}")
        # Fallback to file
        user_data = load_data(USERS_FILE, {})
        logger.info(f"Loaded {len(user_data)} users from file")
    
    regions_data = init_regions_data()
    questions_data = init_test_questions()
    
    logger.info("All data loaded successfully")
    return regions_data, questions_data

# Global variables
regions_data, questions_data = load_all_data()

# 🔧 Keyboard yaratish funksiyalari
def create_main_menu():
    """Asosiy menyu tugmalari"""
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📝 Ro'yxatdan o'tish")],
            [KeyboardButton(text="📚 Loyiha haqida")],
            [KeyboardButton(text="📋 Test topshirish")],
            [KeyboardButton(text="💬 Fikr bildirish")]
        ],
        resize_keyboard=True
    )
    return kb

def create_admin_menu():
    """Admin menyu tugmalari"""
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👥 Foydalanuvchilar ro'yxati")],
            [KeyboardButton(text="📊 Statistika")],
            [KeyboardButton(text="❓ Savollarni boshqarish")],
            [KeyboardButton(text="📢 Xabar yuborish")],
            [KeyboardButton(text="📤 Ma'lumotlarni eksport qilish")],
            [KeyboardButton(text="🔙 Oddiy foydalanuvchi rejimiga qaytish")]
        ],
        resize_keyboard=True
    )
    return kb

def create_export_menu():
    """Eksport menyu tugmalari"""
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📊 Excel formatida", callback_data="export_excel")],
            [InlineKeyboardButton(text="📄 PDF formatida", callback_data="export_pdf")],
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_to_admin")]
        ]
    )
    return kb

def create_question_management_menu():
    """Savol boshqaruv menyu tugmalari"""
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="➕ Savol qo'shish", callback_data="add_question")],
            [InlineKeyboardButton(text="📝 Ko'p savol qo'shish", callback_data="add_bulk_questions")],
            [InlineKeyboardButton(text="📁 Fayl yuklash", callback_data="upload_questions_file")],
            [InlineKeyboardButton(text="👀 Savollarni ko'rish", callback_data="view_questions")],
            [InlineKeyboardButton(text="🗑 Savol o'chirish", callback_data="delete_question")],
            [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_to_admin")]
        ]
    )
    return kb

# 🔧 Excel eksport funksiyasi (yangilangan)
async def export_to_excel() -> Optional[BufferedInputFile]:
    """Foydalanuvchi ma'lumotlarini Excel formatida eksport qilish"""
    if not openpyxl:
        return None
    
    try:
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter
        
        # Load fresh data from database
        users_data = load_users_from_db()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Kitobxon Kids - Foydalanuvchilar"
        
        # Set up colors and styles
        header_fill = PatternFill(start_color="2E5266", end_color="2E5266", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)
        
        # Alternating row colors
        light_blue_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Project header
        ws.merge_cells('A1:J3')
        project_cell = ws['A1']
        project_cell.value = "📚 KITOBXON KIDS LOYIHASI\nFoydalanuvchilar ma'lumotlari\n" + datetime.now().strftime("%d.%m.%Y %H:%M")
        project_cell.font = Font(name="Arial", size=16, bold=True, color="2E5266")
        project_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        project_cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # Headers
        headers = [
            "№", "Bola ismi", "Ota-ona ismi", "Telefon", 
            "Yosh guruhi", "Viloyat", "Tuman", "Mahalla", 
            "Ro'yxatdan o'tgan sana", "Test natijalari"
        ]
        
        # Write headers starting from row 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # Data rows
        row = 6
        for idx, (user_id, user_info) in enumerate(users_data.items(), 1):
            # Determine row color
            fill_color = light_blue_fill if idx % 2 == 0 else white_fill
            
            # Write data
            data = [
                idx,
                user_info.get("child_name", ""),
                user_info.get("parent_name", ""),
                user_info.get("phone", ""),
                user_info.get("age_group", ""),
                user_info.get("region", ""),
                user_info.get("district", ""),
                user_info.get("mahalla", ""),
                user_info.get("registration_date", "").split("T")[0] if user_info.get("registration_date") else "",
                f"{len(user_info.get('test_results', []))} ta test"
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = data_font
                cell.fill = fill_color
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Summary section
        summary_row = row + 2
        ws.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_cell = ws[f'A{summary_row}']
        summary_cell.value = f"📊 JAMI: {len(users_data)} ta foydalanuvchi ro'yxatdan o'tgan"
        summary_cell.font = Font(name="Arial", size=12, bold=True, color="2E5266")
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_cell.fill = PatternFill(start_color="FFE135", end_color="FFE135", fill_type="solid")
        
        # Statistics by region
        region_stats = {}
        for user_info in users_data.values():
            region = user_info.get("region", "Noma'lum")
            region_stats[region] = region_stats.get(region, 0) + 1
        
        stats_row = summary_row + 2
        ws[f'A{stats_row}'] = "VILOYATLAR BO'YICHA STATISTIKA:"
        ws[f'A{stats_row}'].font = Font(name="Arial", size=11, bold=True, color="2E5266")
        
        stats_row += 1
        for region, count in sorted(region_stats.items()):
            ws[f'A{stats_row}'] = f"{region}: {count} ta"
            ws[f'A{stats_row}'].font = Font(name="Arial", size=10)
            stats_row += 1
        
        # Save to BytesIO
        from io import BytesIO
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        filename = f"kitobxon_kids_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return BufferedInputFile(excel_data.read(), filename=filename)
        
    except Exception as e:
        logger.error(f"Excel eksport xatosi: {e}")
        return None

# 🔧 PDF eksport funksiyasi (yangilangan)
async def export_to_pdf() -> Optional[BufferedInputFile]:
    """Foydalanuvchi ma'lumotlarini PDF formatida eksport qilish"""
    if not SimpleDocTemplate:
        return None
    
    try:
        from io import BytesIO
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        # Load fresh data from database
        users_data = load_users_from_db()
        
        # Create PDF buffer
        buffer = BytesIO()
        
        # Use landscape orientation for better table display
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            topMargin=0.5*inch,
            bottomMargin=0.5*inch,
            leftMargin=0.5*inch,
            rightMargin=0.5*inch
        )
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center alignment
            textColor=colors.darkblue
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=15,
            alignment=1,  # Center alignment
            textColor=colors.grey
        )
        
        # Add title
        title = Paragraph("📚 KITOBXON KIDS LOYIHASI", title_style)
        subtitle = Paragraph(f"Foydalanuvchilar ma'lumotlari - {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style)
        elements.append(title)
        elements.append(subtitle)
        elements.append(Spacer(1, 20))
        
        # Prepare table data
        headers = [
            "№", "Bola ismi", "Ota-ona ismi", "Telefon", 
            "Yosh", "Viloyat", "Tuman", "Mahalla", "Sana", "Testlar"
        ]
        
        table_data = [headers]
        
        # Add user data with proper text wrapping
        for idx, (user_id, user_info) in enumerate(users_data.items(), 1):
            row = [
                str(idx),
                user_info.get("child_name", "")[:15] + ("..." if len(user_info.get("child_name", "")) > 15 else ""),
                user_info.get("parent_name", "")[:15] + ("..." if len(user_info.get("parent_name", "")) > 15 else ""),
                user_info.get("phone", "")[:12],
                user_info.get("age_group", "")[:8],
                user_info.get("region", "")[:10],
                user_info.get("district", "")[:12],
                user_info.get("mahalla", "")[:10],
                user_info.get("registration_date", "").split("T")[0] if user_info.get("registration_date") else "",
                f"{len(user_info.get('test_results', []))}"
            ]
            table_data.append(row)
        
        # Create table with proper column widths
        col_widths = [0.4*inch, 1.2*inch, 1.2*inch, 0.9*inch, 0.6*inch, 0.8*inch, 1.0*inch, 0.8*inch, 0.8*inch, 0.5*inch]
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Enhanced table style
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            
            # Data row styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Add summary statistics
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=10,
            alignment=1,
            textColor=colors.darkgreen
        )
        
        summary_text = f"📊 JAMI: {len(users_data)} ta foydalanuvchi ro'yxatdan o'tgan"
        summary = Paragraph(summary_text, summary_style)
        elements.append(summary)
        
        # Regional statistics
        region_stats = {}
        for user_info in users_data.values():
            region = user_info.get("region", "Noma'lum")
            region_stats[region] = region_stats.get(region, 0) + 1
        
        if region_stats:
            elements.append(Spacer(1, 10))
            stats_title = Paragraph("VILOYATLAR BO'YICHA STATISTIKA:", summary_style)
            elements.append(stats_title)
            
            stats_data = [["Viloyat", "Foydalanuvchilar soni"]]
            for region, count in sorted(region_stats.items()):
                stats_data.append([region, str(count)])
            
            stats_table = Table(stats_data, colWidths=[2*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.darkblue),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightcyan),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(stats_table)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF data
        pdf_data = buffer.getvalue()
        buffer.close()
        
        filename = f"kitobxon_kids_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        return BufferedInputFile(pdf_data, filename=filename)
        
    except Exception as e:
        logger.error(f"PDF eksport xatosi: {e}")
        return None

# 🔧 START komandasi
@dp.message(CommandStart())
async def start_cmd(message: Message, state: FSMContext):
    """Botni boshlash komandasi"""
    await state.clear()
    
    user_id = str(message.from_user.id)
    first_name = message.from_user.first_name or "Foydalanuvchi"
    
    # Admin ekanligini tekshirish
    if message.from_user.id in ADMINS:
        await message.answer(
            f"👋 Salom {first_name}!\n\n"
            "🔧 Siz admin sifatida kirdingiz.\n"
            "Admin panelidan foydalanish uchun menyuni tanlang:",
            reply_markup=create_admin_menu()
        )
        return
    
    # Oddiy foydalanuvchi uchun
    welcome_text = f"""
🌟 **Assalomu aleykum, {first_name}!** 🌟

📚 **"Kitobxon Kids"** loyihasiga xush kelibsiz!

🎯 **Loyihaning maqsadi:**
✅ Bolalarning kitob o'qish qiziqishini oshirish
✅ O'zbek adabiyoti va madaniyatini targ'ib qilish
✅ Bilim darajasini baholash va rivojlantirish

👶 **Yosh toifalari:**
• 7-10 yosh guruhi
• 11-14 yosh guruhi

🏆 **Sizni kutayotgan imkoniyatlar:**
📖 Qiziqarli testlar
📊 Natijalarni kuzatish
🎁 G'oliblar uchun sovrinlar
📈 Bilim darajasini oshirish

🚀 Boshlash uchun quyidagi tugmalardan birini tanlang!
"""
    
    await message.answer(
        welcome_text,
        reply_markup=create_main_menu(),
        parse_mode="Markdown"
    )

# 🔧 Ro'yxatdan o'tish tugmasi
@dp.message(F.text == "📝 Ro'yxatdan o'tish")
async def register_cmd(message: Message, state: FSMContext):
    """Ro'yxatdan o'tish jarayonini boshlash"""
    user_id = str(message.from_user.id)
    
    # Allaqachon ro'yxatdan o'tganligini tekshirish
    if user_id in user_data and user_data[user_id].get("registered"):
        user_info = user_data[user_id]
        await message.answer(
            f"✅ Siz allaqachon ro'yxatdan o'tgansiz!\n\n"
            f"👶 Bola: {user_info.get('child_name')}\n"
            f"👨‍👩‍👧 Ota/Ona: {user_info.get('parent_name')}\n"
            f"📌 Yosh toifa: {user_info.get('age_group')}\n"
            f"🏢 Manzil: {user_info.get('region')}, {user_info.get('district')}, {user_info.get('mahalla')}"
        )
        return
    
    # Telefon raqam so'rash o'rniga to'g'ridan-to'g'ri ro'yxatdan o'tishni boshlash
    await message.answer(
        "👶 **Ro'yxatdan o'tish boshlandi!**\n\n"
        "Iltimos, bolaning to'liq ismini kiriting:",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_child_name)

# 🔧 Bolaning ismini olish
@dp.message(RegistrationStates.waiting_for_child_name)
async def get_child_name(message: Message, state: FSMContext):
    """Bolaning ismini saqlash"""
    child_name = message.text.strip()
    
    if len(child_name) < 2:
        await message.answer("⚠️ Iltimos, bolaning to'liq ismini kiriting (kamida 2 ta harf):")
        return
    
    user_id = str(message.from_user.id)
    if user_id not in user_data:
        user_data[user_id] = {}
    
    user_data[user_id]["child_name"] = child_name
    user_data[user_id]["user_id"] = message.from_user.id
    user_data[user_id]["telegram_username"] = message.from_user.username or ""
    user_data[user_id]["telegram_name"] = message.from_user.first_name or ""
    
    await message.answer(
        f"✅ Bolaning ismi: **{child_name}**\n\n"
        "👨‍👩‍👧 Endi ota-onaning to'liq ismini kiriting:",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_parent_name)

# 🔧 Ota-onaning ismini olish
@dp.message(RegistrationStates.waiting_for_parent_name)
async def get_parent_name(message: Message, state: FSMContext):
    """Ota-onaning ismini saqlash"""
    parent_name = message.text.strip()
    
    if len(parent_name) < 2:
        await message.answer("⚠️ Iltimos, ota-onaning to'liq ismini kiriting (kamida 2 ta harf):")
        return
    
    user_id = str(message.from_user.id)
    user_data[user_id]["parent_name"] = parent_name
    
    # Yosh guruhini tanlash
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="7–10 yosh")],
            [KeyboardButton(text="11–14 yosh")]
        ],
        resize_keyboard=True
    )
    
    await message.answer(
        f"✅ Ota-ona: **{parent_name}**\n\n"
        "📌 Bolaning yosh guruhini tanlang:",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_age_group)

# 🔧 Yosh guruhini olish
@dp.message(RegistrationStates.waiting_for_age_group)
async def get_age_group(message: Message, state: FSMContext):
    """Yosh guruhini saqlash"""
    age_group = message.text.strip()
    
    if age_group not in ["7–10 yosh", "11–14 yosh"]:
        await message.answer("⚠️ Iltimos, yuqoridagi tugmalardan birini tanlang:")
        return
    
    user_id = str(message.from_user.id)
    user_data[user_id]["age_group"] = age_group
    
    # Viloyatlarni ko'rsatish
    keyboard = []
    for region in regions_data.keys():
        keyboard.append([KeyboardButton(text=region)])
    keyboard.append([KeyboardButton(text="📝 Qo'lda kiritish")])
    
    kb = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    await message.answer(
        f"✅ Yosh guruhi: **{age_group}**\n\n"
        "🏢 Viloyatingizni tanlang yoki qo'lda kiriting:",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_region)

# 🔧 Viloyatni olish
@dp.message(RegistrationStates.waiting_for_region)
async def get_region(message: Message, state: FSMContext):
    """Viloyatni saqlash"""
    region = message.text.strip()
    user_id = str(message.from_user.id)
    
    if region == "📝 Qo'lda kiritish":
        await message.answer(
            "✍️ Viloyat nomini yozib kiriting:",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(RegistrationStates.waiting_for_manual_region)
        return
    
    if region not in regions_data:
        await message.answer("⚠️ Iltimos, ro'yxatdan viloyatni tanlang yoki 'Qo'lda kiritish' tugmasini bosing.")
        return
    
    user_data[user_id]["region"] = region
    
    # Tumanlarni ko'rsatish
    districts = regions_data[region]["districts"]
    keyboard = []
    for district in districts:
        keyboard.append([KeyboardButton(text=district)])
    keyboard.append([KeyboardButton(text="📝 Qo'lda kiritish")])
    
    kb = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    await message.answer(
        f"✅ Viloyat: **{region}**\n\n"
        "🏙 Tuman/Shaharni tanlang yoki qo'lda kiriting:",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_district)

# 🔧 Qo'lda viloyat kiritish
@dp.message(RegistrationStates.waiting_for_manual_region)
async def get_manual_region(message: Message, state: FSMContext):
    """Qo'lda kiritilgan viloyatni saqlash"""
    region = message.text.strip()
    
    if len(region) < 2:
        await message.answer("⚠️ Iltimos, viloyat nomini to'liq kiriting:")
        return
    
    user_id = str(message.from_user.id)
    user_data[user_id]["region"] = region
    
    await message.answer(
        f"✅ Viloyat: **{region}**\n\n"
        "🏙 Tuman/Shahar nomini yozib kiriting:",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_manual_district)

# 🔧 Tumanni olish
@dp.message(RegistrationStates.waiting_for_district)
async def get_district(message: Message, state: FSMContext):
    """Tumanni saqlash"""
    district = message.text.strip()
    user_id = str(message.from_user.id)
    
    if district == "📝 Qo'lda kiritish":
        await message.answer(
            "✍️ Tuman/Shahar nomini yozib kiriting:",
            reply_markup=ReplyKeyboardRemove()
        )
        await state.set_state(RegistrationStates.waiting_for_manual_district)
        return
    
    user_data[user_id]["district"] = district
    
    await message.answer(
        f"✅ Tuman/Shahar: **{district}**\n\n"
        "🏡 Mahalla nomini yozib kiriting:",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_mahalla)

# 🔧 Qo'lda tuman kiritish
@dp.message(RegistrationStates.waiting_for_manual_district)
async def get_manual_district(message: Message, state: FSMContext):
    """Qo'lda kiritilgan tumanni saqlash"""
    district = message.text.strip()
    
    if len(district) < 2:
        await message.answer("⚠️ Iltimos, tuman/shahar nomini to'liq kiriting:")
        return
    
    user_id = str(message.from_user.id)
    user_data[user_id]["district"] = district
    
    await message.answer(
        f"✅ Tuman/Shahar: **{district}**\n\n"
        "🏡 Mahalla nomini yozib kiriting:",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_mahalla)

# 🔧 Mahallani olish
@dp.message(RegistrationStates.waiting_for_mahalla)
async def get_mahalla(message: Message, state: FSMContext):
    """Mahallani saqlash va ro'yxatdan o'tishni yakunlash"""
    mahalla = message.text.strip()
    
    if len(mahalla) < 2:
        await message.answer("⚠️ Iltimos, mahalla nomini to'liq kiriting:")
        return
    
    user_id = str(message.from_user.id)
    user_data[user_id]["mahalla"] = mahalla
    user_data[user_id]["registered"] = True
    user_data[user_id]["registration_date"] = datetime.now().isoformat()
    user_data[user_id]["test_results"] = []
    
    # Ma'lumotlarni saqlash
    save_user_to_db(user_id, user_data[user_id])
    
    # Ro'yxat tugadi - ma'lumotlarni ko'rsatish
    user_info = user_data[user_id]
    success_text = f"""
✅ **Tabriklaymiz! Siz muvaffaqiyatli ro'yxatdan o'tdingiz!** 🎉

📋 **Sizning ma'lumotlaringiz:**
👶 **Bola ismi:** {user_info['child_name']}
👨‍👩‍👧 **Ota-ona:** {user_info['parent_name']}
📌 **Yosh guruhi:** {user_info['age_group']}
🏢 **Viloyat:** {user_info['region']}
🏙 **Tuman/Shahar:** {user_info['district']}
🏡 **Mahalla:** {user_info['mahalla']}
📅 **Sana:** {datetime.now().strftime('%d.%m.%Y %H:%M')}

🚀 Endi testlarni topshirish, loyiha haqida ma'lumot olish va fikr bildirish imkoniyatiga egasiz!
"""
    
    await message.answer(
        success_text,
        reply_markup=create_main_menu(),
        parse_mode="Markdown"
    )
    
    # Adminlarga xabar yuborish
    admin_text = f"""
📥 **Yangi foydalanuvchi ro'yxatdan o'tdi!**

👤 **Telegram:** @{user_info.get('telegram_username', 'username yo\'q')} ({user_info.get('telegram_name', 'Noma\'lum')})
👶 **Bola:** {user_info['child_name']}
👨‍👩‍👧 **Ota-ona:** {user_info['parent_name']}
📌 **Yosh:** {user_info['age_group']}
📍 **Manzil:** {user_info['region']}, {user_info['district']}, {user_info['mahalla']}
📅 **Sana:** {datetime.now().strftime('%d.%m.%Y %H:%M')}
"""
    
    for admin_id in ADMINS:
        try:
            await bot.send_message(admin_id, admin_text, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Adminga xabar yuborishda xato: {e}")
    
    await state.clear()

# 🔧 Test topshirish tugmasi
@dp.message(F.text == "📋 Test topshirish")
async def test_cmd(message: Message):
    """Test topshirish funksiyasi"""
    user_id = str(message.from_user.id)
    
    # Ro'yxatdan o'tganligini tekshirish
    if user_id not in user_data or not user_data[user_id].get("registered"):
        await message.answer(
            "❌ Test topshirish uchun avval ro'yxatdan o'ting!\n\n"
            "Ro'yxatdan o'tish uchun '📝 Ro'yxatdan o'tish' tugmasini bosing."
        )
        return
    
    # Testlar mavjudligini tekshirish
    questions = load_questions_from_db()
    age_group = user_data[user_id]["age_group"]
    age_key = "7-10" if "7" in age_group else "11-14"
    
    if not questions.get(age_key) or len(questions[age_key]) == 0:
        await message.answer(
            "😔 Afsuski, hozircha sizning yosh guruhingiz uchun testlar mavjud emas.\n\n"
            "Tez orada testlar qo'shiladi. Kuting!"
        )
        return
    
    await message.answer(
        f"📋 **Test topshirish**\n\n"
        f"👶 **Bola:** {user_data[user_id]['child_name']}\n"
        f"📌 **Yosh guruhi:** {age_group}\n\n"
        f"🔢 **Savollar soni:** {len(questions[age_key])} ta\n"
        f"⏰ **Har bir savol uchun vaqt:** 20 soniya\n\n"
        f"❗ **Eslatma:** Test boshlangandan keyin uni to'xtatib bo'lmaydi!\n\n"
        f"Testni boshlashga tayyormisiz?",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="🚀 Testni boshlash", callback_data="start_test")],
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="cancel_test")]
            ]
        ),
        parse_mode="Markdown"
    )

# 🔧 Loyiha haqida tugmasi
@dp.message(F.text == "📚 Loyiha haqida")
async def about_project(message: Message):
    """Loyiha haqida ma'lumot"""
    about_text = """
📚 **"KITOBXON KIDS" LOYIHASI**

🎯 **Maqsad:**
Bolalarning kitobxonlik ko'nikmalarini rivojlantirish va o'zbek adabiyotiga bo'lgan qiziqishini oshirish.

👥 **Ishtirokchilar:**
• 7-10 yosh guruhi
• 11-14 yosh guruhi

📋 **Loyiha tuzilmasi:**
✅ Ro'yxatdan o'tish
✅ Bilim darajasini baholash testlari
✅ Natijalarni tahlil qilish
✅ G'oliblarni aniqlash

🏆 **Mukofotlar:**
🥇 Har yosh guruhidan g'oliblar
🎁 Qimmatbaho sovrinlar
📜 Sertifikatlar
📚 Kitoblar to'plami

📞 **Aloqa:**
Savollar bo'lsa, fikr bildirish bo'limidan foydalaning.

🚀 **Omad tilaymiz!**
"""
    
    await message.answer(about_text, parse_mode="Markdown")

# 🔧 Fikr bildirish tugmasi
@dp.message(F.text == "💬 Fikr bildirish")
async def feedback_cmd(message: Message, state: FSMContext):
    """Fikr bildirish funksiyasi"""
    user_id = str(message.from_user.id)
    
    # Ro'yxatdan o'tganligini tekshirish
    if user_id not in user_data or not user_data[user_id].get("registered"):
        await message.answer(
            "❌ Fikr bildirish uchun avval ro'yxatdan o'ting!\n\n"
            "Ro'yxatdan o'tish uchun '📝 Ro'yxatdan o'tish' tugmasini bosing."
        )
        return
    
    await message.answer(
        "💬 **Fikr bildirish**\n\n"
        "Sizning fikr va takliflaringiz biz uchun muhim!\n\n"
        "Iltimos, o'z fikringizni yozib yuboring:",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(FeedbackStates.waiting_for_feedback)

# 🔧 Fikrni qabul qilish
@dp.message(FeedbackStates.waiting_for_feedback)
async def get_feedback(message: Message, state: FSMContext):
    """Foydalanuvchi fikrini saqlash"""
    feedback_text = message.text.strip()
    
    if len(feedback_text) < 10:
        await message.answer("⚠️ Iltimos, kamida 10 ta belgidan iborat fikr yuboring:")
        return
    
    user_id = str(message.from_user.id)
    user_info = user_data[user_id]
    
    # Fikrni saqlash
    feedback_data = {
        "user_id": user_id,
        "feedback_text": feedback_text,
        "phone": user_info.get("phone", ""),
        "telegram_username": user_info.get("telegram_username", "")
    }
    
    try:
        DatabaseService.save_feedback(feedback_data)
    except Exception as e:
        logger.error(f"Feedback saqlashda xato: {e}")
    
    await message.answer(
        "✅ **Rahmat!**\n\n"
        "Sizning fikringiz qabul qilindi va adminlarga yuborildi.\n"
        "Tez orada sizga javob beramiz!",
        reply_markup=create_main_menu(),
        parse_mode="Markdown"
    )
    
    # Adminlarga yuborish
    admin_text = f"""
💬 **Yangi fikr va taklif!**

👤 **Foydalanuvchi:** {user_info.get('child_name')} ({user_info.get('parent_name')})
📱 **Telegram:** @{user_info.get('telegram_username', 'username yo\'q')}
📞 **Telefon:** {user_info.get('phone', 'Ko\'rsatilmagan')}
📍 **Manzil:** {user_info.get('region')}, {user_info.get('district')}

💭 **Fikr:**
{feedback_text}

📅 **Sana:** {datetime.now().strftime('%d.%m.%Y %H:%M')}
"""
    
    for admin_id in ADMINS:
        try:
            await bot.send_message(admin_id, admin_text, parse_mode="Markdown")
        except Exception as e:
            logger.error(f"Adminga fikr yuborishda xato: {e}")
    
    await state.clear()

# 🔧 ADMIN PANEL FUNKSIYALARI

# Admin panel asosiy menyusi
@dp.message(F.text == "👥 Foydalanuvchilar ro'yxati")
async def admin_users_list(message: Message):
    """Foydalanuvchilar ro'yxati (admin)"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    users_data = load_users_from_db()
    
    if not users_data:
        await message.answer("📭 Hech kim ro'yxatdan o'tmagan.")
        return
    
    # Statistika
    total_users = len(users_data)
    age_groups = {"7–10 yosh": 0, "11–14 yosh": 0}
    regions = {}
    
    for user_info in users_data.values():
        age_group = user_info.get("age_group", "Noma'lum")
        if age_group in age_groups:
            age_groups[age_group] += 1
        
        region = user_info.get("region", "Noma'lum")
        regions[region] = regions.get(region, 0) + 1
    
    # Oxirgi 10 ta foydalanuvchi
    sorted_users = sorted(
        users_data.items(),
        key=lambda x: x[1].get("registration_date", ""),
        reverse=True
    )[:10]
    
    users_text = "👥 **Oxirgi 10 ta foydalanuvchi:**\n\n"
    for i, (user_id, user_info) in enumerate(sorted_users, 1):
        users_text += f"{i}. {user_info.get('child_name', 'Noma\'lum')} ({user_info.get('age_group', 'Noma\'lum')})\n"
        users_text += f"   📍 {user_info.get('region', 'Noma\'lum')}, {user_info.get('district', 'Noma\'lum')}\n\n"
    
    stats_text = f"""
📊 **STATISTIKA:**

👥 **Jami foydalanuvchilar:** {total_users} ta

📊 **Yosh guruhlari bo'yicha:**
• 7-10 yosh: {age_groups['7–10 yosh']} ta
• 11-14 yosh: {age_groups['11–14 yosh']} ta

🏢 **Eng faol viloyatlar:**
"""
    
    # Top 5 viloyat
    top_regions = sorted(regions.items(), key=lambda x: x[1], reverse=True)[:5]
    for region, count in top_regions:
        stats_text += f"• {region}: {count} ta\n"
    
    full_text = stats_text + "\n" + users_text
    
    await message.answer(full_text, parse_mode="Markdown")

@dp.message(F.text == "📊 Statistika")
async def admin_statistics(message: Message):
    """Umumiy statistika (admin)"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    users_data = load_users_from_db()
    questions_data = load_questions_from_db()
    
    # Asosiy statistika
    total_users = len(users_data)
    total_questions_7_10 = len(questions_data.get("7-10", []))
    total_questions_11_14 = len(questions_data.get("11-14", []))
    
    # Yosh guruhlari
    age_stats = {"7–10 yosh": 0, "11–14 yosh": 0}
    
    # Viloyatlar statistikasi
    region_stats = {}
    
    # Test natijalari
    total_tests = 0
    
    for user_info in users_data.values():
        # Yosh guruh statistikasi
        age_group = user_info.get("age_group", "Noma'lum")
        if age_group in age_stats:
            age_stats[age_group] += 1
        
        # Viloyat statistikasi
        region = user_info.get("region", "Noma'lum")
        region_stats[region] = region_stats.get(region, 0) + 1
        
        # Test statistikasi
        test_results = user_info.get("test_results", [])
        total_tests += len(test_results)
    
    stats_text = f"""
📊 **LOYIHA STATISTIKASI**

👥 **FOYDALANUVCHILAR:**
• Jami ro'yxatdan o'tganlar: **{total_users}** ta
• 7-10 yosh guruhi: **{age_stats['7–10 yosh']}** ta
• 11-14 yosh guruhi: **{age_stats['11–14 yosh']}** ta

❓ **SAVOLLAR BAZASI:**
• 7-10 yosh uchun: **{total_questions_7_10}** ta savol
• 11-14 yosh uchun: **{total_questions_11_14}** ta savol

📋 **TESTLAR:**
• Jami topshirilgan testlar: **{total_tests}** ta

🏢 **VILOYATLAR BO'YICHA TAQSIMOT:**
"""
    
    # Viloyatlar bo'yicha tartiblash
    sorted_regions = sorted(region_stats.items(), key=lambda x: x[1], reverse=True)
    for region, count in sorted_regions:
        percentage = (count / total_users * 100) if total_users > 0 else 0
        stats_text += f"• {region}: **{count}** ta ({percentage:.1f}%)\n"
    
    stats_text += f"\n📅 **So'ngi yangilanish:** {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    
    await message.answer(stats_text, parse_mode="Markdown")

@dp.message(F.text == "❓ Savollarni boshqarish")
async def admin_questions(message: Message):
    """Savollarni boshqarish (admin)"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    questions_data = load_questions_from_db()
    
    stats_text = f"""
❓ **SAVOLLAR BOSHQARUVI**

📊 **Hozirgi holat:**
• 7-10 yosh: **{len(questions_data.get('7-10', []))}** ta savol
• 11-14 yosh: **{len(questions_data.get('11-14', []))}** ta savol

🔧 **Amallar:**
"""
    
    await message.answer(
        stats_text,
        reply_markup=create_question_management_menu(),
        parse_mode="Markdown"
    )

@dp.message(F.text == "📢 Xabar yuborish")
async def admin_broadcast(message: Message, state: FSMContext):
    """Umumiy xabar yuborish (admin)"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    users_data = load_users_from_db()
    await message.answer(
        f"📢 **Umumiy xabar yuborish**\n\n"
        f"👥 Jami foydalanuvchilar: **{len(users_data)}** ta\n\n"
        f"Yubormoqchi bo'lgan xabaringizni yozing:",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_broadcast)

@dp.message(AdminStates.waiting_for_broadcast)
async def send_broadcast(message: Message, state: FSMContext):
    """Xabarni barcha foydalanuvchilarga yuborish"""
    if message.from_user.id not in ADMINS:
        return
    
    broadcast_text = message.text
    users_data = load_users_from_db()
    
    await message.answer("📤 Xabar yuborilmoqda...")
    
    sent_count = 0
    failed_count = 0
    
    for user_id in users_data.keys():
        try:
            await bot.send_message(user_id, broadcast_text, parse_mode="Markdown")
            sent_count += 1
            await asyncio.sleep(0.1)  # Rate limiting
        except Exception as e:
            failed_count += 1
            logger.error(f"Xabar yuborishda xato {user_id}: {e}")
    
    await message.answer(
        f"✅ **Xabar yuborish yakunlandi!**\n\n"
        f"📤 Yuborildi: **{sent_count}** ta\n"
        f"❌ Yuborilmadi: **{failed_count}** ta",
        parse_mode="Markdown"
    )
    await state.clear()

@dp.message(F.text == "📤 Ma'lumotlarni eksport qilish")
async def admin_export(message: Message):
    """Ma'lumotlarni eksport qilish (admin)"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    users_data = load_users_from_db()
    await message.answer(
        f"📤 **Ma'lumotlarni eksport qilish**\n\n"
        f"👥 Jami foydalanuvchilar: **{len(users_data)}** ta\n\n"
        f"Qaysi formatda eksport qilmoqchisiz?",
        reply_markup=create_export_menu(),
        parse_mode="Markdown"
    )

@dp.message(F.text == "🔙 Oddiy foydalanuvchi rejimiga qaytish")
async def admin_exit(message: Message):
    """Admin rejimidan chiqish"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqi yo'q!")
        return
    
    await message.answer(
        "👤 Oddiy foydalanuvchi rejimiga qaytdingiz.",
        reply_markup=create_main_menu()
    )

# 🔧 CALLBACK QUERY HANDLERS

@dp.callback_query(F.data == "export_excel")
async def export_excel_callback(callback: CallbackQuery):
    """Excel formatida eksport qilish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.answer("📊 Excel fayl tayyorlanmoqda...")
    await callback.message.edit_text("📊 Excel fayl tayyorlanmoqda...")
    
    excel_file = await export_to_excel()
    
    if excel_file:
        await callback.message.answer_document(
            excel_file,
            caption=f"📊 **Foydalanuvchilar ma'lumotlari (Excel)**\n\n"
                   f"📅 Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                   f"👥 Foydalanuvchilar: {len(load_users_from_db())} ta",
            parse_mode="Markdown"
        )
        await callback.message.edit_text("✅ Excel fayl yuborildi!")
    else:
        await callback.message.edit_text("❌ Excel fayl yaratishda xatolik yuz berdi!")

@dp.callback_query(F.data == "export_pdf")
async def export_pdf_callback(callback: CallbackQuery):
    """PDF formatida eksport qilish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.answer("📄 PDF fayl tayyorlanmoqda...")
    await callback.message.edit_text("📄 PDF fayl tayyorlanmoqda...")
    
    pdf_file = await export_to_pdf()
    
    if pdf_file:
        await callback.message.answer_document(
            pdf_file,
            caption=f"📄 **Foydalanuvchilar ma'lumotlari (PDF)**\n\n"
                   f"📅 Sana: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                   f"👥 Foydalanuvchilar: {len(load_users_from_db())} ta",
            parse_mode="Markdown"
        )
        await callback.message.edit_text("✅ PDF fayl yuborildi!")
    else:
        await callback.message.edit_text("❌ PDF fayl yaratishda xatolik yuz berdi!")

@dp.callback_query(F.data == "back_to_admin")
async def back_to_admin_callback(callback: CallbackQuery):
    """Admin paneliga qaytish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "🔧 **Admin Panel**\n\nKerakli amalni tanlang:",
        reply_markup=None,
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data == "view_questions")
async def view_questions_callback(callback: CallbackQuery):
    """Savollarni ko'rish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    questions_data = load_questions_from_db()
    
    text = "👀 **BARCHA SAVOLLAR:**\n\n"
    
    for age_group in ["7-10", "11-14"]:
        questions = questions_data.get(age_group, [])
        text += f"📚 **{age_group} yosh guruhi:** {len(questions)} ta savol\n\n"
        
        if questions:
            for i, q in enumerate(questions[:5], 1):  # Faqat birinchi 5 ta savolni ko'rsatish
                text += f"{i}. {q['question'][:50]}...\n"
            if len(questions) > 5:
                text += f"   ... va yana {len(questions) - 5} ta savol\n"
        else:
            text += "   ❌ Savollar mavjud emas\n"
        text += "\n"
    
    await callback.message.edit_text(text, parse_mode="Markdown")
    await callback.answer()

@dp.callback_query(F.data == "add_question")
async def add_question_callback(callback: CallbackQuery, state: FSMContext):
    """Bitta savol qo'shish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "➕ **Yangi savol qo'shish**\n\n"
        "Qaysi yosh guruhi uchun savol qo'shmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="7-10 yosh", callback_data="age_7_10")],
                [InlineKeyboardButton(text="11-14 yosh", callback_data="age_11_14")],
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="back_questions")]
            ]
        ),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("age_"))
async def select_age_group_callback(callback: CallbackQuery, state: FSMContext):
    """Yosh guruhini tanlash"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    age_group = "7-10" if callback.data == "age_7_10" else "11-14"
    await state.update_data(selected_age_group=age_group)
    
    await callback.message.edit_text(
        f"➕ **{age_group} yosh guruhi uchun savol**\n\n"
        "Savol matnini yozing:",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_new_question)
    await callback.answer()

@dp.message(AdminStates.waiting_for_new_question)
async def get_new_question(message: Message, state: FSMContext):
    """Yangi savol matnini olish"""
    if message.from_user.id not in ADMINS:
        return
    
    question_text = message.text.strip()
    
    if len(question_text) < 10:
        await message.answer("⚠️ Savol kamida 10 ta belgidan iborat bo'lishi kerak!")
        return
    
    await state.update_data(question_text=question_text)
    
    await message.answer(
        f"✅ Savol: {question_text}\n\n"
        "Endi javob variantlarini yozing.\n"
        "Har bir variant yangi qatorda bo'lsin:\n\n"
        "Masalan:\n"
        "Toshkent\n"
        "Samarqand\n"
        "Buxoro\n"
        "Farg'ona"
    )
    await state.set_state(AdminStates.waiting_for_new_options)

@dp.message(AdminStates.waiting_for_new_options)
async def get_new_options(message: Message, state: FSMContext):
    """Javob variantlarini olish"""
    if message.from_user.id not in ADMINS:
        return
    
    options_text = message.text.strip()
    options = [opt.strip() for opt in options_text.split('\n') if opt.strip()]
    
    if len(options) < 2:
        await message.answer("⚠️ Kamida 2 ta javob varianti bo'lishi kerak!")
        return
    
    await state.update_data(options=options)
    
    # Variantlarni raqamlash va ko'rsatish
    options_display = ""
    for i, option in enumerate(options, 1):
        options_display += f"{i}. {option}\n"
    
    await message.answer(
        f"✅ Javob variantlari:\n{options_display}\n"
        f"To'g'ri javob raqamini kiriting (1-{len(options)}):"
    )
    await state.set_state(AdminStates.waiting_for_correct_answer)

@dp.message(AdminStates.waiting_for_correct_answer)
async def get_correct_answer(message: Message, state: FSMContext):
    """To'g'ri javobni olish"""
    if message.from_user.id not in ADMINS:
        return
    
    try:
        correct_index = int(message.text.strip()) - 1
        data = await state.get_data()
        options = data['options']
        
        if correct_index < 0 or correct_index >= len(options):
            await message.answer(f"⚠️ 1 dan {len(options)} gacha raqam kiriting!")
            return
        
        # Savolni saqlash
        question_data = {
            "question_text": data['question_text'],
            "options": options,
            "correct_answer": correct_index,
            "age_group": data['selected_age_group']
        }
        
        success = save_question_to_db(question_data)
        
        if success:
            await message.answer(
                f"✅ **Savol muvaffaqiyatli qo'shildi!**\n\n"
                f"📚 Yosh guruhi: {data['selected_age_group']}\n"
                f"❓ Savol: {data['question_text']}\n"
                f"✅ To'g'ri javob: {options[correct_index]}",
                reply_markup=create_admin_menu(),
                parse_mode="Markdown"
            )
        else:
            await message.answer(
                "❌ Savolni saqlashda xatolik yuz berdi!",
                reply_markup=create_admin_menu()
            )
        
        await state.clear()
        
    except ValueError:
        await message.answer("⚠️ Faqat raqam kiriting!")

@dp.callback_query(F.data == "add_bulk_questions")
async def add_bulk_questions_callback(callback: CallbackQuery, state: FSMContext):
    """Ko'p savol qo'shish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    help_text = """
📝 **Ko'p savol qo'shish**

Savollarni quyidagi formatda yozing:

1. Savol matni? A) Variant 1 B) Variant 2 C) Variant 3 D) Variant 4 Javob: A

2. Ikkinchi savol? A) Variant 1 B) Variant 2 C) Variant 3 D) Variant 4 Javob: C

**Eslatma:** Har bir savoldan keyin bo'sh qator qoldiring.
"""
    
    await callback.message.edit_text(
        help_text,
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="7-10 yosh", callback_data="bulk_age_7-10")],
                [InlineKeyboardButton(text="11-14 yosh", callback_data="bulk_age_11-14")],
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="questions_menu")]
            ]
        ),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("bulk_age_"))
async def bulk_age_selection_callback(callback: CallbackQuery, state: FSMContext):
    """Yosh guruhi tanlab ko'p savol qo'shish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    age_group = callback.data.split("_", 2)[2]
    await state.update_data(bulk_age_group=age_group)
    await state.set_state(AdminStates.waiting_for_bulk_questions)
    
    await callback.message.edit_text(
        f"📝 **{age_group} yosh guruhi uchun savollarni kiriting:**\n\n"
        "Formatni to'g'ri rioya qiling:\n"
        "1. Savol? A) Variant B) Variant C) Variant D) Variant Javob: A\n\n"
        "Savollarni yozing:",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="questions_menu")]]
        ),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(AdminStates.waiting_for_bulk_questions)
async def process_bulk_questions(message: Message, state: FSMContext):
    """Ko'p savollarni qayta ishlash"""
    if message.from_user.id not in ADMINS:
        return
    
    data = await state.get_data()
    age_group = data.get('bulk_age_group')
    text = message.text.strip()
    
    # Parse questions using regex
    pattern = r'(\d+\.\s*.+?\?\s*A\)\s*.+?\s*B\)\s*.+?\s*C\)\s*.+?\s*D\)\s*.+?\s*Javob:\s*[ABCD])'
    found_questions = re.findall(pattern, text, re.DOTALL | re.IGNORECASE)
    
    if not found_questions:
        await message.answer("❌ Savollar topilmadi! Format to'g'ri emasligini tekshiring.")
        return
    
    added_count = 0
    failed_count = 0
    
    for question_text in found_questions:
        try:
            # Extract question and options
            parts = re.split(r'\s*[ABCD]\)\s*', question_text)
            if len(parts) < 5:
                failed_count += 1
                continue
            
            question = parts[0].split('?')[0].strip() + '?'
            
            # Extract options
            options = []
            for i in range(1, 5):
                option = parts[i].split('Javob:')[0].strip() if 'Javob:' in parts[i] else parts[i].strip()
                options.append(option)
            
            # Extract correct answer
            answer_match = re.search(r'Javob:\s*([ABCD])', question_text, re.IGNORECASE)
            if not answer_match:
                failed_count += 1
                continue
            
            answer_letter = answer_match.group(1).upper()
            correct_answer = ord(answer_letter) - ord('A')
            
            # Save to database
            question_data = {
                "question_text": question,
                "options": options,
                "correct_answer": correct_answer,
                "age_group": age_group
            }
            
            if save_question_to_db(question_data):
                added_count += 1
            else:
                failed_count += 1
                
        except Exception as e:
            logger.error(f"Error processing question: {e}")
            failed_count += 1
    
    result_message = f"✅ **Natijalar:**\n\n"
    result_message += f"➕ Qo'shilgan savollar: {added_count}\n"
    if failed_count > 0:
        result_message += f"❌ Xatolik yuz bergan: {failed_count}\n"
    
    await message.answer(
        result_message,
        reply_markup=create_admin_menu(),
        parse_mode="Markdown"
    )
    await state.clear()

@dp.callback_query(F.data == "upload_questions_file")
async def upload_questions_file_callback(callback: CallbackQuery, state: FSMContext):
    """Fayl yuklash"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "📁 **Savollar faylini yuklang**\n\n"
        "Qo'llab-quvvatlanadigan formatlar:\n"
        "• .txt - matn fayli\n"
        "• .docx - Word hujjati\n"
        "• .xlsx - Excel jadval\n\n"
        "Faylni yuboring:",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="🔙 Orqaga", callback_data="questions_menu")]]
        ),
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_question_file)
    await callback.answer()

@dp.message(AdminStates.waiting_for_question_file)
async def process_questions_file(message: Message, state: FSMContext):
    """Yuklangan faylni qayta ishlash"""
    if message.from_user.id not in ADMINS:
        return
    
    if not message.document:
        await message.answer("❌ Iltimos, fayl yuboring!")
        return
    
    file_name = message.document.file_name.lower()
    supported_formats = ['.txt', '.docx', '.xlsx']
    
    if not any(file_name.endswith(fmt) for fmt in supported_formats):
        await message.answer("❌ Fayl formati qo'llab-quvvatlanmaydi!")
        return
    
    await message.answer("⏳ Fayl yuklanmoqda va qayta ishlanmoqda...")
    
    try:
        # Download file
        file_info = await bot.get_file(message.document.file_id)
        file_data = await bot.download_file(file_info.file_path)
        
        # Process file based on type
        questions_text = ""
        
        if file_name.endswith('.txt'):
            questions_text = file_data.read().decode('utf-8')
            
        elif file_name.endswith('.docx'):
            try:
                from docx import Document
                import io
                doc = Document(io.BytesIO(file_data.read()))
                questions_text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            except ImportError:
                await message.answer("❌ Word fayllari uchun qo'shimcha kutubxona kerak!")
                await state.clear()
                return
                
        elif file_name.endswith('.xlsx'):
            try:
                import io
                import pandas as pd
                df = pd.read_excel(io.BytesIO(file_data.read()))
                questions_text = df.to_string(index=False)
            except ImportError:
                await message.answer("❌ Excel fayllari uchun qo'shimcha kutubxona kerak!")
                await state.clear()
                return
        
        # Parse questions
        pattern = r'(\d+\.\s*.+?\?\s*A\)\s*.+?\s*B\)\s*.+?\s*C\)\s*.+?\s*D\)\s*.+?\s*Javob:\s*[ABCD])'
        found_questions = re.findall(pattern, questions_text, re.DOTALL | re.IGNORECASE)
        
        if not found_questions:
            await message.answer("❌ Faylda savollar topilmadi! Format to'g'ri emasligini tekshiring.")
            await state.clear()
            return
        
        # Ask for age group
        await message.answer(
            f"📊 **{len(found_questions)} ta savol topildi!**\n\n"
            "Yosh guruhini tanlang:",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="7-10 yosh", callback_data="file_age_7-10")],
                    [InlineKeyboardButton(text="11-14 yosh", callback_data="file_age_11-14")],
                    [InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="questions_menu")]
                ]
            ),
            parse_mode="Markdown"
        )
        
        await state.update_data(parsed_questions=found_questions)
        
    except Exception as e:
        logger.error(f"Error processing file: {e}")
        await message.answer("❌ Faylni qayta ishlashda xatolik yuz berdi!")
        await state.clear()

@dp.callback_query(F.data.startswith("file_age_"))
async def file_age_selection_callback(callback: CallbackQuery, state: FSMContext):
    """Fayl uchun yosh guruhi tanlab savollarni saqlash"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    age_group = callback.data.split("_", 2)[2]
    data = await state.get_data()
    found_questions = data.get('parsed_questions', [])
    
    added_count = 0
    failed_count = 0
    
    for question_text in found_questions:
        try:
            # Extract question and options (same logic as bulk questions)
            parts = re.split(r'\s*[ABCD]\)\s*', question_text)
            if len(parts) < 5:
                failed_count += 1
                continue
            
            question = parts[0].split('?')[0].strip() + '?'
            
            options = []
            for i in range(1, 5):
                option = parts[i].split('Javob:')[0].strip() if 'Javob:' in parts[i] else parts[i].strip()
                options.append(option)
            
            answer_match = re.search(r'Javob:\s*([ABCD])', question_text, re.IGNORECASE)
            if not answer_match:
                failed_count += 1
                continue
            
            answer_letter = answer_match.group(1).upper()
            correct_answer = ord(answer_letter) - ord('A')
            
            question_data = {
                "question_text": question,
                "options": options,
                "correct_answer": correct_answer,
                "age_group": age_group
            }
            
            if save_question_to_db(question_data):
                added_count += 1
            else:
                failed_count += 1
                
        except Exception as e:
            logger.error(f"Error processing question: {e}")
            failed_count += 1
    
    result_message = f"✅ **Fayl qayta ishlandi!**\n\n"
    result_message += f"📚 Yosh guruhi: {age_group}\n"
    result_message += f"➕ Qo'shilgan savollar: {added_count}\n"
    if failed_count > 0:
        result_message += f"❌ Xatolik yuz bergan: {failed_count}\n"
    
    await callback.message.edit_text(
        result_message,
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[[InlineKeyboardButton(text="🔙 Admin panel", callback_data="back_to_admin")]]
        ),
        parse_mode="Markdown"
    )
    await callback.answer()
    await state.clear()

@dp.callback_query(F.data == "view_questions")
async def view_questions_callback(callback: CallbackQuery):
    """Savollarni ko'rish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    try:
        questions_7_10 = DatabaseService.get_questions_by_age_group("7-10")
        questions_11_14 = DatabaseService.get_questions_by_age_group("11-14")
        
        stats_text = f"📊 **Savollar statistikasi:**\n\n"
        stats_text += f"👶 7-10 yosh: {len(questions_7_10)} ta savol\n"
        stats_text += f"🧒 11-14 yosh: {len(questions_11_14)} ta savol\n"
        stats_text += f"📝 Jami: {len(questions_7_10) + len(questions_11_14)} ta savol"
        
        await callback.message.edit_text(
            stats_text,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="👶 7-10 yosh savollar", callback_data="show_7-10")],
                    [InlineKeyboardButton(text="🧒 11-14 yosh savollar", callback_data="show_11-14")],
                    [InlineKeyboardButton(text="🔙 Orqaga", callback_data="questions_menu")]
                ]
            ),
            parse_mode="Markdown"
        )
        await callback.answer()
        
    except Exception as e:
        logger.error(f"Error viewing questions: {e}")
        await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)

@dp.callback_query(F.data.startswith("show_"))
async def show_age_group_questions(callback: CallbackQuery):
    """Yosh guruhi savollarini ko'rsatish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    age_group = callback.data.split("_")[1]
    
    try:
        questions = DatabaseService.get_questions_by_age_group(age_group)
        
        if not questions:
            await callback.message.edit_text(
                f"📝 **{age_group} yosh guruhi:**\n\n❌ Hozircha savollar yo'q",
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[[InlineKeyboardButton(text="🔙 Orqaga", callback_data="view_questions")]]
                ),
                parse_mode="Markdown"
            )
        else:
            # Show first few questions
            text = f"📝 **{age_group} yosh guruhi ({len(questions)} ta savol):**\n\n"
            
            for i, q in enumerate(questions[:5], 1):
                text += f"{i}. {q.question_text}\n"
                for j, option in enumerate(q.options):
                    marker = "✅" if j == q.correct_answer else "◦"
                    text += f"   {marker} {chr(65+j)}) {option}\n"
                text += "\n"
            
            if len(questions) > 5:
                text += f"... va yana {len(questions) - 5} ta savol"
            
            await callback.message.edit_text(
                text,
                reply_markup=InlineKeyboardMarkup(
                    inline_keyboard=[
                        [InlineKeyboardButton(text="🗑 Hammasi o'chirish", callback_data=f"clear_all_{age_group}")],
                        [InlineKeyboardButton(text="🔙 Orqaga", callback_data="view_questions")]
                    ]
                ),
                parse_mode="Markdown"
            )
        
        await callback.answer()
        
    except Exception as e:
        logger.error(f"Error showing questions: {e}")
        await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)

@dp.callback_query(F.data.startswith("clear_all_"))
async def clear_all_questions_callback(callback: CallbackQuery):
    """Barcha savollarni o'chirish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    age_group = callback.data.split("_", 2)[2]
    
    await callback.message.edit_text(
        f"⚠️ **Diqqat!**\n\n"
        f"Haqiqatan ham {age_group} yosh guruhi uchun barcha savollarni o'chirmoqchimisiz?\n\n"
        f"Bu amalni bekor qilib bo'lmaydi!",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="✅ Ha, o'chir", callback_data=f"confirm_clear_{age_group}")],
                [InlineKeyboardButton(text="❌ Yo'q, bekor qil", callback_data=f"show_{age_group}")]
            ]
        ),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("confirm_clear_"))
async def confirm_clear_questions(callback: CallbackQuery):
    """Savollarni o'chirishni tasdiqlash"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    age_group = callback.data.split("_", 2)[2]
    
    try:
        deleted_count = DatabaseService.clear_questions_by_age_group(age_group)
        
        await callback.message.edit_text(
            f"✅ **Muvaffaqiyatli o'chirildi!**\n\n"
            f"📚 Yosh guruhi: {age_group}\n"
            f"🗑 O'chirilgan savollar: {deleted_count} ta",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[[InlineKeyboardButton(text="🔙 Savollar menu", callback_data="questions_menu")]]
            ),
            parse_mode="Markdown"
        )
        await callback.answer()
        
    except Exception as e:
        logger.error(f"Error clearing questions: {e}")
        await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)

@dp.callback_query(F.data == "delete_question")
async def delete_question_callback(callback: CallbackQuery):
    """Savol o'chirish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "🗑 **Savol o'chirish**\n\n"
        "Avval yosh guruhini tanlang:",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="7-10 yosh", callback_data="del_age_7-10")],
                [InlineKeyboardButton(text="11-14 yosh", callback_data="del_age_11-14")],
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="questions_menu")]
            ]
        ),
        parse_mode="Markdown"
    )
    await callback.answer()

# Main menu handlers
@dp.message(F.text == "📚 Loyiha haqida")
async def project_info(message: Message):
    """Loyiha haqida ma'lumot"""
    info_text = """
📚 **Kitobxon Kids Loyihasi**

🎯 **Maqsad:** O'zbekiston bolalarining kitobxonlik va umumiy bilim darajasini oshirish

👶 **Yosh guruhlari:**
• 7-10 yosh
• 11-14 yosh

📋 **Imkoniyatlar:**
• Ro'yxatdan o'tish
• Bilim darajasini sinash
• Natijalarni ko'rish
• Fikr bildirish

🏆 **Maqsadlar:**
• Bolalarda o'qish sevgisini rivojlantirish
• Bilimlarni baholash va mustahkamlash
• Ta'limiy jarayonni qiziqarli qilish

📞 **Qo'llab-quvvatlash:**
Savollaringiz bo'lsa, admin bilan bog'laning.
"""
    
    await message.answer(
        info_text,
        parse_mode="Markdown"
    )

@dp.message(F.text == "📝 Ro'yxatdan o'tish")
async def start_registration(message: Message, state: FSMContext):
    """Ro'yxatdan o'tish jarayonini boshlash"""
    user_id = str(message.from_user.id)
    
    # Check if user is already registered
    existing_user = DatabaseService.get_user(user_id)
    if existing_user:
        await message.answer(
            f"✅ Siz allaqachon ro'yxatdan o'tgansiz!\n\n"
            f"👶 Bola ismi: {existing_user.child_name}\n"
            f"👨‍👩‍👧‍👦 Ota-ona: {existing_user.parent_name}\n"
            f"📚 Yosh guruhi: {existing_user.age_group}\n"
            f"🏢 Viloyat: {existing_user.region}\n"
            f"🏙 Tuman: {existing_user.district}",
            parse_mode="Markdown"
        )
        return
    
    await message.answer(
        "📝 **Ro'yxatdan o'tish boshlandi!**\n\n"
        "👶 Bolangizning to'liq ismini kiriting:",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_child_name)

@dp.message(RegistrationStates.waiting_for_child_name)
async def get_child_name(message: Message, state: FSMContext):
    """Bola ismini olish"""
    await state.update_data(child_name=message.text.strip())
    
    await message.answer(
        "👨‍👩‍👧‍👦 **Ota-ona ismini kiriting:**\n\n"
        "To'liq ism-sharifingizni yozing:",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_parent_name)

@dp.message(RegistrationStates.waiting_for_parent_name)
async def get_parent_name(message: Message, state: FSMContext):
    """Ota-ona ismini olish"""
    await state.update_data(parent_name=message.text.strip())
    
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="7-10 yosh")],
            [KeyboardButton(text="11-14 yosh")]
        ],
        resize_keyboard=True
    )
    
    await message.answer(
        "📚 **Yosh guruhini tanlang:**",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_age_group)

@dp.message(RegistrationStates.waiting_for_age_group)
async def get_age_group(message: Message, state: FSMContext):
    """Yosh guruhini olish"""
    age_group = message.text.strip()
    
    if age_group not in ["7-10 yosh", "11-14 yosh"]:
        await message.answer("❌ Iltimos, tugmalardan birini tanlang!")
        return
    
    await state.update_data(age_group=age_group)
    
    # Load regions
    regions = init_regions_data()
    region_buttons = [[KeyboardButton(text=region)] for region in regions.keys()]
    region_buttons.append([KeyboardButton(text="🖊 Qo'lda kiriting")])
    
    kb = ReplyKeyboardMarkup(
        keyboard=region_buttons,
        resize_keyboard=True
    )
    
    await message.answer(
        "🏢 **Viloyatni tanlang:**",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_region)

@dp.message(RegistrationStates.waiting_for_region)
async def get_region(message: Message, state: FSMContext):
    """Viloyatni olish"""
    region = message.text.strip()
    
    if region == "🖊 Qo'lda kiriting":
        await message.answer(
            "🖊 **Viloyat nomini yozing:**",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode="Markdown"
        )
        await state.set_state(RegistrationStates.waiting_for_manual_region)
        return
    
    await state.update_data(region=region)
    
    # Get districts for this region
    regions = init_regions_data()
    districts = regions.get(region, {}).get("districts", [])
    
    if districts:
        district_buttons = [[KeyboardButton(text=district)] for district in districts]
        district_buttons.append([KeyboardButton(text="🖊 Qo'lda kiriting")])
        
        kb = ReplyKeyboardMarkup(
            keyboard=district_buttons,
            resize_keyboard=True
        )
        
        await message.answer(
            "🏙 **Tuman/Shaharni tanlang:**",
            reply_markup=kb,
            parse_mode="Markdown"
        )
        await state.set_state(RegistrationStates.waiting_for_district)
    else:
        await message.answer(
            "🏙 **Tuman/Shahar nomini yozing:**",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode="Markdown"
        )
        await state.set_state(RegistrationStates.waiting_for_manual_district)

@dp.message(RegistrationStates.waiting_for_manual_region)
async def get_manual_region(message: Message, state: FSMContext):
    """Qo'lda kiritilgan viloyatni olish"""
    await state.update_data(region=message.text.strip())
    
    await message.answer(
        "🏙 **Tuman/Shahar nomini yozing:**",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_manual_district)

@dp.message(RegistrationStates.waiting_for_district)
async def get_district(message: Message, state: FSMContext):
    """Tumanni olish"""
    district = message.text.strip()
    
    if district == "🖊 Qo'lda kiriting":
        await message.answer(
            "🖊 **Tuman/Shahar nomini yozing:**",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode="Markdown"
        )
        await state.set_state(RegistrationStates.waiting_for_manual_district)
        return
    
    await state.update_data(district=district)
    
    await message.answer(
        "🏡 **Mahalla nomini yozing:**",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_mahalla)

@dp.message(RegistrationStates.waiting_for_manual_district)
async def get_manual_district(message: Message, state: FSMContext):
    """Qo'lda kiritilgan tumanni olish"""
    await state.update_data(district=message.text.strip())
    
    await message.answer(
        "🏡 **Mahalla nomini yozing:**",
        parse_mode="Markdown"
    )
    await state.set_state(RegistrationStates.waiting_for_mahalla)

@dp.message(RegistrationStates.waiting_for_mahalla)
async def get_mahalla(message: Message, state: FSMContext):
    """Mahallani olish va ro'yxatdan o'tishni yakunlash"""
    data = await state.get_data()
    data['mahalla'] = message.text.strip()
    
    # Save to database
    user_data = {
        "user_id": str(message.from_user.id),
        "child_name": data['child_name'],
        "parent_name": data['parent_name'],
        "phone": "",  # Will be updated when phone is shared
        "age_group": data['age_group'].replace(" yosh", ""),  # Convert to database format
        "region": data['region'],
        "district": data['district'],
        "mahalla": data['mahalla'],
        "telegram_username": message.from_user.username or "",
        "telegram_name": f"{message.from_user.first_name or ''} {message.from_user.last_name or ''}".strip()
    }
    
    try:
        DatabaseService.create_user(user_data)
        
        success_text = f"""
✅ **Ro'yxatdan o'tish yakunlandi!**

👶 **Bola:** {data['child_name']}
👨‍👩‍👧‍👦 **Ota-ona:** {data['parent_name']}
📚 **Yosh guruhi:** {data['age_group']}
🏢 **Viloyat:** {data['region']}
🏙 **Tuman:** {data['district']}
🏡 **Mahalla:** {data['mahalla']}

🎉 Tabriklaymiz! Endi testlarni topshirishingiz mumkin.
"""
        
        await message.answer(
            success_text,
            reply_markup=create_main_menu(),
            parse_mode="Markdown"
        )
        
        # Notify admins
        admin_text = f"""
📥 **Yangi foydalanuvchi ro'yxatdan o'tdi!**

👤 **Telegram:** @{message.from_user.username or 'noma\'lum'} ({message.from_user.id})
{success_text}
"""
        
        for admin_id in ADMINS:
            try:
                await bot.send_message(admin_id, admin_text, parse_mode="Markdown")
            except:
                pass
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error saving user: {e}")
        await message.answer(
            "❌ Ro'yxatdan o'tishda xatolik yuz berdi. Qaytadan urinib ko'ring.",
            reply_markup=create_main_menu()
        )
        await state.clear()

@dp.message(F.text == "📋 Test topshirish")
async def start_test(message: Message):
    """Test topshirishni boshlash"""
    user_id = str(message.from_user.id)
    
    # Check if user is registered
    user = DatabaseService.get_user(user_id)
    if not user:
        await message.answer(
            "❌ Avval ro'yxatdan o'ting!\n\n"
            "📝 Ro'yxatdan o'tish tugmasini bosing.",
            reply_markup=create_main_menu()
        )
        return
    
    # Check if there are questions for user's age group
    questions = DatabaseService.get_questions_by_age_group(user.age_group)
    if not questions:
        await message.answer(
            "❌ Sizning yosh guruhingiz uchun hozircha savollar mavjud emas.\n\n"
            "Iltimos, keyinroq qaytadan urinib ko'ring.",
            reply_markup=create_main_menu()
        )
        return
    
    await message.answer(
        f"📋 **Test boshlash**\n\n"
        f"📚 Yosh guruhi: {user.age_group} yosh\n"
        f"❓ Savollar soni: {len(questions)}\n"
        f"⏱ Har bir savol uchun 20 soniya vaqt\n\n"
        f"Tayyormisiz?",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="▶️ Testni boshlash", callback_data="start_test")],
                [InlineKeyboardButton(text="🔙 Orqaga", callback_data="main_menu")]
            ]
        ),
        parse_mode="Markdown"
    )

@dp.callback_query(F.data == "start_test")
async def begin_test(callback: CallbackQuery, state: FSMContext):
    """Testni boshlash"""
    user_id = str(callback.from_user.id)
    
    user = DatabaseService.get_user(user_id)
    questions = DatabaseService.get_questions_by_age_group(user.age_group)
    
    if not questions:
        await callback.answer("❌ Savollar mavjud emas!", show_alert=True)
        return
    
    # Shuffle questions and take 25 (or all if less than 25)
    import random
    random.shuffle(questions)
    test_questions = questions[:min(25, len(questions))]
    
    # Initialize test data
    test_data = {
        "questions": [q.id for q in test_questions],
        "current_question": 0,
        "answers": {},
        "start_time": time.time(),
        "age_group": user.age_group
    }
    
    await state.update_data(test_data=test_data)
    await state.set_state(TestStates.taking_test)
    
    # Show first question
    await show_test_question(callback.message, state, callback.from_user.id)
    await callback.answer()

async def show_test_question(message: Message, state: FSMContext, user_id: int):
    """Test savolini ko'rsatish"""
    data = await state.get_data()
    test_data = data['test_data']
    
    current_index = test_data['current_question']
    questions_ids = test_data['questions']
    
    if current_index >= len(questions_ids):
        # Test finished
        await finish_test(message, state, user_id)
        return
    
    # Get current question
    try:
        question_id = questions_ids[current_index]
        question = DatabaseService.get_session().query(Question).filter(Question.id == question_id).first()
        
        if not question:
            await message.edit_text("❌ Savol topilmadi!")
            return
        
        # Create options keyboard
        options_keyboard = []
        for i, option in enumerate(question.options):
            options_keyboard.append([
                InlineKeyboardButton(
                    text=f"{chr(65+i)}) {option}", 
                    callback_data=f"answer_{current_index}_{i}"
                )
            ])
        
        options_keyboard.append([
            InlineKeyboardButton(text="⏭ O'tkazib yuborish", callback_data=f"skip_{current_index}")
        ])
        
        question_text = f"""
📋 **Test** ({current_index + 1}/{len(questions_ids)})

❓ **Savol:** {question.question_text}

⏱ Vaqt: 20 soniya
"""
        
        await message.edit_text(
            question_text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=options_keyboard),
            parse_mode="Markdown"
        )
        
    except Exception as e:
        logger.error(f"Error showing question: {e}")
        await message.edit_text("❌ Savol ko'rsatishda xatolik!")

@dp.callback_query(F.data.startswith("answer_"))
async def process_answer(callback: CallbackQuery, state: FSMContext):
    """Javobni qayta ishlash"""
    try:
        parts = callback.data.split("_")
        question_index = int(parts[1])
        selected_answer = int(parts[2])
        
        data = await state.get_data()
        test_data = data['test_data']
        
        # Save answer
        test_data['answers'][question_index] = selected_answer
        
        # Move to next question
        test_data['current_question'] += 1
        
        await state.update_data(test_data=test_data)
        
        # Show next question or finish test
        await show_test_question(callback.message, state, callback.from_user.id)
        await callback.answer()
        
    except Exception as e:
        logger.error(f"Error processing answer: {e}")
        await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)

@dp.callback_query(F.data.startswith("skip_"))
async def skip_question(callback: CallbackQuery, state: FSMContext):
    """Savolni o'tkazib yuborish"""
    try:
        question_index = int(callback.data.split("_")[1])
        
        data = await state.get_data()
        test_data = data['test_data']
        
        # Mark as skipped (no answer saved)
        test_data['current_question'] += 1
        
        await state.update_data(test_data=test_data)
        
        # Show next question or finish test
        await show_test_question(callback.message, state, callback.from_user.id)
        await callback.answer("⏭ Savol o'tkazib yuborildi")
        
    except Exception as e:
        logger.error(f"Error skipping question: {e}")
        await callback.answer("❌ Xatolik yuz berdi!", show_alert=True)

async def finish_test(message: Message, state: FSMContext, user_id: int):
    """Testni yakunlash"""
    try:
        data = await state.get_data()
        test_data = data['test_data']
        
        # Calculate results
        questions_ids = test_data['questions']
        answers = test_data['answers']
        
        correct_count = 0
        total_questions = len(questions_ids)
        
        session = DatabaseService.get_session()
        
        for i, question_id in enumerate(questions_ids):
            if i in answers:
                question = session.query(Question).filter(Question.id == question_id).first()
                if question and answers[i] == question.correct_answer:
                    correct_count += 1
        
        session.close()
        
        percentage = round((correct_count / total_questions) * 100) if total_questions > 0 else 0
        duration = int(time.time() - test_data['start_time'])
        
        # Save test result to database
        result_data = {
            "user_id": str(user_id),
            "age_group": test_data['age_group'],
            "total_questions": total_questions,
            "correct_answers": correct_count,
            "percentage": percentage,
            "duration_seconds": duration
        }
        
        DatabaseService.save_test_result(result_data)
        
        # Show results
        result_text = f"""
🎉 **Test yakunlandi!**

📊 **Natijalar:**
✅ To'g'ri javoblar: {correct_count}/{total_questions}
📈 Foiz: {percentage}%
⏱ Vaqt: {duration // 60}:{duration % 60:02d}

{"🏆 A'lo natija!" if percentage >= 80 else "👍 Yaxshi natija!" if percentage >= 60 else "📚 Ko'proq o'qing!"}
"""
        
        await message.edit_text(
            result_text,
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="🔙 Bosh menyu", callback_data="main_menu")]
                ]
            ),
            parse_mode="Markdown"
        )
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error finishing test: {e}")
        await message.edit_text("❌ Testni yakunlashda xatolik yuz berdi!")

@dp.callback_query(F.data == "main_menu")
async def main_menu_callback(callback: CallbackQuery):
    """Bosh menyuga qaytish"""
    await callback.message.delete()
    await callback.message.answer(
        "🏠 **Bosh menyu**\n\nKerakli bo'limni tanlang:",
        reply_markup=create_main_menu(),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.message(F.text == "💬 Fikr bildirish")
async def feedback_start(message: Message, state: FSMContext):
    """Fikr bildirish"""
    user_id = str(message.from_user.id)
    
    # Check if user is registered
    user = DatabaseService.get_user(user_id)
    if not user:
        await message.answer(
            "❌ Avval ro'yxatdan o'ting!\n\n"
            "📝 Ro'yxatdan o'tish tugmasini bosing.",
            reply_markup=create_main_menu()
        )
        return
    
    await message.answer(
        "💬 **Fikr bildirish**\n\n"
        "Loyiha haqida fikr va takliflaringizni yozing.\n"
        "Barcha xabarlar adminlarga yuboriladi:",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(FeedbackStates.waiting_for_feedback)

@dp.message(FeedbackStates.waiting_for_feedback)
async def process_feedback(message: Message, state: FSMContext):
    """Fikr-mulohazani qayta ishlash"""
    user_id = str(message.from_user.id)
    feedback_text = message.text.strip()
    
    # Get user data
    user = DatabaseService.get_user(user_id)
    
    # Save feedback to database
    feedback_data = {
        "user_id": user_id,
        "feedback_text": feedback_text,
        "phone": user.phone if user else "",
        "telegram_username": message.from_user.username or ""
    }
    
    try:
        DatabaseService.save_feedback(feedback_data)
        
        await message.answer(
            "✅ **Fikr-mulohaza yuborildi!**\n\n"
            "Rahmat! Sizning fikringiz adminlarga yetkazildi.",
            reply_markup=create_main_menu(),
            parse_mode="Markdown"
        )
        
        # Send to admins
        admin_text = f"""
💬 **Yangi fikr-mulohaza:**

👤 **Foydalanuvchi:** {user.child_name if user else 'Noma\'lum'} ({message.from_user.id})
📱 **Username:** @{message.from_user.username or 'noma\'lum'}
📞 **Telefon:** {user.phone if user else 'noma\'lum'}

💭 **Xabar:**
{feedback_text}

📅 **Sana:** {datetime.now().strftime('%d.%m.%Y %H:%M')}
"""
        
        for admin_id in ADMINS:
            try:
                await bot.send_message(admin_id, admin_text, parse_mode="Markdown")
            except:
                pass
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error saving feedback: {e}")
        await message.answer(
            "❌ Fikr yuborishda xatolik yuz berdi. Qaytadan urinib ko'ring.",
            reply_markup=create_main_menu()
        )
        await state.clear()

# Admin handlers (continue from the previous implementation...)

@dp.message(F.text == "🔧 Admin panel")
async def admin_panel(message: Message):
    """Admin panel"""
    if message.from_user.id not in ADMINS:
        return
    
    await message.answer(
        "🔧 **Admin Panel**\n\n"
        "Boshqaruv paneliga xush kelibsiz!",
        reply_markup=create_admin_menu(),
        parse_mode="Markdown"
    )

# Callback handlers for navigation
@dp.callback_query(F.data == "questions_menu")
async def questions_menu_callback(callback: CallbackQuery):
    """Savollar menyusiga qaytish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.edit_text(
        "❓ **Savollarni boshqarish**\n\nKerakli amalni tanlang:",
        reply_markup=create_question_management_menu(),
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data == "back_to_admin")
async def back_to_admin_callback(callback: CallbackQuery):
    """Admin panelga qaytish"""
    if callback.from_user.id not in ADMINS:
        await callback.answer("❌ Ruxsat yo'q!", show_alert=True)
        return
    
    await callback.message.delete()
    await callback.message.answer(
        "🔧 **Admin Panel**\n\n"
        "Boshqaruv paneliga xush kelibsiz!",
        reply_markup=create_admin_menu(),
        parse_mode="Markdown"
    )
    await callback.answer()

# Start command and main function
@dp.message(CommandStart())
async def start_command(message: Message):
    """Botni boshlash komandasi"""
    user_id = str(message.from_user.id)
    first_name = message.from_user.first_name or "Foydalanuvchi"
    
    # Check if user is admin
    if message.from_user.id in ADMINS:
        welcome_text = f"""
🤖 **Admin rejimida xush kelibsiz, {first_name}!**

Siz admin huquqlariga egasiz.

🔧 Admin panel uchun: /admin
👤 Oddiy foydalanuvchi rejimi uchun: pastdagi tugmalardan foydalaning
"""
    else:
        # Check if user is already registered
        user = DatabaseService.get_user(user_id)
        
        if user:
            welcome_text = f"""
🎉 **Qaytib kelganingizdan xursandmiz, {first_name}!**

Siz allaqachon ro'yxatdan o'tgansiz. Testlarni topshirishni boshlashingiz yoki loyiha haqida ko'proq ma'lumot olishingiz mumkin.
"""
        else:
            welcome_text = f"""
🌟 **Kitobxon Kids loyihasiga xush kelibsiz, {first_name}!**

📚 Bu loyiha orqali siz:
• Bilimlaringizni sinab ko'rishingiz
• Kitobxonlik darajangizni baholashingiz
• Qiziqarli savollar bilan tanishishingiz mumkin

🎯 **Boshlash uchun ro'yxatdan o'ting!**
"""
    
    await message.answer(
        welcome_text,
        reply_markup=create_main_menu(),
        parse_mode="Markdown"
    )

@dp.message(Command("admin"))
async def admin_command(message: Message):
    """Admin komandasi"""
    if message.from_user.id not in ADMINS:
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    await admin_panel(message)

# Main function
async def main():
    """Botni ishga tushirish"""
    logger.info("Bot ishga tushirilmoqda...")
    
    # Initialize database
    try:
        init_database()
        logger.info("Database initialized successfully")
    except Exception as e:
        logger.error(f"Database initialization failed: {e}")
        return
    
    # Load initial data
    try:
        global regions_data, questions_data
        regions_data = init_regions_data()
        questions_data = init_test_questions()
        logger.info("Initial data loaded successfully")
    except Exception as e:
        logger.error(f"Failed to load initial data: {e}")
    
    logger.info("Bot started successfully!")
    await dp.start_polling(bot, skip_updates=True)

if __name__ == "__main__":
    asyncio.run(main())

