import logging
import asyncio
import json
import os
import random
import uuid
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Any
import tempfile

from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import (
    ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, 
    InlineKeyboardButton, BufferedInputFile
)
from aiogram.enums import ParseMode
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.client.default import DefaultBotProperties

# Excel library for admin exports only
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
except ImportError:
    print("Required libraries not installed. Install with: pip install openpyxl")

# PDF library for admin reports only
try:
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
except ImportError:
    print("Required libraries not installed. Install with: pip install reportlab")

# 🔑 Configuration
TOKEN = os.getenv("BOT_TOKEN", "7570796885:AAHHfpXanemNYvW-wVT2Rv40U0xq-XjxSwk")
SUPER_ADMIN_ID = int(os.getenv("SUPER_ADMIN_ID", "6578706277"))
SPECIAL_ADMIN_IDS = [6578706277, 7853664401]  # Special privilege admin IDs
CHANNEL_USERNAME = "@Kitobxon_Kids"

# O'zbekiston vaqti (UTC+5)
UZBEKISTAN_TZ = timezone(timedelta(hours=5))

def get_uzbekistan_time():
    """O'zbekiston vaqti bo'yicha hozirgi vaqtni qaytaradi"""
    return datetime.now(UZBEKISTAN_TZ)

# 🛠 Logging
logging.basicConfig(level=logging.INFO)

# 🤖 Bot and Dispatcher
bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher()

# 📁 Data storage files
DATA_DIR = "bot_data"
os.makedirs(DATA_DIR, exist_ok=True)

USERS_FILE = os.path.join(DATA_DIR, "users.json")
ADMINS_FILE = os.path.join(DATA_DIR, "admins.json")
TESTS_FILE = os.path.join(DATA_DIR, "tests.json")
RESULTS_FILE = os.path.join(DATA_DIR, "results.json")
BROADCASTS_FILE = os.path.join(DATA_DIR, "broadcasts.json")
STATISTICS_FILE = os.path.join(DATA_DIR, "statistics.json")

# 📌 Regions data
REGIONS = {
    "Toshkent shahri": ["Bektemir", "Chilonzor", "Mirzo Ulug'bek", "Mirobod", "Olmazor", "Shayxontohur", "Sergeli", "Uchtepa", "Yashnobod", "Yakkasaroy", "Yunusobod"],
    "Toshkent viloyati": ["Bekabad", "Bo'ka", "Bo'stonliq", "Chinoz", "Chirchiq", "Ohangaron", "Oqqo'rg'on", "Parkent", "Piskent", "Quyichirchiq", "O'rtachirchiq", "Yangiyo'l", "Toshkent", "Yuqorichirchiq", "Zangiota", "Nurafshon", "Olmaliq", "Angren"],
    "Andijon": ["Andijon shahri", "Asaka", "Baliqchi", "Bo'ston", "Buloqboshi", "Izboskan", "Jalaquduq", "Marhamat", "Oltinko'l", "Paxtaobod", "Paytug'", "Qo'rg'ontepa", "Shahriston", "Xo'jaobod"],
    "Farg'ona": ["Beshariq", "Buvayda", "Dang'ara", "Farg'ona shahri", "Ferghana tumani", "Furqat", "Qo'qon", "Quva", "Rishton", "So'x", "Toshloq", "Uchko'prik", "Yozyovon", "Oltiariq"],
    "Namangan": ["Chortoq", "Chust", "Kosonsoy", "Namangan shahri", "Norin", "Pop", "To'raqo'rg'on", "Uychi", "Uchqo'rg'on", "Yangiqo'rg'on", "Yangihayot"],
    "Samarqand": ["Bulung'ur", "Ishtixon", "Jomboy", "Kattakurgan", "Oqdaryo", "Payariq", "Pastdarg'om", "Qo'shrabot", "Samarqand shahri", "Toyloq", "Urgut"],
    "Buxoro": ["Buxoro shahri", "Buxoro tumani", "G'ijduvon", "Jondor", "Kogon", "Olot", "Peshku", "Qorako'l", "Qorovulbozor", "Romitan", "Shofirkon", "Vobkent"],
    "Jizzax": ["Baxmal", "Chiroqchi", "Do'stlik", "Forish", "G'allaorol", "Zafarobod", "Zarbdor", "Zomin", "Zafar", "Yangiobod", "Jizzax shahri", "Mirzacho'l"],
    "Navoiy": ["Bespah", "Karmana", "Konimex", "Navbahor", "Nurota", "Tomdi", "Xatirchi", "Uchquduq", "Navoiy shahri", "Zarafshon"],
    "Qashqadaryo": ["Chiroqchi", "G'uzor", "Qarshi", "Kitob", "Koson", "Mirishkor", "Muborak", "Nishon", "Shahrisabz", "Dehqonobod", "Yakkabog'"],
    "Surxondaryo": ["Angor", "Bandixon", "Denov", "Jarqo'rg'on", "Muzrabot", "Oltinsoy", "Sariosiyo", "Sherobod", "Sho'rchi", "Termiz", "Uzun", "Boysun"],
    "Sirdaryo": ["Guliston", "Guliston tumani", "Mirzaobod", "Oqoltin", "Sardoba", "Sayxunobod", "Sirdaryo tumani", "Xovos", "Boyovut", "Yangiyer"],
    "Xorazm": ["Bog'ot", "Gurlan", "Hazorasp", "Khiva", "Qo'shko'pir", "Shovot", "Urganch tumani", "Xonqa", "Yangiariq", "Yangibozor", "Tuproqqal'a", "Urganch shahri"],
    "Qoraqalpog'iston": ["Amudaryo", "Beruniy", "Chimboy", "Ellikqala", "Kegeyli", "Mo'ynoq", "Nukus", "Qanliko'l", "Qo'ng'irot", "Taxiatosh", "To'rtko'l", "Xo'jayli"]
}

# 📌 Data management functions
def load_json_data(file_path: str, default_data: Any = None) -> Any:
    """Load data from JSON file"""
    if not os.path.exists(file_path):
        return default_data or {}
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, FileNotFoundError):
        return default_data or {}

def save_json_data(file_path: str, data: Any) -> None:
    """Save data to JSON file"""
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_users() -> Dict:
    """Get all registered users"""
    return load_json_data(USERS_FILE, {})

def save_user(user_id: str, user_data: Dict) -> None:
    """Save user data"""
    users = get_users()
    users[user_id] = user_data
    save_json_data(USERS_FILE, users)

def get_admins() -> Dict:
    """Get all admins"""
    default_admins = {str(SUPER_ADMIN_ID): {"role": "super_admin", "added_by": "system", "added_date": get_uzbekistan_time().isoformat()}}
    return load_json_data(ADMINS_FILE, default_admins)

def save_admin(admin_id: str, admin_data: Dict) -> None:
    """Save admin data"""
    admins = get_admins()
    admins[admin_id] = admin_data
    save_json_data(ADMINS_FILE, admins)

def remove_admin(admin_id: str) -> bool:
    """Remove admin by ID"""
    try:
        admins = get_admins()
        if admin_id in admins:
            del admins[admin_id]
            save_json_data(ADMINS_FILE, admins)
            return True
        return False
    except Exception as e:
        logging.error(f"Error removing admin data: {e}")
        return False

def get_tests() -> Dict:
    """Get all tests"""
    return load_json_data(TESTS_FILE, {"7-10": {}, "11-14": {}})

def save_test(test_data: Dict) -> None:
    """Save test data"""
    tests = get_tests()
    age_group = test_data["age_group"]
    test_id = str(uuid.uuid4())
    
    if age_group not in tests:
        tests[age_group] = {}
    
    tests[age_group][test_id] = test_data
    save_json_data(TESTS_FILE, tests)

def get_results() -> List:
    """Get all test results"""
    return load_json_data(RESULTS_FILE, [])

def save_result(result_data: Dict) -> None:
    """Save test result"""
    results = get_results()
    results.append(result_data)
    save_json_data(RESULTS_FILE, results)



def get_broadcasts() -> List:
    """Get all broadcast history"""
    return load_json_data(BROADCASTS_FILE, [])

def save_broadcast(broadcast_data: Dict) -> None:
    """Save broadcast data"""
    broadcasts = get_broadcasts()
    broadcasts.append(broadcast_data)
    save_json_data(BROADCASTS_FILE, broadcasts)

def get_statistics() -> Dict:
    """Get statistics data"""
    return load_json_data(STATISTICS_FILE, {})

def update_statistics() -> None:
    """Update comprehensive statistics"""
    users = get_users()
    results = get_results()
    
    # Regional statistics
    regional_stats = {}
    for region in REGIONS.keys():
        regional_stats[region] = {
            "total_users": 0,
            "districts": {}
        }
        
        for district in REGIONS[region]:
            regional_stats[region]["districts"][district] = 0
    
    # Count users by region and district
    for user_data in users.values():
        region = user_data.get("region", "Unknown")
        district = user_data.get("district", "Unknown")
        
        if region in regional_stats:
            regional_stats[region]["total_users"] += 1
            if district in regional_stats[region]["districts"]:
                regional_stats[region]["districts"][district] += 1
    
    # Test statistics
    test_stats = {
        "total_tests_taken": len(results),
        "average_score": 0,
        "high_scorers_70plus": 0,
        "age_group_stats": {
            "7-10": {"count": 0, "avg_score": 0},
            "11-14": {"count": 0, "avg_score": 0}
        }
    }
    
    if results:
        total_score = sum(result.get("score", 0) for result in results)
        test_stats["average_score"] = round(total_score / len(results), 2)
        test_stats["high_scorers_70plus"] = len([r for r in results if r.get("score", 0) >= 70])
        
        # Age group statistics
        age_7_10 = [r for r in results if r.get("age") in ["7", "8", "9", "10"]]
        age_11_14 = [r for r in results if r.get("age") in ["11", "12", "13", "14"]]
        
        if age_7_10:
            test_stats["age_group_stats"]["7-10"]["count"] = len(age_7_10)
            test_stats["age_group_stats"]["7-10"]["avg_score"] = round(
                sum(r.get("score", 0) for r in age_7_10) / len(age_7_10), 2
            )
        
        if age_11_14:
            test_stats["age_group_stats"]["11-14"]["count"] = len(age_11_14)
            test_stats["age_group_stats"]["11-14"]["avg_score"] = round(
                sum(r.get("score", 0) for r in age_11_14) / len(age_11_14), 2
            )
    
    # Top performers ranking
    ranking_data = []
    for result in results:
        ranking_data.append({
            "user_name": result.get("user_name", "Unknown"),
            "score": result.get("score", 0),
            "percentage": result.get("percentage", 0),
            "age": result.get("age", "Unknown"),
            "region": result.get("region", "Unknown"),
            "date": result.get("date", "Unknown")
        })
    
    # Sort by score (descending)
    ranking_data.sort(key=lambda x: x["score"], reverse=True)
    
    statistics = {
        "last_updated": get_uzbekistan_time().isoformat(),
        "total_registered_users": len(users),
        "regional_statistics": regional_stats,
        "test_statistics": test_stats,
        "top_performers": ranking_data[:100]  # Top 100 performers
    }
    
    save_json_data(STATISTICS_FILE, statistics)

def is_admin(user_id: int) -> bool:
    """Check if user is admin"""
    admins = get_admins()
    return str(user_id) in admins

def is_super_admin(user_id: int) -> bool:
    """Check if user is super admin"""
    admins = get_admins()
    return str(user_id) in admins and admins[str(user_id)].get("role") == "super_admin"

def has_special_privileges(user_id: int) -> bool:
    """Check if user has special admin privileges"""
    return user_id in SPECIAL_ADMIN_IDS

# 📌 FSM States
class Registration(StatesGroup):
    check_subscription = State()
    child_name = State()
    parent_name = State()
    region = State()
    district = State()
    mahalla = State()
    age = State()
    phone = State()
    feedback = State()

class AdminStates(StatesGroup):
    add_admin = State()
    remove_admin = State()
    promote_super_admin = State()
    add_test_age = State()
    add_test_book = State()
    add_test_content = State()
    add_test_questions = State()
    delete_test_age = State()
    delete_test_select = State()
    broadcast_message = State()
    broadcast_confirm = State()

    view_statistics = State()
    view_rankings = State()

class TestStates(StatesGroup):
    taking_test = State()
    test_question = State()

# Certificate generation removed as requested

# 📌 Keyboards
def get_main_menu():
    """Main menu keyboard"""
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="📋 Ro'yxatdan o'tish")],
        [KeyboardButton(text="📝 Test topshirish")],
        [KeyboardButton(text="💬 Fikr va maslahatlar")],
        [KeyboardButton(text="📚 Loyiha haqida")]
    ], resize_keyboard=True)

def get_admin_menu(is_super: bool = False):
    """Admin menu keyboard"""
    keyboard = [
        [KeyboardButton(text="👥 Foydalanuvchilar ro'yxati")],
        [KeyboardButton(text="➕ Test qo'shish")]
    ]
    
    if is_super:
        keyboard.extend([
            [KeyboardButton(text="👨‍💼 Adminlar ro'yxati")],
            [KeyboardButton(text="➕ Admin qo'shish")],
            [KeyboardButton(text="⬆️ Super Admin tayinlash")],
            [KeyboardButton(text="➖ Admin o'chirish")],
            [KeyboardButton(text="🗑 Test o'chirish")],
            [KeyboardButton(text="📊 Test natijalarini yuklab olish")],
            [KeyboardButton(text="📋 Foydalanuvchi ma'lumotlarini yuklab olish")],
            [KeyboardButton(text="📢 Xabar yuborish")],
            [KeyboardButton(text="📊 Statistika va monitoring")],
            [KeyboardButton(text="🏆 Reytinglar ro'yxati")]
        ])
    
    keyboard.append([KeyboardButton(text="🔙 Asosiy menyu")])
    return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)

def get_age_group_keyboard():
    """Age group selection keyboard"""
    return ReplyKeyboardMarkup(keyboard=[
        [KeyboardButton(text="7-10 yosh")],
        [KeyboardButton(text="11-14 yosh")],
        [KeyboardButton(text="🔙 Orqaga")]
    ], resize_keyboard=True)

# 📌 PDF and Excel generation functions
def generate_pdf_report(results: List[Dict]) -> bytes:
    """Generate PDF report of test results"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("KITOBXON KIDS - Test Natijalari", styles['Title'])
    
    # Create table data
    table_data = [['Foydalanuvchi', 'Yosh', 'Ball', 'Vaqt', 'Foiz', 'Sana']]
    
    for result in results:
        table_data.append([
            result.get('user_name', 'N/A'),
            result.get('age', 'N/A'),
            f"{result.get('score', 0)}/100",
            result.get('time_taken', 'N/A'),
            f"{result.get('percentage', 0)}%",
            result.get('date', 'N/A')
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    # Build PDF
    doc.build([title, table])
    buffer.seek(0)
    return buffer.getvalue()

def generate_users_pdf_report(users_data: Dict) -> bytes:
    """Generate PDF report of registered users"""
    buffer = BytesIO()
    # Use landscape orientation and larger page size for more space
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("KITOBXON KIDS - Ro'yxatdan o'tgan foydalanuvchilar", styles['Title'])
    
    # Create table data with wrapped text
    table_data = [['Farzand nomi', 'Ota-ona', 'Yosh', 'Viloyat', 'Tuman', 'Telegram ID', 'Username', 'Telefon', 'Ro\'yxat sanasi']]
    
    for user_id, user_data in users_data.items():
        username = user_data.get('username', 'N/A')
        if username != 'N/A' and not username.startswith('@'):
            username = f"@{username}"
        
        # Wrap long text in Paragraphs for better display
        child_name = Paragraph(user_data.get('child_name', 'N/A'), styles['Normal'])
        parent_name = Paragraph(user_data.get('parent_name', 'N/A'), styles['Normal'])
        region = Paragraph(user_data.get('region', 'N/A'), styles['Normal'])
        district = Paragraph(user_data.get('district', 'N/A'), styles['Normal'])
        username_p = Paragraph(username, styles['Normal'])
        phone = Paragraph(user_data.get('phone', 'N/A'), styles['Normal'])
        
        table_data.append([
            child_name,
            parent_name,
            user_data.get('age', 'N/A'),
            region,
            district,
            user_id,
            username_p,
            phone,
            user_data.get('registration_date', 'N/A')[:10] if user_data.get('registration_date') else 'N/A'
        ])
    
    # Create table with specific column widths
    col_widths = [80, 80, 35, 70, 70, 60, 70, 70, 65]  # Adjusted widths
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    # Build PDF
    doc.build([title, table])
    buffer.seek(0)
    return buffer.getvalue()

def generate_excel_report(results: List[Dict]) -> bytes:
    """Generate Excel report of test results"""
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Natijalari"
    
    # Headers
    headers = ['Foydalanuvchi', 'Yosh', 'Ball', 'Vaqt', 'Foiz', 'Sana']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('user_name', 'N/A'))
        ws.cell(row=row, column=2, value=result.get('age', 'N/A'))
        ws.cell(row=row, column=3, value=f"{result.get('score', 0)}/100")
        ws.cell(row=row, column=4, value=result.get('time_taken', 'N/A'))
        ws.cell(row=row, column=5, value=f"{result.get('percentage', 0)}%")
        ws.cell(row=row, column=6, value=result.get('date', 'N/A'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_users_excel_report(users_data: Dict) -> bytes:
    """Generate Excel report of registered users"""
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    
    # Headers
    headers = ['Farzand nomi', 'Ota-ona nomi', 'Yosh', 'Viloyat', 'Tuman', 'Mahalla', 'Telegram ID', 'Username', 'Telefon', 'Ro\'yxat sanasi']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data
    for row, (user_id, user_data) in enumerate(users_data.items(), 2):
        username = user_data.get('username', 'N/A')
        if username != 'N/A' and not username.startswith('@'):
            username = f"@{username}"
        
        ws.cell(row=row, column=1, value=user_data.get('child_name', 'N/A'))
        ws.cell(row=row, column=2, value=user_data.get('parent_name', 'N/A'))
        ws.cell(row=row, column=3, value=user_data.get('age', 'N/A'))
        ws.cell(row=row, column=4, value=user_data.get('region', 'N/A'))
        ws.cell(row=row, column=5, value=user_data.get('district', 'N/A'))
        ws.cell(row=row, column=6, value=user_data.get('mahalla', 'N/A'))
        ws.cell(row=row, column=7, value=user_id)
        ws.cell(row=row, column=8, value=username)
        ws.cell(row=row, column=9, value=user_data.get('phone', 'N/A'))
        ws.cell(row=row, column=10, value=user_data.get('registration_date', 'N/A')[:10] if user_data.get('registration_date') else 'N/A')
        
        # Set text wrapping for all cells in this row
        for col in range(1, 11):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set minimum row height for readability
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 30
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# 🔘 Start command and subscription check
@dp.message(Command("start"))
async def start(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    
    # Check if user is admin
    if is_admin(user_id):
        is_super = is_super_admin(user_id)
        await message.answer(
            f"👋 Salom! Admin panelga xush kelibsiz!\n"
            f"{'🔑 Super Admin' if is_super else '👨‍💼 Admin'} huquqlaringiz mavjud.",
            reply_markup=get_admin_menu(is_super)
        )
        return
    
    try:
        chat_member = await bot.get_chat_member(chat_id=CHANNEL_USERNAME, user_id=user_id)
        if chat_member.status not in ["member", "administrator", "creator"]:
            keyboard = InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="✉️ Obuna bo'lish", url=f"https://t.me/{CHANNEL_USERNAME[1:]}")],
                [InlineKeyboardButton(text="✅ Obuna bo'ldim", callback_data="check_sub")]
            ])
            await message.answer("✉️ Iltimos, quyidagi kanalga obuna bo'ling:", reply_markup=keyboard)
            await state.set_state(Registration.check_subscription)
        else:
            await message.answer("👋 Salom! 'KITOBXON KIDS' botiga xush kelibsiz!", reply_markup=get_main_menu())
    except Exception as e:
        logging.error(f"Error checking subscription: {e}")
        await message.answer("👋 Salom! 'KITOBXON KIDS' botiga xush kelibsiz!", reply_markup=get_main_menu())

@dp.callback_query(lambda c: c.data == "check_sub")
async def check_subscription(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.from_user.id
    try:
        chat_member = await bot.get_chat_member(chat_id=CHANNEL_USERNAME, user_id=user_id)
        if chat_member.status not in ["member", "administrator", "creator"]:
            await callback_query.answer("❌ Hali ham obuna emassiz!", show_alert=True)
        else:
            await bot.send_message(user_id, "👋 Salom! 'KITOBXON KIDS' botiga xush kelibsiz!", reply_markup=get_main_menu())
            await state.clear()
    except Exception as e:
        logging.error(f"Error checking subscription: {e}")
        await bot.send_message(user_id, "👋 Salom! 'KITOBXON KIDS' botiga xush kelibsiz!", reply_markup=get_main_menu())
        await state.clear()

# 📋 Registration process
@dp.message(lambda message: message.text == "📋 Ro'yxatdan o'tish")
async def register_start(message: types.Message, state: FSMContext):
    user_id = str(message.from_user.id)
    users = get_users()
    
    await message.answer("👶 Farzandingiz ism familiyasini kiriting:")
    await state.set_state(Registration.child_name)

@dp.message(Registration.child_name)
async def register_child_name(message: types.Message, state: FSMContext):
    await state.update_data(child_name=message.text)
    await message.answer("👨‍👩‍👦 Ota-onaning ism familiyasini kiriting:")
    await state.set_state(Registration.parent_name)

@dp.message(Registration.parent_name)
async def register_parent_name(message: types.Message, state: FSMContext):
    await state.update_data(parent_name=message.text)
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=region)] for region in REGIONS.keys()],
        resize_keyboard=True
    )
    await message.answer("🌍 Viloyatni tanlang:", reply_markup=keyboard)
    await state.set_state(Registration.region)

@dp.message(Registration.region)
async def register_region(message: types.Message, state: FSMContext):
    region = message.text
    if region not in REGIONS:
        await message.answer("❌ Iltimos, ro'yxatdan viloyat tanlang!")
        return
    
    await state.update_data(region=region)
    districts = REGIONS.get(region, [])
    keyboard = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text=dist)] for dist in districts],
        resize_keyboard=True
    )
    await message.answer("🏙 Tumaningizni tanlang:", reply_markup=keyboard)
    await state.set_state(Registration.district)

@dp.message(Registration.district)
async def register_district(message: types.Message, state: FSMContext):
    await state.update_data(district=message.text)
    await message.answer("🏘 Mahallangiz nomini kiriting:")
    await state.set_state(Registration.mahalla)

@dp.message(Registration.mahalla)
async def register_mahalla(message: types.Message, state: FSMContext):
    await state.update_data(mahalla=message.text)
    age_keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="7"), KeyboardButton(text="8"), KeyboardButton(text="9"), KeyboardButton(text="10")],
            [KeyboardButton(text="11"), KeyboardButton(text="12"), KeyboardButton(text="13"), KeyboardButton(text="14")]
        ],
        resize_keyboard=True
    )
    await message.answer("📅 Yoshni tanlang:", reply_markup=age_keyboard)
    await state.set_state(Registration.age)

@dp.message(Registration.age)
async def register_age(message: types.Message, state: FSMContext):
    try:
        age = int(message.text)
        if age < 7 or age > 14:
            await message.answer("❌ Yosh 7 dan 14 gacha bo'lishi kerak!")
            return
    except ValueError:
        await message.answer("❌ Iltimos, to'g'ri yosh kiriting!")
        return
    
    await state.update_data(age=message.text)
    phone_button = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="📞 Telefon raqamni yuborish", request_contact=True)]],
        resize_keyboard=True
    )
    await message.answer("📞 Telefon raqamingizni yuboring:", reply_markup=phone_button)
    await state.set_state(Registration.phone)

@dp.message(Registration.phone)
async def register_phone(message: types.Message, state: FSMContext):
    if not message.contact:
        await message.answer("📞 Iltimos, tugma orqali telefon raqam yuboring.")
        return

    user_data = await state.get_data()
    phone_number = message.contact.phone_number
    user_id = str(message.from_user.id)
    
    # Save user data
    user_info = {
        'child_name': user_data['child_name'],
        'parent_name': user_data['parent_name'],
        'region': user_data['region'],
        'district': user_data['district'],
        'mahalla': user_data['mahalla'],
        'age': user_data['age'],
        'phone': phone_number,
        'telegram_id': message.from_user.id,
        'username': message.from_user.username or "No username",
        'registration_date': get_uzbekistan_time().isoformat()
    }
    
    save_user(user_id, user_info)

    # Notify all admins
    reg_info = (
        f"📋 Yangi ro'yxatdan o'tish:\n"
        f"👶 Farzand: {user_data['child_name']}\n"
        f"👨‍👩‍👦 Ota-ona: {user_data['parent_name']}\n"
        f"🌍 Viloyat: {user_data['region']}\n"
        f"🏙 Tuman: {user_data['district']}\n"
        f"🏘 Mahalla: {user_data['mahalla']}\n"
        f"📅 Yosh: {user_data['age']}\n"
        f"📞 Telefon: {phone_number}\n"
        f"🆔 Telegram ID: {message.from_user.id}\n"
        f"👤 Username: @{message.from_user.username or 'mavjud emas'}"
    )

    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), reg_info)
        except Exception as e:
            logging.error(f"Error sending notification to admin {admin_id}: {e}")

    await message.answer("✅ Ro'yxatdan o'tish muvaffaqiyatli yakunlandi!", reply_markup=get_main_menu())
    await state.clear()

# 📝 Test system
@dp.message(lambda message: message.text == "📝 Test topshirish")
async def start_test(message: types.Message, state: FSMContext):
    user_id = str(message.from_user.id)
    users = get_users()
    
    if user_id not in users:
        await message.answer("❌ Avval ro'yxatdan o'ting!", reply_markup=get_main_menu())
        return
    
    user_age = int(users[user_id]['age'])
    age_group = "7-10" if user_age <= 10 else "11-14"
    
    tests = get_tests()
    if not tests.get(age_group) or not any(tests[age_group].values()):
        await message.answer("❌ Sizning yosh guruhingiz uchun hozircha testlar mavjud emas.")
        return
    
    # Collect all questions from all tests in the age group
    all_questions = []
    for test_data in tests[age_group].values():
        if 'questions' in test_data:
            all_questions.extend(test_data['questions'])
    
    if len(all_questions) < 25:
        await message.answer("❌ Yetarli savollar mavjud emas. Test uchun kamida 25 ta savol kerak.")
        return
    
    # Select 25 random questions
    selected_questions = random.sample(all_questions, 25)
    
    await state.update_data(
        questions=selected_questions,
        current_question=0,
        score=0,
        start_time=get_uzbekistan_time().isoformat(),
        age_group=age_group,
        user_answers=[]
    )
    
    await message.answer(
        f"📝 Test boshlandi!\n"
        f"👶 Yosh guruhi: {age_group}\n"
        f"❓ Savollar soni: 25\n"
        f"⏱ Har bir savol uchun: 1 daqiqa\n"
        f"💯 Har bir to'g'ri javob: 4 ball\n\n"
        "Tayyor bo'lsangiz, birinchi savolga o'ting:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="▶️ Boshlash")]],
            resize_keyboard=True
        )
    )
    await state.set_state(TestStates.taking_test)

@dp.message(TestStates.taking_test)
async def show_question(message: types.Message, state: FSMContext):
    if message.text != "▶️ Boshlash" and message.text != "➡️ Keyingisi":
        return
    
    data = await state.get_data()
    questions = data['questions']
    current_q = data['current_question']
    
    if current_q >= len(questions):
        await finish_test(message, state)
        return
    
    question = questions[current_q]
    
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="A"), KeyboardButton(text="B")],
            [KeyboardButton(text="C"), KeyboardButton(text="D")]
        ],
        resize_keyboard=True
    )
    
    question_text = (
        f"❓ Savol {current_q + 1}/25\n\n"
        f"{question['question']}\n\n"
        f"A) {question['option_a']}\n"
        f"B) {question['option_b']}\n"
        f"C) {question['option_c']}\n"
        f"D) {question['option_d']}\n\n"
        "⏱ 1 daqiqa vaqtingiz bor!"
    )
    
    await message.answer(question_text, reply_markup=keyboard)
    await state.set_state(TestStates.test_question)
    
    # Set timer for 60 seconds
    await asyncio.sleep(60)
    
    # Check if user is still on this question
    current_data = await state.get_data()
    if current_data.get('current_question') == current_q:
        await handle_timeout(message, state)

async def handle_timeout(message: types.Message, state: FSMContext):
    """Handle question timeout"""
    data = await state.get_data()
    current_q = data['current_question']
    user_answers = data.get('user_answers', [])
    
    # Add timeout answer
    user_answers.append({
        'question_number': current_q + 1,
        'user_answer': 'TIMEOUT',
        'correct_answer': data['questions'][current_q]['correct_answer'],
        'is_correct': False
    })
    
    await state.update_data(
        current_question=current_q + 1,
        user_answers=user_answers
    )
    
    if current_q + 1 >= 25:
        await finish_test(message, state)
    else:
        await message.answer(
            "⏰ Vaqt tugadi! Keyingi savolga o'tamiz...",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="➡️ Keyingisi")]],
                resize_keyboard=True
            )
        )
        await state.set_state(TestStates.taking_test)

@dp.message(TestStates.test_question)
async def process_answer(message: types.Message, state: FSMContext):
    user_answer = message.text.upper()
    if user_answer not in ['A', 'B', 'C', 'D']:
        await message.answer("❌ Iltimos, A, B, C yoki D harfini tanlang!")
        return
    
    data = await state.get_data()
    current_q = data['current_question']
    questions = data['questions']
    score = data['score']
    user_answers = data.get('user_answers', [])
    
    current_question = questions[current_q]
    correct_answer = current_question['correct_answer'].upper()
    is_correct = user_answer == correct_answer
    
    if is_correct:
        score += 4
        await message.answer("✅ To'g'ri!")
    else:
        await message.answer(f"❌ Noto'g'ri! To'g'ri javob: {correct_answer}")
    
    # Record user answer
    user_answers.append({
        'question_number': current_q + 1,
        'user_answer': user_answer,
        'correct_answer': correct_answer,
        'is_correct': is_correct
    })
    
    current_q += 1
    
    await state.update_data(
        current_question=current_q,
        score=score,
        user_answers=user_answers
    )
    
    if current_q >= 25:
        await finish_test(message, state)
    else:
        await message.answer(
            f"Keyingi savol uchun tayyor bo'ling... ({current_q}/25)",
            reply_markup=ReplyKeyboardMarkup(
                keyboard=[[KeyboardButton(text="➡️ Keyingisi")]],
                resize_keyboard=True
            )
        )
        await state.set_state(TestStates.taking_test)

async def finish_test(message: types.Message, state: FSMContext):
    """Finish the test and show results"""
    data = await state.get_data()
    score = data['score']
    start_time = datetime.fromisoformat(data['start_time']).replace(tzinfo=UZBEKISTAN_TZ)
    end_time = get_uzbekistan_time()
    time_taken = str(end_time - start_time).split('.')[0]  # Remove microseconds
    percentage = (score / 100) * 100
    
    user_id = str(message.from_user.id)
    users = get_users()
    user_data = users.get(user_id, {})
    child_name = user_data.get('child_name', 'Unknown')
    child_last_name = user_data.get('child_last_name', '')
    full_child_name = f"{child_name} {child_last_name}".strip() if child_last_name else child_name
    
    # Save result with regional information
    result_data = {
        'user_id': user_id,
        'user_name': child_name,
        'age': user_data.get('age', 'Unknown'),
        'region': user_data.get('region', 'Unknown'),
        'district': user_data.get('district', 'Unknown'),
        'score': score,
        'percentage': percentage,
        'time_taken': time_taken,
        'date': end_time.strftime('%Y-%m-%d %H:%M:%S'),
        'age_group': data['age_group'],
        'answers': data.get('user_answers', [])
    }
    
    save_result(result_data)
    
    # Update statistics after saving result
    update_statistics()
    
    # Certificate generation removed as requested
    
    # Send results
    result_text = (
        f"🏁 Test yakunlandi!\n\n"
        f"👶 Farzand: {full_child_name}\n"
        f"💯 Ball: {score}/100\n"
        f"📊 Foiz: {percentage:.1f}%\n"
        f"⏱ Vaqt: {time_taken}\n"
        f"📅 Sana: {end_time.strftime('%Y-%m-%d %H:%M')}\n\n"
    )
    
    if score >= 70:
        result_text += "🏆 Ajoyib! Yaxshi natija ko'rsatdingiz!\n\n"
    
    result_text += "Rahmat! Qatnashganingiz uchun!"
    
    await message.answer(result_text, reply_markup=get_main_menu())
    
    # Notify all admins
    admin_notification = (
        f"📊 Yangi test natijasi:\n"
        f"👶 Farzand: {child_name}\n"
        f"👤 Username: @{message.from_user.username or 'mavjud emas'}\n"
        f"🆔 Telegram ID: {user_id}\n"
        f"💯 Ball: {score}/100\n"
        f"📊 Foiz: {percentage:.1f}%\n"
        f"⏱ Vaqt: {time_taken}\n"
        f"📅 Sana: {end_time.strftime('%Y-%m-%d %H:%M')}\n"
        f"👶 Yosh guruhi: {data['age_group']}\n"
    )
    
    if score >= 70:
        admin_notification += "🏆 Yaxshi natija (70+ ball)\n"
    
    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), admin_notification)
        except Exception as e:
            logging.error(f"Error sending test result to admin {admin_id}: {e}")
    
    await state.clear()

# 💬 Feedback system
@dp.message(lambda message: message.text == "💬 Fikr va maslahatlar")
async def feedback_prompt(message: types.Message, state: FSMContext):
    await message.answer("✍️ Fikringizni yozing:")
    await state.set_state(Registration.feedback)

@dp.message(Registration.feedback)
async def save_feedback(message: types.Message, state: FSMContext):
    feedback_text = (
        f"💬 Yangi fikr:\n"
        f"👤 {message.from_user.full_name}\n"
        f"👤 Username: @{message.from_user.username or 'mavjud emas'}\n"
        f"🆔 Telegram ID: {message.from_user.id}\n"
        f"💭 Fikr: {message.text}"
    )
    
    # Send to all admins
    admins = get_admins()
    for admin_id in admins.keys():
        try:
            await bot.send_message(int(admin_id), feedback_text)
        except Exception as e:
            logging.error(f"Error sending feedback to admin {admin_id}: {e}")
    
    await message.answer("✅ Fikringiz uchun rahmat!", reply_markup=get_main_menu())
    await state.clear()

# 📚 Project info
@dp.message(lambda message: message.text == "📚 Loyiha haqida")
async def project_info(message: types.Message):
    text = """<b>"Kitobxon kids" tanlovini tashkil etish va o'tkazish to'g'risidagi NIZOM</b>

🔹 <b>Umumiy qoidalar:</b>
• Mazkur Nizom yoshlar o'rtasida "Kitobxon Kids" tanlovini o'tkazish tartibini belgilaydi.
• Tanlov 7–10 va 11–14 yoshdagi bolalar uchun mo'ljallangan.
• Tanlov kitobxonlik madaniyatini oshirishga qaratilgan.

🔹 <b>Tashkilotchilar:</b>
• Yoshlar ishlari agentligi,
• Maktabgacha va maktab ta'limi vazirligi,
• O'zbekiston bolalar tashkiloti.

🔹 <b>Ishtirokchilar:</b>
• 7–14 yoshdagi barcha bolalar qatnasha oladi.
• Qoraqalpoq va rus tillarida ham qatnashish mumkin.

🔹 <b>Maqsad va vazifalar:</b>
• Kitob o'qishga qiziqish uyg'otish, mustaqil o'qish ko'nikmasini shakllantirish.
• Adiblar merosini o'rganish, o'zlikni anglashga chorlash.

🔹 <b>Tanlov bosqichlari:</b>
1. Saralash (oy boshida test, 25 ta savol, har biri 4 ball).
2. Hududiy (30 ta savol, har biri 30 soniya, top scorer keyingi bosqichga o'tadi).
3. Respublika (Fantaziya festivali, Taassurotlar, Savollar - 100 ballik tizim).

🔹 <b>G'oliblar:</b>
• 1-o'rin: Noutbuk
• 2-o'rin: Planshet
• 3-o'rin: Telefon
• Barcha qatnashchilarga velosiped

🔹 <b>Moliya manbalari:</b>
• Agentlik mablag'lari, homiylar, qonuniy xayriyalar.

Batafsil: @Kitobxon_Kids kanali orqali kuzatib boring.
"""
    await message.answer(text)

# 🔙 Back to main menu
@dp.message(lambda message: message.text == "🔙 Asosiy menyu")
async def back_to_main(message: types.Message, state: FSMContext):
    await state.clear()
    if is_admin(message.from_user.id):
        is_super = is_super_admin(message.from_user.id)
        await message.answer("Admin panel", reply_markup=get_admin_menu(is_super))
    else:
        await message.answer("Asosiy menyu", reply_markup=get_main_menu())

# 👥 View users (Admin only)
@dp.message(lambda message: message.text == "👥 Foydalanuvchilar ro'yxati" and is_admin(message.from_user.id))
async def view_users(message: types.Message):
    users = get_users()
    
    if not users:
        await message.answer("📋 Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q.")
        return
    
    user_list = "👥 Ro'yxatdan o'tgan foydalanuvchilar:\n\n"
    for i, (user_id, user_data) in enumerate(users.items(), 1):
        username = user_data.get('username', 'N/A')
        if username != 'N/A' and not username.startswith('@'):
            username = f"@{username}"
        
        user_list += (
            f"{i}. 👶 {user_data.get('child_name', 'N/A')}\n"
            f"   👨‍👩‍👦 {user_data.get('parent_name', 'N/A')}\n"
            f"   📅 {user_data.get('age', 'N/A')} yosh\n"
            f"   🌍 {user_data.get('region', 'N/A')}, {user_data.get('district', 'N/A')}\n"
            f"   🆔 {user_id}\n"
            f"   👤 {username}\n\n"
        )
        
        if len(user_list) > 3500:  # Telegram message limit
            await message.answer(user_list)
            user_list = ""
    
    if user_list:
        await message.answer(user_list)

# 📊 Download test results (Super Admin only)
@dp.message(lambda message: message.text == "📊 Test natijalarini yuklab olish" and is_super_admin(message.from_user.id))
async def download_test_results(message: types.Message):
    results = get_results()
    
    if not results:
        await message.answer("📊 Hozircha test natijalari mavjud emas.")
        return
    
    try:
        # Generate PDF
        pdf_data = generate_pdf_report(results)
        pdf_file = BufferedInputFile(pdf_data, filename="test_natijalari.pdf")
        
        # Generate Excel
        excel_data = generate_excel_report(results)
        excel_file = BufferedInputFile(excel_data, filename="test_natijalari.xlsx")
        
        await message.answer("📊 Test natijalari:")
        await message.answer_document(pdf_file, caption="📄 PDF format")
        await message.answer_document(excel_file, caption="📊 Excel format")
        
    except Exception as e:
        logging.error(f"Error generating test results: {e}")
        await message.answer("❌ Hisobot yaratishda xatolik yuz berdi.")

# 📋 Download user data (Super Admin only)
@dp.message(lambda message: message.text == "📋 Foydalanuvchi ma'lumotlarini yuklab olish" and is_super_admin(message.from_user.id))
async def download_user_data(message: types.Message):
    users = get_users()
    
    if not users:
        await message.answer("📋 Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q.")
        return
    
    try:
        # Generate PDF
        pdf_data = generate_users_pdf_report(users)
        pdf_file = BufferedInputFile(pdf_data, filename="foydalanuvchilar.pdf")
        
        # Generate Excel
        excel_data = generate_users_excel_report(users)
        excel_file = BufferedInputFile(excel_data, filename="foydalanuvchilar.xlsx")
        
        await message.answer("📋 Foydalanuvchilar ma'lumoti:")
        await message.answer_document(pdf_file, caption="📄 PDF format")
        await message.answer_document(excel_file, caption="📊 Excel format")
        
    except Exception as e:
        logging.error(f"Error generating user data: {e}")
        await message.answer("❌ Hisobot yaratishda xatolik yuz berdi.")

# 👨‍💼 View admins (Super Admin only)
@dp.message(lambda message: message.text == "👨‍💼 Adminlar ro'yxati" and is_super_admin(message.from_user.id))
async def view_admins(message: types.Message):
    admins = get_admins()
    
    admin_list = "👨‍💼 Adminlar ro'yxati:\n\n"
    for i, (admin_id, admin_data) in enumerate(admins.items(), 1):
        role = "🔑 Super Admin" if admin_data.get("role") == "super_admin" else "👨‍💼 Admin"
        
        try:
            admin_info = await bot.get_chat(int(admin_id))
            full_name = admin_info.full_name or "N/A"
            username = f"@{admin_info.username}" if admin_info.username else "N/A"
        except:
            full_name = "N/A"
            username = "N/A"
        
        admin_list += (
            f"{i}. {role}\n"
            f"   👤 {full_name}\n"
            f"   👤 {username}\n"
            f"   🆔 {admin_id}\n"
            f"   📅 {admin_data.get('added_date', 'N/A')[:10]}\n\n"
        )
    
    await message.answer(admin_list)

# ➕ Add admin (Super Admin only)
@dp.message(lambda message: message.text == "➕ Admin qo'shish" and is_super_admin(message.from_user.id))
async def add_admin_start(message: types.Message, state: FSMContext):
    await message.answer("🆔 Yangi admin Telegram ID sini kiriting:")
    await state.set_state(AdminStates.add_admin)

@dp.message(AdminStates.add_admin)
async def add_admin_process(message: types.Message, state: FSMContext):
    try:
        admin_id = int(message.text)
        
        # Check if already admin
        admins = get_admins()
        if str(admin_id) in admins:
            await message.answer("❌ Bu foydalanuvchi allaqachon admin!")
            await state.clear()
            return
        
        # Add admin
        admin_data = {
            "role": "admin",
            "added_by": str(message.from_user.id),
            "added_date": get_uzbekistan_time().isoformat()
        }
        
        save_admin(str(admin_id), admin_data)
        
        try:
            admin_info = await bot.get_chat(admin_id)
            admin_name = admin_info.full_name or "N/A"
            username = f"@{admin_info.username}" if admin_info.username else "N/A"
        except:
            admin_name = "N/A"
            username = "N/A"
        
        await message.answer(f"✅ Yangi admin qo'shildi!\n👤 {admin_name}\n🆔 {admin_id}")
        
        # Notify new admin
        try:
            await bot.send_message(
                admin_id,
                "🎉 Tabriklaymiz! Siz Kitobxon Kids botida admin etib tayinlandingiz!\n"
                "Botni qayta ishga tushiring: /start"
            )
        except:
            pass
        
        # Notify all admins
        notification = (
            f"👨‍💼 Yangi admin qo'shildi:\n"
            f"👤 {admin_name}\n"
            f"👤 {username}\n"
            f"🆔 {admin_id}\n"
            f"➕ Qo'shgan: {message.from_user.full_name}"
        )
        
        for aid in admins.keys():
            if aid != str(message.from_user.id):
                try:
                    await bot.send_message(int(aid), notification)
                except:
                    pass
        
    except ValueError:
        await message.answer("❌ Noto'g'ri ID format! Raqam kiriting.")
        return
    except Exception as e:
        logging.error(f"Error adding admin: {e}")
        await message.answer("❌ Admin qo'shishda xatolik yuz berdi.")
    
    await state.clear()

# ⬆️ Promote to Super Admin (Super Admin only)
@dp.message(lambda message: message.text == "⬆️ Super Admin tayinlash" and is_super_admin(message.from_user.id))
async def promote_super_admin_start(message: types.Message, state: FSMContext):
    await message.answer("🆔 Super Admin etib tayinlash uchun Telegram ID kiriting:")
    await state.set_state(AdminStates.promote_super_admin)

@dp.message(AdminStates.promote_super_admin)
async def promote_super_admin_process(message: types.Message, state: FSMContext):
    try:
        admin_id = str(message.text)
        
        admins = get_admins()
        if admin_id not in admins:
            await message.answer("❌ Bu foydalanuvchi admin emas!")
            await state.clear()
            return
        
        if admins[admin_id].get("role") == "super_admin":
            await message.answer("❌ Bu foydalanuvchi allaqachon Super Admin!")
            await state.clear()
            return
        
        # Promote to super admin
        admins[admin_id]["role"] = "super_admin"
        admins[admin_id]["promoted_by"] = str(message.from_user.id)
        admins[admin_id]["promoted_date"] = get_uzbekistan_time().isoformat()
        
        save_admin(admin_id, admins[admin_id])
        
        try:
            admin_info = await bot.get_chat(int(admin_id))
            admin_name = admin_info.full_name or "N/A"
        except:
            admin_name = "N/A"
        
        await message.answer(f"✅ {admin_name} Super Admin etib tayinlandi!")
        
        # Notify promoted admin
        try:
            await bot.send_message(
                int(admin_id),
                "🔑 Tabriklaymiz! Siz Super Admin etib tayinlandingiz!\n"
                "Botni qayta ishga tushiring: /start"
            )
        except:
            pass
        
    except ValueError:
        await message.answer("❌ Noto'g'ri ID format! Raqam kiriting.")
    except Exception as e:
        logging.error(f"Error promoting admin: {e}")
        await message.answer("❌ Super Admin tayinlashda xatolik yuz berdi.")
    
    await state.clear()

# ➖ Remove admin (Super Admin or Special Admin only)
@dp.message(lambda message: message.text == "➖ Admin o'chirish" and (is_super_admin(message.from_user.id) or has_special_privileges(message.from_user.id)))
async def remove_admin_start(message: types.Message, state: FSMContext):
    await message.answer("🆔 O'chirish uchun admin Telegram ID sini kiriting:")
    await state.set_state(AdminStates.remove_admin)

@dp.message(AdminStates.remove_admin)
async def remove_admin_process(message: types.Message, state: FSMContext):
    try:
        admin_id = str(message.text)
        current_user = message.from_user.id
        
        # Special privilege check
        if not (is_super_admin(current_user) or has_special_privileges(current_user)):
            await message.answer("❌ Sizda bu amalni bajarish huquqi yo'q!")
            await state.clear()
            return
        
        # Can't remove yourself
        if admin_id == str(current_user):
            await message.answer("❌ O'zingizni admin ro'yxatidan o'chira olmaysiz!")
            await state.clear()
            return
        
        # Can't remove system super admin
        if admin_id == str(SUPER_ADMIN_ID) and current_user != SUPER_ADMIN_ID:
            await message.answer("❌ Asosiy super adminni o'chira olmaysiz!")
            await state.clear()
            return
        
        if remove_admin(admin_id):
            try:
                admin_info = await bot.get_chat(int(admin_id))
                admin_name = admin_info.full_name or "N/A"
            except:
                admin_name = "N/A"
            
            await message.answer(f"✅ Admin o'chirildi: {admin_name} ({admin_id})")
            
            # Notify removed admin
            try:
                await bot.send_message(
                    int(admin_id),
                    "❌ Sizning admin huquqlaringiz olib tashlandi!"
                )
            except:
                pass
            
        else:
            await message.answer("❌ Admin topilmadi yoki o'chirishda xatolik!")
        
    except ValueError:
        await message.answer("❌ Noto'g'ri ID format! Raqam kiriting.")
    except Exception as e:
        logging.error(f"Error removing admin: {e}")
        await message.answer("❌ Admin o'chirishda xatolik yuz berdi.")
    
    await state.clear()

# ➕ Add test (Admin only)
@dp.message(lambda message: message.text == "➕ Test qo'shish" and is_admin(message.from_user.id))
async def add_test_start(message: types.Message, state: FSMContext):
    await message.answer("👶 Qaysi yosh guruhi uchun test qo'shasiz?", reply_markup=get_age_group_keyboard())
    await state.set_state(AdminStates.add_test_age)

@dp.message(AdminStates.add_test_age)
async def add_test_age(message: types.Message, state: FSMContext):
    if message.text == "🔙 Orqaga":
        is_super = is_super_admin(message.from_user.id)
        await message.answer("Admin panel", reply_markup=get_admin_menu(is_super))
        await state.clear()
        return
    
    if message.text not in ["7-10 yosh", "11-14 yosh"]:
        await message.answer("❌ Iltimos, to'g'ri yosh guruhini tanlang!")
        return
    
    age_group = "7-10" if message.text == "7-10 yosh" else "11-14"
    await state.update_data(age_group=age_group)
    
    await message.answer("📚 Kitob nomini kiriting:")
    await state.set_state(AdminStates.add_test_book)

@dp.message(AdminStates.add_test_book)
async def add_test_book(message: types.Message, state: FSMContext):
    await state.update_data(book_name=message.text)
    
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📝 Matn")],
            [KeyboardButton(text="📄 PDF")],
            [KeyboardButton(text="🔙 Orqaga")]
        ],
        resize_keyboard=True
    )
    
    await message.answer("📄 Test turini tanlang:", reply_markup=keyboard)
    await state.set_state(AdminStates.add_test_content)

@dp.message(AdminStates.add_test_content)
async def add_test_content(message: types.Message, state: FSMContext):
    if message.text == "🔙 Orqaga":
        await message.answer("👶 Qaysi yosh guruhi uchun test qo'shasiz?", reply_markup=get_age_group_keyboard())
        await state.set_state(AdminStates.add_test_age)
        return
    
    if message.text not in ["📝 Matn", "📄 PDF"]:
        await message.answer("❌ Iltimos, to'g'ri formatni tanlang!")
        return
    
    content_type = "text" if message.text == "📝 Matn" else "pdf"
    await state.update_data(content_type=content_type)
    
    if content_type == "text":
        await message.answer(
            "📝 Savollarni quyidagi formatda kiriting:\n\n"
            "Savol matni?\n"
            "A) Variant A\n"
            "B) Variant B\n"
            "C) Variant C\n"
            "D) Variant D\n"
            "Javob: B\n\n"
            "Har bir savolni alohida xabar sifatida yuboring. Tugatish uchun 'TUGADI' yozing."
        )
    else:
        await message.answer("📄 PDF faylni yuboring:")
    
    await state.set_state(AdminStates.add_test_questions)

@dp.message(AdminStates.add_test_questions)
async def add_test_questions(message: types.Message, state: FSMContext):
    data = await state.get_data()
    content_type = data['content_type']
    
    if content_type == "pdf":
        if not message.document or message.document.mime_type != 'application/pdf':
            await message.answer("❌ Iltimos, PDF fayl yuboring!")
            return
        
        # For PDF, we'll create a placeholder test since PDF parsing is complex
        # In a real implementation, you'd use PyPDF2 or similar to extract text
        await message.answer("📄 PDF qabul qilindi. PDF dan savollarni ajratib olish...")
        
        # Placeholder questions (in real app, extract from PDF)
        questions = [{
            "question": f"PDF dan ajratilgan savol (kitob: {data['book_name']})",
            "option_a": "Variant A",
            "option_b": "Variant B", 
            "option_c": "Variant C",
            "option_d": "Variant D",
            "correct_answer": "B"
        }]
        
        test_data = {
            "book_name": data['book_name'],
            "age_group": data['age_group'],
            "questions": questions,
            "added_by": str(message.from_user.id),
            "added_date": get_uzbekistan_time().isoformat(),
            "content_type": "pdf"
        }
        
        save_test(test_data)
        
        is_super = is_super_admin(message.from_user.id)
        await message.answer("✅ PDF test muvaffaqiyatli qo'shildi!", reply_markup=get_admin_menu(is_super))
        await state.clear()
        
    else:  # text
        if message.text == "TUGADI":
            questions = data.get('questions', [])
            if not questions:
                await message.answer("❌ Hech qanday savol qo'shilmadi!")
                return
            
            test_data = {
                "book_name": data['book_name'],
                "age_group": data['age_group'],
                "questions": questions,
                "added_by": str(message.from_user.id),
                "added_date": get_uzbekistan_time().isoformat(),
                "content_type": "text"
            }
            
            save_test(test_data)
            
            is_super = is_super_admin(message.from_user.id)
            await message.answer(f"✅ Test muvaffaqiyatli qo'shildi! Jami {len(questions)} ta savol.", reply_markup=get_admin_menu(is_super))
            await state.clear()
            return
        
        # Parse question
        try:
            lines = message.text.strip().split('\n')
            if len(lines) < 6:
                await message.answer("❌ Noto'g'ri format! Qaytadan kiriting.")
                return
            
            question_text = lines[0]
            options = {}
            correct_answer = None
            
            for line in lines[1:]:
                line = line.strip()
                if line.startswith('A)'):
                    options['option_a'] = line[2:].strip()
                elif line.startswith('B)'):
                    options['option_b'] = line[2:].strip()
                elif line.startswith('C)'):
                    options['option_c'] = line[2:].strip()
                elif line.startswith('D)'):
                    options['option_d'] = line[2:].strip()
                elif line.startswith('Javob:'):
                    correct_answer = line.split(':')[1].strip().upper()
            
            if not all([question_text, options.get('option_a'), options.get('option_b'), 
                       options.get('option_c'), options.get('option_d'), correct_answer]):
                await message.answer("❌ Noto'g'ri format! Barcha maydonlarni to'ldiring.")
                return
            
            if correct_answer not in ['A', 'B', 'C', 'D']:
                await message.answer("❌ To'g'ri javob A, B, C yoki D bo'lishi kerak!")
                return
            
            question_data = {
                "question": question_text,
                "option_a": options['option_a'],
                "option_b": options['option_b'],
                "option_c": options['option_c'],
                "option_d": options['option_d'],
                "correct_answer": correct_answer
            }
            
            current_questions = data.get('questions', [])
            current_questions.append(question_data)
            await state.update_data(questions=current_questions)
            
            await message.answer(f"✅ Savol qo'shildi! Jami: {len(current_questions)}\nKeyingi savolni kiriting yoki 'TUGADI' yozing.")
            
        except Exception as e:
            logging.error(f"Error parsing question: {e}")
            await message.answer("❌ Savolni tahlil qilishda xatolik! Qaytadan kiriting.")

# 🗑 Delete test (Super Admin only)
@dp.message(lambda message: message.text == "🗑 Test o'chirish" and is_super_admin(message.from_user.id))
async def delete_test_start(message: types.Message, state: FSMContext):
    await message.answer("👶 Qaysi yosh guruhidan test o'chirasiz?", reply_markup=get_age_group_keyboard())
    await state.set_state(AdminStates.delete_test_age)

@dp.message(AdminStates.delete_test_age)
async def delete_test_age(message: types.Message, state: FSMContext):
    if message.text == "🔙 Orqaga":
        await message.answer("Admin panel", reply_markup=get_admin_menu(True))
        await state.clear()
        return
    
    if message.text not in ["7-10 yosh", "11-14 yosh"]:
        await message.answer("❌ Iltimos, to'g'ri yosh guruhini tanlang!")
        return
    
    age_group = "7-10" if message.text == "7-10 yosh" else "11-14"
    tests = get_tests()
    
    if not tests.get(age_group) or not tests[age_group]:
        await message.answer("❌ Bu yosh guruhida testlar mavjud emas!")
        await state.clear()
        return
    
    await state.update_data(age_group=age_group)
    
    # Show available tests
    test_list = f"📚 {age_group} yosh guruhi testlari:\n\n"
    keyboard_buttons = []
    
    for i, (test_id, test_data) in enumerate(tests[age_group].items(), 1):
        book_name = test_data.get('book_name', 'N/A')
        question_count = len(test_data.get('questions', []))
        test_list += f"{i}. {book_name} ({question_count} ta savol)\n"
        keyboard_buttons.append([KeyboardButton(text=f"{i}. {book_name}")])
    
    keyboard_buttons.append([KeyboardButton(text="🔙 Orqaga")])
    keyboard = ReplyKeyboardMarkup(keyboard=keyboard_buttons, resize_keyboard=True)
    
    await message.answer(test_list + "\nO'chirish uchun testni tanlang:", reply_markup=keyboard)
    await state.set_state(AdminStates.delete_test_select)

@dp.message(AdminStates.delete_test_select)
async def delete_test_select(message: types.Message, state: FSMContext):
    if message.text == "🔙 Orqaga":
        await message.answer("👶 Qaysi yosh guruhidan test o'chirasiz?", reply_markup=get_age_group_keyboard())
        await state.set_state(AdminStates.delete_test_age)
        return
    
    data = await state.get_data()
    age_group = data['age_group']
    tests = get_tests()
    
    # Find selected test
    selected_test_id = None
    for test_id, test_data in tests[age_group].items():
        book_name = test_data.get('book_name', 'N/A')
        if book_name in message.text:
            selected_test_id = test_id
            break
    
    if not selected_test_id:
        await message.answer("❌ Test topilmadi! Qaytadan tanlang.")
        return
    
    # Delete test
    del tests[age_group][selected_test_id]
    save_json_data(TESTS_FILE, tests)
    
    await message.answer("✅ Test muvaffaqiyatli o'chirildi!", reply_markup=get_admin_menu(True))
    await state.clear()

# 📢 Broadcast message (Super Admin only)
@dp.message(lambda message: message.text == "📢 Xabar yuborish" and is_super_admin(message.from_user.id))
async def broadcast_start(message: types.Message, state: FSMContext):
    await message.answer(
        "📢 Barcha ro'yxatdan o'tgan foydalanuvchilarga yubormoqchi bo'lgan xabaringizni yozing:\n\n"
        "❗️ Xabar barcha foydalanuvchilarga yuboriladi!"
    )
    await state.set_state(AdminStates.broadcast_message)

@dp.message(AdminStates.broadcast_message)
async def broadcast_message(message: types.Message, state: FSMContext):
    broadcast_text = message.text
    await state.update_data(broadcast_text=broadcast_text)
    
    # Show preview
    users = get_users()
    user_count = len(users)
    
    preview_text = (
        f"📢 Xabar ko'rinishi:\n\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"{broadcast_text}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n\n"
        f"👥 Jami {user_count} ta foydalanuvchiga yuboriladi.\n\n"
        f"Tasdiqlaysizmi?"
    )
    
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="✅ Ha, yuborish")],
            [KeyboardButton(text="❌ Yo'q, bekor qilish")]
        ],
        resize_keyboard=True
    )
    
    await message.answer(preview_text, reply_markup=keyboard)
    await state.set_state(AdminStates.broadcast_confirm)

@dp.message(AdminStates.broadcast_confirm)
async def broadcast_confirm(message: types.Message, state: FSMContext):
    if message.text == "❌ Yo'q, bekor qilish":
        await message.answer("❌ Xabar yuborish bekor qilindi.", reply_markup=get_admin_menu(True))
        await state.clear()
        return
    
    if message.text != "✅ Ha, yuborish":
        await message.answer("❌ Iltimos, tugmalardan birini tanlang!")
        return
    
    data = await state.get_data()
    broadcast_text = data['broadcast_text']
    users = get_users()
    
    if not users:
        await message.answer("❌ Xabar yuborish uchun foydalanuvchilar mavjud emas!")
        await state.clear()
        return
    
    # Start broadcasting
    await message.answer("📤 Xabar yuborish boshlandi...", reply_markup=get_admin_menu(True))
    
    success_count = 0
    failed_count = 0
    
    for user_id in users.keys():
        try:
            await bot.send_message(int(user_id), f"📢 E'lon:\n\n{broadcast_text}")
            success_count += 1
            await asyncio.sleep(0.05)  # Small delay to avoid rate limits
        except Exception as e:
            failed_count += 1
            logging.error(f"Failed to send broadcast to {user_id}: {e}")
    
    # Save broadcast history
    broadcast_data = {
        'admin_id': str(message.from_user.id),
        'admin_name': message.from_user.full_name,
        'message': broadcast_text,
        'date': get_uzbekistan_time().isoformat(),
        'success_count': success_count,
        'failed_count': failed_count,
        'total_users': len(users)
    }
    save_broadcast(broadcast_data)
    
    result_text = (
        f"📊 Xabar yuborish yakunlandi!\n\n"
        f"✅ Muvaffaqiyatli: {success_count}\n"
        f"❌ Xatolik: {failed_count}\n"
        f"👥 Jami: {len(users)} foydalanuvchi"
    )
    
    await message.answer(result_text)
    await state.clear()

# Certificate viewing removed as requested

# 📊 Statistics and Monitoring (Super Admin only)
@dp.message(lambda message: message.text == "📊 Statistika va monitoring" and is_super_admin(message.from_user.id))
async def view_statistics(message: types.Message):
    # Update statistics first
    update_statistics()
    stats = get_statistics()
    
    if not stats:
        await message.answer("📊 Statistika ma'lumotlari mavjud emas.")
        return
    
    # General statistics
    general_text = (
        f"📊 <b>KITOBXON KIDS - UMUMIY STATISTIKA</b>\n\n"
        f"👥 Jami ro'yxatdan o'tganlar: <b>{stats.get('total_registered_users', 0)}</b>\n"
        f"📝 Jami testlar topshirildi: <b>{stats.get('test_statistics', {}).get('total_tests_taken', 0)}</b>\n"
        f"📈 O'rtacha ball: <b>{stats.get('test_statistics', {}).get('average_score', 0)}</b>\n"
        f"🏆 70%+ ball olganlar: <b>{stats.get('test_statistics', {}).get('high_scorers_70plus', 0)}</b>\n\n"
        f"📅 So'nggi yangilanish: <b>{stats.get('last_updated', 'N/A')[:16]}</b>"
    )
    
    await message.answer(general_text)
    
    # Age group statistics
    age_stats = stats.get('test_statistics', {}).get('age_group_stats', {})
    age_text = (
        f"📊 <b>YOSH GURUHLARI BO'YICHA STATISTIKA</b>\n\n"
        f"👶 <b>7-10 yosh:</b>\n"
        f"   • Test topshirganlar: {age_stats.get('7-10', {}).get('count', 0)}\n"
        f"   • O'rtacha ball: {age_stats.get('7-10', {}).get('avg_score', 0)}\n\n"
        f"🧒 <b>11-14 yosh:</b>\n"
        f"   • Test topshirganlar: {age_stats.get('11-14', {}).get('count', 0)}\n"
        f"   • O'rtacha ball: {age_stats.get('11-14', {}).get('avg_score', 0)}"
    )
    
    await message.answer(age_text)
    
    # Regional statistics
    regional_stats = stats.get('regional_statistics', {})
    region_text = "🌍 <b>VILOYATLAR BO'YICHA STATISTIKA</b>\n\n"
    
    # Sort regions by user count
    sorted_regions = sorted(
        regional_stats.items(),
        key=lambda x: x[1].get('total_users', 0),
        reverse=True
    )
    
    for region, data in sorted_regions[:10]:  # Top 10 regions
        user_count = data.get('total_users', 0)
        if user_count > 0:
            region_text += f"📍 <b>{region}:</b> {user_count} foydalanuvchi\n"
    
    region_text += f"\n📱 <i>Barcha viloyatlar ma'lumotlari mavjud</i>"
    
    await message.answer(region_text)

# 🏆 Rankings (Super Admin only)
@dp.message(lambda message: message.text == "🏆 Reytinglar ro'yxati" and is_super_admin(message.from_user.id))
async def view_rankings(message: types.Message):
    # Update statistics first
    update_statistics()
    stats = get_statistics()
    
    top_performers = stats.get('top_performers', [])
    
    if not top_performers:
        await message.answer("🏆 Hozircha test natijalar mavjud emas.")
        return
    
    # Top 20 performers
    ranking_text = "🏆 <b>TOP NATIJALAR (Eng yaxshi 20 ta)</b>\n\n"
    
    for i, performer in enumerate(top_performers[:20], 1):
        medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
        
        ranking_text += (
            f"{medal} <b>{performer.get('user_name', 'N/A')}</b>\n"
            f"   💯 {performer.get('score', 0)}/100 ({performer.get('percentage', 0):.1f}%)\n"
            f"   👶 {performer.get('age', 'N/A')} yosh\n"
            f"   🌍 {performer.get('region', 'N/A')}\n"
            f"   📅 {performer.get('date', 'N/A')[:10]}\n\n"
        )
        
        # Split long messages
        if len(ranking_text) > 3500:
            await message.answer(ranking_text)
            ranking_text = ""
    
    if ranking_text:
        await message.answer(ranking_text)
    
    # Regional rankings
    regional_ranking_text = "🌍 <b>VILOYATLAR REYTINGI</b>\n\n"
    regional_stats = stats.get('regional_statistics', {})
    
    # Sort regions by user count
    sorted_regions = sorted(
        regional_stats.items(),
        key=lambda x: x[1].get('total_users', 0),
        reverse=True
    )
    
    for i, (region, data) in enumerate(sorted_regions[:15], 1):
        user_count = data.get('total_users', 0)
        if user_count > 0:
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            regional_ranking_text += f"{medal} <b>{region}:</b> {user_count} foydalanuvchi\n"
    
    await message.answer(regional_ranking_text)

# Certificate template upload removed as requested

# 🚫 Block unauthorized access to admin commands
@dp.message(lambda message: message.text in [
    "👥 Foydalanuvchilar ro'yxati", "➕ Test qo'shish", "👨‍💼 Adminlar ro'yxati",
    "➕ Admin qo'shish", "⬆️ Super Admin tayinlash", "➖ Admin o'chirish",
    "🗑 Test o'chirish", "📊 Test natijalarini yuklab olish", 
    "📋 Foydalanuvchi ma'lumotlarini yuklab olish", "📢 Xabar yuborish",
    "📊 Statistika va monitoring", "🏆 Reytinglar ro'yxati"
] and not is_admin(message.from_user.id))
async def block_unauthorized_admin_commands(message: types.Message):
    await message.answer("❌ Bu buyruq faqat adminlar uchun!")

# 🛡 Anti-spam filter
@dp.message(lambda message: message.chat.type == "private" and any(
    x in message.text.lower() for x in ["t.me", "http", "@"] if message.text and len(message.text.split()) > 5
))
async def block_spam(message: types.Message):
    await message.delete()
    await message.answer("⚠️ Reklama va spam taqiqlanadi!")

# 📣 Main function
async def main():
    """Start the bot"""
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
