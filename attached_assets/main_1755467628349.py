import logging
import asyncio
import json
import os
import re
import time
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
import random

# Database imports
from database import (
    DatabaseService, init_database, User, Question, TestResult, 
    TestAnswer, Feedback, RegionData
)

from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message, KeyboardButton, ReplyKeyboardMarkup,
    ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup,
    CallbackQuery, BufferedInputFile
)
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage

# Excel va PDF uchun kutubxonalar
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:
    SimpleDocTemplate = None

# 🔑 Bot token va admin ID'lar
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN")
ADMINS = [int(x) for x in os.getenv("ADMIN_IDS", "6578706277,7853664401").split(",")]

# 🔧 Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 🔧 Bot va Dispatcher
storage = MemoryStorage()
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=storage)

# 🔧 Holatlar (FSM States)
class RegistrationStates(StatesGroup):
    waiting_for_child_name = State()
    waiting_for_parent_name = State()
    waiting_for_age_group = State()
    waiting_for_region = State()
    waiting_for_manual_region = State()
    waiting_for_district = State()
    waiting_for_manual_district = State()
    waiting_for_mahalla = State()

class TestStates(StatesGroup):
    taking_test = State()
    test_question = State()

class AdminStates(StatesGroup):
    waiting_for_broadcast = State()
    waiting_for_question_text = State()
    waiting_for_question_file = State()
    waiting_for_new_question = State()
    waiting_for_new_options = State()
    waiting_for_correct_answer = State()
    waiting_for_age_group_selection = State()
    waiting_for_delete_question_id = State()
    waiting_for_bulk_questions = State()

class FeedbackStates(StatesGroup):
    waiting_for_feedback = State()

# 🔧 Ma'lumotlar saqlash
user_data = {}
active_tests = {}
test_timers = {}

# 🔧 Legacy data files (for migration if needed)
DATA_DIR = "data"
USERS_FILE = os.path.join(DATA_DIR, "users.json")
QUESTIONS_FILE = os.path.join(DATA_DIR, "test_questions.json")
REGIONS_FILE = os.path.join(DATA_DIR, "regions.json")

def ensure_data_dir():
    """Ma'lumotlar papkasini yaratish (legacy support)"""
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

def load_data(filename: str, default: Any = None) -> Any:
    """JSON fayldan ma'lumot yuklash (legacy support)"""
    ensure_data_dir()
    if os.path.exists(filename):
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    return default or {}

def save_data(filename: str, data: Any):
    """JSON faylga ma'lumot saqlash (legacy support)"""
    ensure_data_dir()
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# 🔧 Database helper functions
def load_users_from_db() -> Dict[str, Any]:
    """Load users from database in legacy format for compatibility"""
    users = DatabaseService.get_all_users()
    result = {}
    for user in users:
        result[user.user_id] = {
            "child_name": user.child_name,
            "parent_name": user.parent_name,
            "phone": user.phone,
            "age_group": user.age_group,
            "region": user.region,
            "district": user.district,
            "mahalla": user.mahalla,
            "telegram_username": user.telegram_username,
            "telegram_name": user.telegram_name,
            "registered": user.registered,
            "registration_date": user.registration_date.isoformat(),
            "test_results": [
                {
                    "age_group": result.age_group,
                    "total_questions": result.total_questions,
                    "correct_answers": result.correct_answers,
                    "percentage": result.percentage,
                    "duration": result.duration_seconds,
                    "date": result.completed_at.isoformat()
                }
                for result in user.test_results
            ]
        }
    return result

def load_questions_from_db() -> Dict[str, List[Dict]]:
    """Load questions from database in legacy format"""
    result = {}
    for age_group in ["7-10", "11-14"]:
        questions = DatabaseService.get_questions_by_age_group(age_group)
        result[age_group] = [
            {
                "question": q.question_text,
                "options": q.options,
                "correct": q.correct_answer
            }
            for q in questions
        ]
    return result

def save_user_to_db(user_id: str, user_data: Dict[str, Any]):
    """Save user to database"""
    try:
        existing_user = DatabaseService.get_user(user_id)
        if existing_user:
            DatabaseService.update_user(user_id, user_data)
        else:
            user_data["user_id"] = user_id
            DatabaseService.create_user(user_data)
    except Exception as e:
        logger.error(f"Error saving user to database: {e}")

def save_question_to_db(question_data: Dict[str, Any]) -> bool:
    """Save question to database"""
    try:
        DatabaseService.add_question(question_data)
        return True
    except Exception as e:
        logger.error(f"Error saving question to database: {e}")
        return False

# 🔧 Viloyatlar va tumanlar ma'lumotlarini yuklash/yaratish
def init_regions_data():
    """Viloyatlar va tumanlar ma'lumotlarini boshlang'ich holatga keltirish"""
    # Try to load from database first
    try:
        regions_from_db = DatabaseService.get_regions()
        if regions_from_db:
            return regions_from_db
    except Exception as e:
        logger.warning(f"Could not load regions from database: {e}")
    
    # Fallback to default data
    regions_data = {
        "Toshkent": {
            "districts": ["Chilonzor", "Shayxontohur", "Yunusobod", "Mirzo Ulug'bek", "Yakkasaroy", "Mirobod"],
            "mahallas": {
                "Chilonzor": ["1-mahalla", "2-mahalla", "3-mahalla"],
                "Shayxontohur": ["Markaziy", "Sharqiy", "G'arbiy"],
                "Yunusobod": ["Yangi Yunusobod", "Eski Yunusobod"]
            }
        },
        "Samarqand": {
            "districts": ["Samarqand shahar", "Urgut", "Bulung'ur", "Ishtixon", "Kattaqo'rg'on"],
            "mahallas": {
                "Samarqand shahar": ["Registon", "Siob", "Darvozaboy"],
                "Urgut": ["Markaz", "Yangiobod"],
                "Bulung'ur": ["Markaz", "Qishloq"]
            }
        },
        "Farg'ona": {
            "districts": ["Farg'ona shahar", "Marg'ilon", "Qo'qon", "Beshariq", "Bog'dod"],
            "mahallas": {
                "Farg'ona shahar": ["Markaz", "Yangi shahar"],
                "Marg'ilon": ["Ipak yo'li", "Hunarmandchilik"],
                "Qo'qon": ["Eski shahar", "Yangi shahar"]
            }
        },
        "Andijon": {
            "districts": ["Andijon shahar", "Xo'jaobod", "Asaka", "Baliqchi", "Bo'z"],
            "mahallas": {
                "Andijon shahar": ["Markaz", "Sanoat"],
                "Xo'jaobod": ["Qishloq", "Shahar"],
                "Asaka": ["Markaz"]
            }
        },
        "Namangan": {
            "districts": ["Namangan shahar", "To'raqo'rg'on", "Uchqo'rg'on", "Chust", "Pop"],
            "mahallas": {
                "Namangan shahar": ["Markaz", "Sanoat zone"],
                "To'raqo'rg'on": ["Markaz", "Qishloq"],
                "Chust": ["Hunarmandchilik", "Dehqonchilik"]
            }
        },
        "Qashqadaryo": {
            "districts": ["Qarshi", "Shahrisabz", "Kitob", "Yakkabog'", "Chiroqchi"],
            "mahallas": {
                "Qarshi": ["Markaz", "Sanoat"],
                "Shahrisabz": ["Tarixiy markaz", "Yangi shahar"],
                "Kitob": ["Markaz"]
            }
        },
        "Surxondaryo": {
            "districts": ["Termiz", "Denov", "Qaraqo'l", "Sho'rchi", "Muzrabot"],
            "mahallas": {
                "Termiz": ["Markaz", "Port zone"],
                "Denov": ["Markaz", "Qishloq"],
                "Qaraqo'l": ["Markaz"]
            }
        },
        "Sirdaryo": {
            "districts": ["Guliston", "Yangiyer", "Sirdaryo", "Boyovut", "Mirzaobod"],
            "mahallas": {
                "Guliston": ["Markaz", "Sanoat"],
                "Yangiyer": ["Markaz", "Qishloq"],
                "Sirdaryo": ["Markaz"]
            }
        },
        "Jizzax": {
            "districts": ["Jizzax shahar", "G'allaorol", "Zomin", "Baxtiyor", "Mirzacho'l"],
            "mahallas": {
                "Jizzax shahar": ["Markaz", "Yangi shahar"],
                "G'allaorol": ["Markaz", "Qishloq"],
                "Zomin": ["Tog'li", "Tekislik"]
            }
        },
        "Navoiy": {
            "districts": ["Navoiy shahar", "Zarafshon", "Nurota", "Karmana", "Qiziltepa"],
            "mahallas": {
                "Navoiy shahar": ["Markaz", "Sanoat"],
                "Zarafshon": ["Oltin qazish", "Shahar"],
                "Nurota": ["Tarixiy", "Markaz"]
            }
        },
        "Buxoro": {
            "districts": ["Buxoro shahar", "G'ijduvon", "Kogon", "Olot", "Peshku"],
            "mahallas": {
                "Buxoro shahar": ["Eski shahar", "Yangi shahar"],
                "G'ijduvon": ["Markaz", "Qishloq"],
                "Kogon": ["Temir yo'l", "Markaz"]
            }
        },
        "Xorazm": {
            "districts": ["Urganch", "Xiva", "Bog'ot", "Gurlan", "Qo'shko'pir"],
            "mahallas": {
                "Urganch": ["Markaz", "Sanoat"],
                "Xiva": ["Ichan qal'a", "Tashqi shahar"],
                "Bog'ot": ["Markaz"]
            }
        }
    }
    
    existing_data = load_data(REGIONS_FILE, {})
    if not existing_data:
        save_data(REGIONS_FILE, regions_data)
    return regions_data

def init_test_questions():
    """Test savollarini boshlang'ich holatga keltirish"""
    # Try to load from database first
    try:
        questions_from_db = load_questions_from_db()
        if questions_from_db and any(questions_from_db.values()):
            return questions_from_db
    except Exception as e:
        logger.warning(f"Could not load questions from database: {e}")
    
    # Fallback to default data
    questions_data = {
        "7-10": [
            {
                "question": "O'zbekistonning poytaxti qaysi shahar?",
                "options": ["Toshkent", "Samarqand", "Buxoro", "Farg'ona"],
                "correct": 0
            },
            {
                "question": "Alisher Navoiyning mashhur asari qaysi?",
                "options": ["Xamsa", "Lison ut-tayr", "Mahbub ul-qulub", "Barchasi"],
                "correct": 3
            },
            {
                "question": "O'zbekistonda nechta viloyat bor?",
                "options": ["12", "13", "14", "15"],
                "correct": 0
            },
            {
                "question": "\"Alpomish\" dostonining bosh qahramoni kim?",
                "options": ["Alpomish", "Barchin", "Qaldirg'och", "Ultontoz"],
                "correct": 0
            },
            {
                "question": "O'zbekistonning eng katta ko'li qaysi?",
                "options": ["Orol dengizi", "Aydar ko'li", "Sarmishsoy", "Charvak"],
                "correct": 1
            }
        ],
        "11-14": [
            {
                "question": "Amir Temurning tug'ilgan yili qachon?",
                "options": ["1336", "1340", "1350", "1360"],
                "correct": 0
            },
            {
                "question": "\"Xamsa\" asari necha qismdan iborat?",
                "options": ["3", "4", "5", "6"],
                "correct": 2
            },
            {
                "question": "O'zbekiston mustaqillik bayramini qachon nishonlaymiz?",
                "options": ["1-sentabr", "31-avgust", "1-yanvar", "8-dekabr"],
                "correct": 1
            },
            {
                "question": "Mirzo Ulug'bekning mashhur kuzatuvxonasi qayerda joylashgan?",
                "options": ["Toshkent", "Samarqand", "Buxoro", "Xiva"],
                "correct": 1
            },
            {
                "question": "\"Yodgor\" she'rining muallifi kim?",
                "options": ["Abdulla Oripov", "Erkin Vohidov", "Halima Xudoyberdiyeva", "Zulfiya"],
                "correct": 1
            }
        ]
    }
    
    existing_questions = load_data(QUESTIONS_FILE, {})
    if not existing_questions:
        save_data(QUESTIONS_FILE, questions_data)
    return questions_data

# Ma'lumotlarni yuklash
regions_data = init_regions_data()
test_questions = init_test_questions()

# Database initialization
try:
    init_database()
    logger.info("Database initialized successfully")
except Exception as e:
    logger.error(f"Database initialization failed: {e}")
    logger.info("Falling back to JSON file storage")

def get_main_keyboard():
    """Asosiy klaviatura"""
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="☎️ Telefon raqamni yuborish", request_contact=True)],
            [KeyboardButton(text="📚 Loyiha haqida"), KeyboardButton(text="📝 Test topshirish")],
            [KeyboardButton(text="💭 Fikr bildirish"), KeyboardButton(text="📊 Mening natijalarim")]
        ],
        resize_keyboard=True,
        input_field_placeholder="Tugmani tanlang..."
    )
    return kb

def get_admin_keyboard():
    """Admin klaviaturasi"""
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="👥 Foydalanuvchilar ro'yxati"), KeyboardButton(text="📊 Statistika")],
            [KeyboardButton(text="📝 Savollar boshqaruvi"), KeyboardButton(text="🔴 Test monitoring")],
            [KeyboardButton(text="📢 Xabar yuborish"), KeyboardButton(text="📤 Ma'lumot eksport")],
            [KeyboardButton(text="🔙 Oddiy foydalanuvchi")]
        ],
        resize_keyboard=True
    )
    return kb

def is_admin(user_id: int) -> bool:
    """Foydalanuvchi admin ekanligini tekshirish"""
    return user_id in ADMINS

def get_user_info(user_id: int) -> Dict:
    """Foydalanuvchi ma'lumotlarini olish"""
    try:
        user = DatabaseService.get_user(str(user_id))
        if user:
            return {
                "user_id": user.user_id,
                "child_name": user.child_name,
                "parent_name": user.parent_name,
                "phone": user.phone,
                "age_group": user.age_group,
                "region": user.region,
                "district": user.district,
                "mahalla": user.mahalla,
                "telegram_username": user.telegram_username,
                "telegram_name": user.telegram_name,
                "registered": user.registered,
                "registration_date": user.registration_date.isoformat() if user.registration_date else "",
                "test_results": [
                    {
                        "age_group": result.age_group,
                        "total_questions": result.total_questions,
                        "correct_answers": result.correct_answers,
                        "percentage": result.percentage,
                        "duration": result.duration_seconds,
                        "date": result.completed_at.isoformat()
                    }
                    for result in user.test_results
                ]
            }
        return {}
    except Exception as e:
        logger.error(f"Error getting user info: {e}")
        # Fallback to JSON file
        users = load_data(USERS_FILE, {})
        return users.get(str(user_id), {})

def save_user_info(user_id: int, data: Dict):
    """Foydalanuvchi ma'lumotlarini saqlash"""
    try:
        # Remove test_results from data before saving user info
        user_data = data.copy()
        if "test_results" in user_data:
            del user_data["test_results"]
        
        # Convert user_id to string if it's in the data
        if "user_id" in user_data:
            user_data["user_id"] = str(user_data["user_id"])
        
        save_user_to_db(str(user_id), user_data)
    except Exception as e:
        logger.error(f"Error saving user to database: {e}")
        # Fallback to JSON file
        users = load_data(USERS_FILE, {})
        users[str(user_id)] = data
        save_data(USERS_FILE, users)

# 📌 START komandasi
@dp.message(CommandStart())
async def start_cmd(message: Message, state: FSMContext):
    await state.clear()
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    
    if user_info.get("registered"):
        # Ro'yxatdan o'tgan foydalanuvchi
        kb = get_admin_keyboard() if is_admin(user_id) else get_main_keyboard()
        admin_text = "👑 Admin" if is_admin(user_id) else ""
        await message.answer(
            f"👋 Salom, {message.from_user.first_name}! {admin_text}\n\n"
            f"Siz allaqachon ro'yxatdan o'tgansiz. Kerakli bo'limni tanlang:",
            reply_markup=kb
        )
    else:
        # Yangi foydalanuvchi
        kb = ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="☎️ Telefon raqamni yuborish", request_contact=True)],
                [KeyboardButton(text="📚 Loyiha haqida")]
            ],
            resize_keyboard=True
        )
        await message.answer(
            "🌟 **Salom va xush kelibsiz!** 🌟\n\n"
            "📚 **'Kitobxon Kids'** - O'zbekiston bolalari uchun maxsus kitobxonlik tanloviga qadam qo'yasiz!\n\n"
            "🎯 **Sizni quyidagilar kutmoqda:**\n"
            "✨ Qiziqarli test savollari\n"
            "🏆 Bilimingizni sinash imkoniyati\n"
            "🎁 Ajoyib sovrinlar va sertifikatlar\n"
            "👨‍👩‍👧‍👦 Oila a'zolari bilan birgalikda o'rganish\n\n"
            "🚀 **Keling, birgalikda bilim olaylik va o'rganaylik!**\n\n"
            "Boshlash uchun telefon raqamingizni yuboring 👇",
            reply_markup=kb,
            parse_mode="Markdown"
        )

# 📌 Telefon raqam qabul qilish
@dp.message(F.contact)
async def get_contact(message: Message, state: FSMContext):
    user_id = message.from_user.id
    phone = message.contact.phone_number
    
    # Vaqtincha ma'lumotlarni saqlash
    await state.update_data(
        user_id=user_id,
        phone=phone,
        telegram_username=message.from_user.username or "",
        telegram_name=message.from_user.first_name or ""
    )
    
    await message.answer(
        "👶 Bolaning to'liq ism-sharifini kiriting:",
        reply_markup=ReplyKeyboardRemove()
    )
    await state.set_state(RegistrationStates.waiting_for_child_name)

# 📌 Bolaning ismi
@dp.message(RegistrationStates.waiting_for_child_name)
async def get_child_name(message: Message, state: FSMContext):
    await state.update_data(child_name=message.text.strip())
    await message.answer("👨‍👩‍👧 Ota-onaning to'liq ism-sharifini kiriting:")
    await state.set_state(RegistrationStates.waiting_for_parent_name)

# 📌 Ota-onaning ismi
@dp.message(RegistrationStates.waiting_for_parent_name)
async def get_parent_name(message: Message, state: FSMContext):
    await state.update_data(parent_name=message.text.strip())
    
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="7–10 yosh")],
            [KeyboardButton(text="11–14 yosh")]
        ],
        resize_keyboard=True
    )
    await message.answer("📌 Yosh toifasini tanlang:", reply_markup=kb)
    await state.set_state(RegistrationStates.waiting_for_age_group)

# 📌 Yosh toifasi
@dp.message(RegistrationStates.waiting_for_age_group)
async def get_age_group(message: Message, state: FSMContext):
    if message.text not in ["7–10 yosh", "11–14 yosh"]:
        await message.answer("Iltimos, tugmalardan birini tanlang!")
        return
    
    await state.update_data(age_group=message.text)
    
    # Viloyatlar ro'yxati
    keyboard = []
    for region in regions_data.keys():
        keyboard.append([KeyboardButton(text=region)])
    keyboard.append([KeyboardButton(text="✍️ Qo'lda kiritish")])
    
    kb = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    await message.answer("🏢 Viloyatni tanlang yoki qo'lda kiriting:", reply_markup=kb)
    await state.set_state(RegistrationStates.waiting_for_region)

# 📌 Viloyat tanlash
@dp.message(RegistrationStates.waiting_for_region)
async def get_region(message: Message, state: FSMContext):
    if message.text == "✍️ Qo'lda kiritish":
        await message.answer("✍️ Viloyat nomini yozib kiriting:", reply_markup=ReplyKeyboardRemove())
        await state.set_state(RegistrationStates.waiting_for_manual_region)
        return
    
    if message.text not in regions_data:
        await message.answer("Iltimos, ro'yxatdan viloyatni tanlang yoki 'Qo'lda kiritish' tugmasini bosing!")
        return
    
    await state.update_data(region=message.text)
    
    # Tumanlar ro'yxati
    districts = regions_data[message.text]["districts"]
    keyboard = []
    for district in districts:
        keyboard.append([KeyboardButton(text=district)])
    keyboard.append([KeyboardButton(text="✍️ Qo'lda kiritish")])
    
    kb = ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    await message.answer("🏙 Tuman/Shaharni tanlang yoki qo'lda kiriting:", reply_markup=kb)
    await state.set_state(RegistrationStates.waiting_for_district)

# 📌 Qo'lda viloyat kiritish
@dp.message(RegistrationStates.waiting_for_manual_region)
async def get_manual_region(message: Message, state: FSMContext):
    await state.update_data(region=message.text.strip())
    await message.answer("🏙 Tuman/Shahar nomini yozib kiriting:")
    await state.set_state(RegistrationStates.waiting_for_manual_district)

# 📌 Tuman tanlash
@dp.message(RegistrationStates.waiting_for_district)
async def get_district(message: Message, state: FSMContext):
    user_data = await state.get_data()
    region = user_data.get("region")
    
    if message.text == "✍️ Qo'lda kiritish":
        await message.answer("✍️ Tuman/Shahar nomini yozib kiriting:", reply_markup=ReplyKeyboardRemove())
        await state.set_state(RegistrationStates.waiting_for_manual_district)
        return
    
    # Tekshirish
    if region in regions_data and message.text not in regions_data[region]["districts"]:
        await message.answer("Iltimos, ro'yxatdan tumanni tanlang yoki 'Qo'lda kiritish' tugmasini bosing!")
        return
    
    await state.update_data(district=message.text)
    await message.answer("🏡 Mahallangizni yozib kiriting:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(RegistrationStates.waiting_for_mahalla)

# 📌 Qo'lda tuman kiritish
@dp.message(RegistrationStates.waiting_for_manual_district)
async def get_manual_district(message: Message, state: FSMContext):
    await state.update_data(district=message.text.strip())
    await message.answer("🏡 Mahallangizni yozib kiriting:", reply_markup=ReplyKeyboardRemove())
    await state.set_state(RegistrationStates.waiting_for_mahalla)

# 📌 Mahalla va ro'yxatni yakunlash
@dp.message(RegistrationStates.waiting_for_mahalla)
async def get_mahalla(message: Message, state: FSMContext):
    await state.update_data(mahalla=message.text.strip())
    
    # Barcha ma'lumotlarni olish
    data = await state.get_data()
    user_id = data["user_id"]
    
    # Ro'yxat ma'lumotlari
    user_info = {
        "user_id": user_id,
        "child_name": data["child_name"],
        "parent_name": data["parent_name"],
        "phone": data["phone"],
        "age_group": data["age_group"],
        "region": data["region"],
        "district": data["district"],
        "mahalla": data["mahalla"],
        "telegram_username": data["telegram_username"],
        "telegram_name": data["telegram_name"],
        "registration_date": datetime.now().isoformat(),
        "registered": True,
        "test_results": []
    }
    
    # Saqlash
    save_user_info(user_id, user_info)
    
    # Xabar matnini tayyorlash
    text = (
        f"✅ Siz ro'yxatdan muvaffaqiyatli o'tdingiz!\n\n"
        f"👶 Bola: {data['child_name']}\n"
        f"👨‍👩‍👧 Ota/Ona: {data['parent_name']}\n"
        f"📞 Telefon: {data['phone']}\n"
        f"📌 Yosh toifa: {data['age_group']}\n"
        f"🏢 Viloyat: {data['region']}\n"
        f"🏙 Tuman/Shahar: {data['district']}\n"
        f"🏡 Mahalla: {data['mahalla']}\n\n"
        f"Endi siz test topshira olasiz!"
    )
    
    kb = get_main_keyboard()
    await message.answer(text, reply_markup=kb)
    
    # Adminlarga xabar yuborish
    admin_text = f"📥 Yangi ro'yxatdan o'tuvchi:\n\n{text}"
    for admin_id in ADMINS:
        try:
            await bot.send_message(admin_id, admin_text)
        except Exception as e:
            logger.error(f"Adminga xabar yuborishda xatolik: {e}")
    
    await state.clear()

# 📌 Loyiha haqida
@dp.message(F.text == "📚 Loyiha haqida")
async def about_project(message: Message):
    text = (
        "📚 *Kitobxon Kids* loyihasi haqida ma'lumot:\n\n"
        "🎯 **Maqsad:** Bolalarda kitobxonlik madaniyatini rivojlantirish\n\n"
        "👥 **Ishtirokchilar:** 7-10 va 11-14 yosh toifasidagi bolalar\n\n"
        "📝 **Test formati:**\n"
        "• Har bir test 25 ta savoldan iborat\n"
        "• Har bir savol uchun 20 soniya vaqt\n"
        "• Savollar tasodifiy tartibda beriladi\n\n"
        "🏆 **Mukofotlar:**\n"
        "• Eng yaxshi natija ko'rsatganlar rag'batlantiriladi\n"
        "• Barcha ishtirokchilar sertifikat oladi\n\n"
        "📞 **Aloqa:** Savollar bo'lsa admin bilan bog'laning\n\n"
        "🔗 **Qo'shimcha ma'lumot:** Tez orada e'lon qilinadi"
    )
    await message.answer(text, parse_mode="Markdown")

# 📌 Test topshirish
@dp.message(F.text == "📝 Test topshirish")
async def start_test(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    
    if not user_info.get("registered"):
        await message.answer("❌ Avval ro'yxatdan o'tishingiz kerak!")
        return
    
    if user_id in active_tests:
        await message.answer("⚠️ Siz allaqachon test topshirayotgansiz!")
        return
    
    age_group = user_info["age_group"]
    test_type = "7-10" if "7–10" in age_group else "11-14"
    
    questions = test_questions.get(test_type, [])
    if not questions:
        await message.answer("❌ Sizning yosh toifangiz uchun savollar hali tayyorlanmagan!")
        return
    
    # 25 ta tasodifiy savol tanlash
    selected_questions = random.sample(questions, min(25, len(questions)))
    
    active_tests[user_id] = {
        "questions": selected_questions,
        "current_question": 0,
        "correct_answers": 0,
        "start_time": time.time(),
        "answers": []
    }
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Testni boshlash", callback_data="start_test_confirm")],
            [InlineKeyboardButton(text="❌ Bekor qilish", callback_data="cancel_test")]
        ]
    )
    
    await message.answer(
        f"📝 **Test topshirish**\n\n"
        f"🎯 Yosh toifa: {age_group}\n"
        f"📊 Savollar soni: {len(selected_questions)}\n"
        f"⏱ Har savol uchun: 20 soniya\n"
        f"⚠️ Diqqat: Test boshlangandan so'ng chiqib ketmaslik kerak!\n\n"
        f"Tayyormisiz?",
        reply_markup=kb,
        parse_mode="Markdown"
    )

# 📌 Test tasdiqlash
@dp.callback_query(F.data == "start_test_confirm")
async def confirm_test_start(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    
    if user_id not in active_tests:
        await callback.answer("❌ Test ma'lumotlari topilmadi!")
        return
    
    await callback.answer()
    await callback.message.delete()
    
    # Birinchi savolni yuborish
    await send_question(callback.message, user_id, state)

@dp.callback_query(F.data == "cancel_test")
async def cancel_test(callback: CallbackQuery):
    user_id = callback.from_user.id
    if user_id in active_tests:
        del active_tests[user_id]
    
    await callback.answer("Test bekor qilindi!")
    await callback.message.delete()

async def send_question(message: Message, user_id: int, state: FSMContext):
    """Savolni yuborish"""
    if user_id not in active_tests:
        return
    
    test_data = active_tests[user_id]
    current_q = test_data["current_question"]
    questions = test_data["questions"]
    
    if current_q >= len(questions):
        await finish_test(message, user_id)
        return
    
    question_data = questions[current_q]
    question_text = question_data["question"]
    options = question_data["options"]
    
    # Tugmalarni yaratish
    keyboard = []
    for i, option in enumerate(options):
        keyboard.append([InlineKeyboardButton(
            text=f"{chr(65 + i)}) {option}", 
            callback_data=f"answer_{i}"
        )])
    
    kb = InlineKeyboardMarkup(inline_keyboard=keyboard)
    
    await message.answer(
        f"❓ **Savol {current_q + 1}/{len(questions)}**\n\n"
        f"{question_text}\n\n"
        f"⏱ 20 soniya",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    
    # 20 soniyalik timer
    await asyncio.sleep(20)
    
    # Agar foydalanuvchi javob bermagan bo'lsa
    if user_id in active_tests and active_tests[user_id]["current_question"] == current_q:
        active_tests[user_id]["current_question"] += 1
        active_tests[user_id]["answers"].append(-1)  # Javob berilmagan
        
        await message.answer("⏰ Vaqt tugadi! Keyingi savol...")
        await asyncio.sleep(1)
        await send_question(message, user_id, state)

# 📌 Test javoblarini qabul qilish
@dp.callback_query(F.data.startswith("answer_"))
async def handle_answer(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    
    if user_id not in active_tests:
        await callback.answer("❌ Test ma'lumotlari topilmadi!")
        return
    
    answer_index = int(callback.data.split("_")[1])
    test_data = active_tests[user_id]
    current_q = test_data["current_question"]
    
    # Javobni tekshirish
    correct_answer = test_data["questions"][current_q]["correct"]
    is_correct = answer_index == correct_answer
    
    if is_correct:
        test_data["correct_answers"] += 1
        await callback.answer("✅ To'g'ri!")
    else:
        await callback.answer("❌ Noto'g'ri!")
    
    # Javobni saqlash
    test_data["answers"].append(answer_index)
    test_data["current_question"] += 1
    
    await callback.message.delete()
    
    # Keyingi savol yoki test tugashi
    if test_data["current_question"] >= len(test_data["questions"]):
        await finish_test(callback.message, user_id)
    else:
        await asyncio.sleep(1)
        await send_question(callback.message, user_id, state)

async def finish_test(message: Message, user_id: int):
    """Testni yakunlash"""
    if user_id not in active_tests:
        return
    
    test_data = active_tests[user_id]
    user_info = get_user_info(user_id)
    
    total_questions = len(test_data["questions"])
    correct_answers = test_data["correct_answers"]
    percentage = round((correct_answers / total_questions) * 100, 1)
    duration = round(time.time() - test_data["start_time"], 1)
    
    # Natijani saqlash
    result = {
        "date": datetime.now().isoformat(),
        "total_questions": total_questions,
        "correct_answers": correct_answers,
        "percentage": percentage,
        "duration": duration,
        "age_group": user_info["age_group"]
    }
    
    user_info["test_results"].append(result)
    save_user_info(user_id, user_info)
    
    # Natijani ko'rsatish
    grade = "A" if percentage >= 90 else "B" if percentage >= 70 else "C" if percentage >= 50 else "D"
    
    text = (
        f"🎉 **Test yakunlandi!**\n\n"
        f"📊 **Natijangiz:**\n"
        f"✅ To'g'ri javoblar: {correct_answers}/{total_questions}\n"
        f"📈 Foiz: {percentage}%\n"
        f"🏆 Baho: {grade}\n"
        f"⏱ Vaqt: {duration} soniya\n\n"
        f"{'🎊 Ajoyib natija!' if percentage >= 80 else '👍 Yaxshi natija!' if percentage >= 60 else '💪 Keyingi safar yaxshiroq!'}"
    )
    
    kb = get_main_keyboard()
    await message.answer(text, reply_markup=kb, parse_mode="Markdown")
    
    # Adminlarga xabar
    admin_text = (
        f"📝 Yangi test natijasi:\n\n"
        f"👤 {user_info['child_name']}\n"
        f"📊 Natija: {correct_answers}/{total_questions} ({percentage}%)\n"
        f"🎯 Yosh toifa: {user_info['age_group']}\n"
        f"📍 {user_info['region']}, {user_info['district']}"
    )
    
    for admin_id in ADMINS:
        try:
            await bot.send_message(admin_id, admin_text)
        except:
            pass
    
    # Testni o'chirish
    del active_tests[user_id]

# 📌 Natijalarni ko'rish
@dp.message(F.text == "📊 Mening natijalarim")
async def show_my_results(message: Message):
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    
    if not user_info.get("registered"):
        await message.answer("❌ Avval ro'yxatdan o'tishingiz kerak!")
        return
    
    results = user_info.get("test_results", [])
    
    if not results:
        await message.answer("📊 Siz hali test topshirmagansiz!")
        return
    
    text = f"📊 **{user_info['child_name']} ning natijalari:**\n\n"
    
    for i, result in enumerate(results[-5:], 1):  # So'nggi 5 ta natija
        date = datetime.fromisoformat(result["date"]).strftime("%d.%m.%Y %H:%M")
        text += (
            f"**{i}. Test ({date})**\n"
            f"✅ {result['correct_answers']}/{result['total_questions']} ({result['percentage']}%)\n"
            f"⏱ {result['duration']} soniya\n\n"
        )
    
    if len(results) > 5:
        text += f"📝 Jami {len(results)} ta test topshirilgan"
    
    await message.answer(text, parse_mode="Markdown")

# 📌 Fikr bildirish
@dp.message(F.text == "💭 Fikr bildirish")
async def feedback_start(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    
    if not user_info.get("registered"):
        await message.answer("❌ Avval ro'yxatdan o'tishingiz kerak!")
        return
    
    await message.answer(
        "💭 **Fikr bildirish**\n\n"
        "Loyiha haqida fikr, taklif yoki shikoyatingizni yozing.\n"
        "Sizning fikringiz biz uchun muhim!",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(FeedbackStates.waiting_for_feedback)

@dp.message(FeedbackStates.waiting_for_feedback)
async def get_feedback(message: Message, state: FSMContext):
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    feedback_text = message.text
    
    # Fikrni adminlarga yuborish
    admin_text = (
        f"💭 **Yangi fikr-mulohaza:**\n\n"
        f"👤 **Foydalanuvchi:** {user_info['child_name']}\n"
        f"👨‍👩‍👧 **Ota-ona:** {user_info['parent_name']}\n"
        f"📞 **Telefon:** {user_info['phone']}\n"
        f"📍 **Joylashuv:** {user_info['region']}, {user_info['district']}\n"
        f"📱 **Telegram:** @{user_info.get('telegram_username', 'mavjud emas')}\n\n"
        f"💬 **Fikr:**\n{feedback_text}"
    )
    
    for admin_id in ADMINS:
        try:
            await bot.send_message(admin_id, admin_text, parse_mode="Markdown")
        except:
            pass
    
    kb = get_main_keyboard()
    await message.answer(
        "✅ Fikringiz muvaffaqiyatli yuborildi!\n"
        "Tez orada adminlar ko'rib chiqishadi.",
        reply_markup=kb
    )
    await state.clear()

# 📌 ADMIN KOMANDALAR

@dp.message(F.text == "👑 Admin panel")
@dp.message(Command("admin"))
async def admin_panel(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    kb = get_admin_keyboard()
    await message.answer("👑 **Admin panel**\n\nKerakli bo'limni tanlang:", reply_markup=kb, parse_mode="Markdown")

@dp.message(F.text == "🔙 Oddiy foydalanuvchi")
async def back_to_user(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    kb = get_main_keyboard()
    await message.answer("👤 Oddiy foydalanuvchi rejimiga qaytdingiz", reply_markup=kb)

@dp.message(F.text == "👥 Foydalanuvchilar ro'yxati")
async def users_list(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    users = load_data(USERS_FILE, {})
    
    if not users:
        await message.answer("📝 Hech kim ro'yxatdan o'tmagan!")
        return
    
    text = "👥 **Ro'yxatdan o'tganlar:**\n\n"
    
    for i, (uid, user_info) in enumerate(users.items(), 1):
        if i > 50:  # Birinchi 50 tani ko'rsatish
            text += f"\n... va yana {len(users) - 50} kishi"
            break
        
        text += (
            f"**{i}. {user_info['child_name']}**\n"
            f"👨‍👩‍👧 {user_info['parent_name']}\n"
            f"📞 {user_info['phone']}\n"
            f"🎯 {user_info['age_group']}\n"
            f"📍 {user_info['region']}, {user_info['district']}\n"
            f"📅 {datetime.fromisoformat(user_info['registration_date']).strftime('%d.%m.%Y')}\n\n"
        )
    
    await message.answer(text, parse_mode="Markdown")

@dp.message(F.text == "📊 Statistika")
async def show_statistics(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    users = load_data(USERS_FILE, {})
    
    if not users:
        await message.answer("📊 Ma'lumotlar yo'q!")
        return
    
    # Umumiy statistika
    total_users = len(users)
    age_7_10 = sum(1 for u in users.values() if "7–10" in u.get("age_group", ""))
    age_11_14 = sum(1 for u in users.values() if "11–14" in u.get("age_group", ""))
    
    total_tests = sum(len(u.get("test_results", [])) for u in users.values())
    
    # Viloyat bo'yicha statistika
    region_stats = {}
    for user_info in users.values():
        region = user_info.get("region", "Noma'lum")
        if region not in region_stats:
            region_stats[region] = {"users": 0, "tests": 0}
        region_stats[region]["users"] += 1
        region_stats[region]["tests"] += len(user_info.get("test_results", []))
    
    text = (
        f"📊 **Umumiy statistika:**\n\n"
        f"👥 Jami foydalanuvchilar: {total_users}\n"
        f"👶 7-10 yosh: {age_7_10}\n"
        f"🧒 11-14 yosh: {age_11_14}\n"
        f"📝 Jami testlar: {total_tests}\n\n"
        f"📍 **Viloyatlar bo'yicha:**\n"
    )
    
    for region, stats in sorted(region_stats.items()):
        text += f"• {region}: {stats['users']} kishi, {stats['tests']} test\n"
    
    await message.answer(text, parse_mode="Markdown")

@dp.message(F.text == "🔴 Test monitoring")
async def test_monitoring(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    if not active_tests:
        await message.answer("🔴 Hozirda hech kim test topshirmayapti!")
        return
    
    text = "🔴 **Faol testlar:**\n\n"
    
    for test_user_id, test_data in active_tests.items():
        users = load_data(USERS_FILE, {})
        user_info = users.get(str(test_user_id), {})
        name = user_info.get("child_name", "Noma'lum")
        current_q = test_data["current_question"]
        total_q = len(test_data["questions"])
        duration = round(time.time() - test_data["start_time"], 1)
        
        text += (
            f"👤 **{name}**\n"
            f"📊 {current_q}/{total_q} savol\n"
            f"⏱ {duration} soniya\n"
            f"✅ {test_data['correct_answers']} to'g'ri\n\n"
        )
    
    await message.answer(text, parse_mode="Markdown")

@dp.message(F.text == "📝 Savollar boshqaruvi")
async def question_management(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    questions = load_data(QUESTIONS_FILE, {})
    total_7_10 = len(questions.get("7-10", []))
    total_11_14 = len(questions.get("11-14", []))
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="➕ Savol qo'shish", callback_data="add_question")],
            [InlineKeyboardButton(text="📝 Ko'p savol qo'shish", callback_data="bulk_add_questions")],
            [InlineKeyboardButton(text="❌ Savol o'chirish", callback_data="delete_question")],
            [InlineKeyboardButton(text="📊 Savollarni ko'rish", callback_data="view_questions")],
            [InlineKeyboardButton(text="📁 Fayl orqali yuklash", callback_data="upload_questions")]
        ]
    )
    
    await message.answer(
        f"📝 **Savollar boshqaruvi**\n\n"
        f"📊 **Hozirgi holat:**\n"
        f"👶 7-10 yosh: {total_7_10} ta savol\n"
        f"🧒 11-14 yosh: {total_11_14} ta savol\n\n"
        f"Kerakli amalni tanlang:",
        reply_markup=kb,
        parse_mode="Markdown"
    )

@dp.callback_query(F.data == "add_question")
async def add_question_start(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q!")
        return
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👶 7-10 yosh", callback_data="age_7_10")],
            [InlineKeyboardButton(text="🧒 11-14 yosh", callback_data="age_11_14")]
        ]
    )
    
    await callback.message.edit_text(
        "➕ **Yangi savol qo'shish**\n\n"
        "Qaysi yosh toifasi uchun savol qo'shmoqchisiz?",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.in_(["age_7_10", "age_11_14"]))
async def age_group_selected(callback: CallbackQuery, state: FSMContext):
    age_group = "7-10" if callback.data == "age_7_10" else "11-14"
    await state.update_data(age_group=age_group)
    
    await callback.message.edit_text(
        f"➕ **Yangi savol qo'shish** ({age_group} yosh)\n\n"
        "**Format 1 (To'liq format - tavsiya etiladi):**\n"
        "```\n"
        "1. Savol matni?\n"
        "A) Variant 1\n"
        "B) Variant 2\n"
        "C) Variant 3\n"
        "D) Variant 4\n"
        "Javob: C\n"
        "```\n\n"
        "**Format 2 (Bosqichma-bosqich):**\n"
        "Faqat savol matnini kiriting:",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_new_question)
    await callback.answer()

@dp.message(AdminStates.waiting_for_new_question)
async def get_new_question(message: Message, state: FSMContext):
    text = message.text.strip()
    
    # Check if full question format is provided
    if "A)" in text and "B)" in text and "C)" in text and "D)" in text and "Javob:" in text:
        # Parse the complete question format
        lines = text.split('\n')
        question_text = ""
        options = []
        correct_answer = None
        
        i = 0
        # Extract question text (everything before A))
        while i < len(lines) and not lines[i].strip().startswith('A)'):
            if lines[i].strip():
                question_text += lines[i].strip() + " "
            i += 1
        
        question_text = question_text.strip()
        
        # Extract options and answer
        for line in lines:
            line = line.strip()
            if line.startswith('A)'):
                options.append(line[2:].strip())
            elif line.startswith('B)'):
                options.append(line[2:].strip())
            elif line.startswith('C)'):
                options.append(line[2:].strip())
            elif line.startswith('D)'):
                options.append(line[2:].strip())
            elif line.startswith('Javob:'):
                answer_letter = line[6:].strip().upper()
                if answer_letter in ['A', 'B', 'C', 'D']:
                    correct_answer = ord(answer_letter) - ord('A')
        
        if question_text and len(options) == 4 and correct_answer is not None:
            # Save the complete question
            data = await state.get_data()
            new_question = {
                "question": question_text,
                "options": options,
                "correct": correct_answer
            }
            
            questions = load_data(QUESTIONS_FILE, {})
            if data["age_group"] not in questions:
                questions[data["age_group"]] = []
            
            questions[data["age_group"]].append(new_question)
            save_data(QUESTIONS_FILE, questions)
            
            await message.answer(
                f"✅ **Savol muvaffaqiyatli qo'shildi!**\n\n"
                f"**Yosh toifasi:** {data['age_group']}\n"
                f"**Savol:** {question_text}\n"
                f"**To'g'ri javob:** {options[correct_answer]}\n\n"
                f"Jami savollar: {len(questions[data['age_group']])} ta",
                parse_mode="Markdown"
            )
            await state.clear()
            return
    
    # If not complete format, store question text and ask for options
    await state.update_data(question_text=text)
    await message.answer(
        "✏️ **Javob variantlarini kiriting**\n\n"
        "**Format 1 (Tavsiya etiladi):**\n"
        "```\n"
        "A) Variant 1\n"
        "B) Variant 2\n"
        "C) Variant 3\n"
        "D) Variant 4\n"
        "Javob: C\n"
        "```\n\n"
        "**Format 2:**\n"
        "4 ta javob variantini alohida qatorlarda:\n"
        "Toshkent\n"
        "Samarqand\n"
        "Buxoro\n"
        "Farg'ona",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_new_options)

@dp.message(AdminStates.waiting_for_new_options)
async def get_new_options(message: Message, state: FSMContext):
    text = message.text.strip()
    
    # Check if it's in the specified format
    if "A)" in text and "B)" in text and "C)" in text and "D)" in text:
        # Parse the formatted question
        lines = text.split('\n')
        options = []
        correct_answer = None
        
        for line in lines:
            line = line.strip()
            if line.startswith('A)'):
                options.append(line[2:].strip())
            elif line.startswith('B)'):
                options.append(line[2:].strip())
            elif line.startswith('C)'):
                options.append(line[2:].strip())
            elif line.startswith('D)'):
                options.append(line[2:].strip())
            elif line.startswith('Javob:'):
                answer_letter = line[6:].strip().upper()
                if answer_letter in ['A', 'B', 'C', 'D']:
                    correct_answer = ord(answer_letter) - ord('A')
        
        if len(options) == 4 and correct_answer is not None:
            await state.update_data(options=options, correct_answer=correct_answer)
            
            # Save the question immediately
            data = await state.get_data()
            new_question = {
                "question": data["question_text"],
                "options": options,
                "correct": correct_answer
            }
            
            questions = load_data(QUESTIONS_FILE, {})
            if data["age_group"] not in questions:
                questions[data["age_group"]] = []
            
            questions[data["age_group"]].append(new_question)
            save_data(QUESTIONS_FILE, questions)
            
            await message.answer(
                f"✅ **Savol muvaffaqiyatli qo'shildi!**\n\n"
                f"**Yosh toifasi:** {data['age_group']}\n"
                f"**Savol:** {data['question_text']}\n"
                f"**To'g'ri javob:** {options[correct_answer]}\n\n"
                f"Jami savollar: {len(questions[data['age_group']])} ta",
                parse_mode="Markdown"
            )
            await state.clear()
            return
    
    # If not in the specified format, use the old method
    options = [opt.strip() for opt in text.split('\n') if opt.strip()]
    
    if len(options) != 4:
        await message.answer(
            "❌ **Noto'g'ri format!**\n\n"
            "Quyidagi formatda kiriting:\n"
            "```\n"
            "A) Variant 1\n"
            "B) Variant 2\n"
            "C) Variant 3\n"
            "D) Variant 4\n"
            "Javob: C\n"
            "```\n\n"
            "Yoki 4 ta alohida qatorda yozing.",
            parse_mode="Markdown"
        )
        return
    
    await state.update_data(options=options)
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=f"{i+1}. {opt}", callback_data=f"correct_{i}")]
            for i, opt in enumerate(options)
        ]
    )
    
    await message.answer(
        "✅ **To'g'ri javobni tanlang:**\n\n" + 
        "\n".join([f"{i+1}. {opt}" for i, opt in enumerate(options)]),
        reply_markup=kb,
        parse_mode="Markdown"
    )

@dp.callback_query(F.data.startswith("correct_"))
async def save_new_question(callback: CallbackQuery, state: FSMContext):
    correct_index = int(callback.data.split("_")[1])
    data = await state.get_data()
    
    new_question = {
        "question": data["question_text"],
        "options": data["options"],
        "correct": correct_index
    }
    
    # Save question to database
    question_data = {
        "question_text": new_question["question"],
        "age_group": data["age_group"],
        "options": new_question["options"],
        "correct_answer": new_question["correct"]
    }
    success = save_question_to_db(question_data)
    
    if success:
        # Update global test_questions
        global test_questions
        test_questions = load_questions_from_db()
    
    # Get updated count
    current_questions = test_questions.get(data["age_group"], [])
    
    await callback.message.edit_text(
        f"✅ **Savol muvaffaqiyatli qo'shildi!**\n\n"
        f"**Yosh toifasi:** {data['age_group']}\n"
        f"**Savol:** {data['question_text']}\n"
        f"**To'g'ri javob:** {data['options'][correct_index]}\n\n"
        f"Jami savollar: {len(current_questions)} ta",
        parse_mode="Markdown"
    )
    await state.clear()
    await callback.answer("✅ Savol qo'shildi!")

@dp.callback_query(F.data == "delete_question")
async def delete_question_start(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q!")
        return
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👶 7-10 yosh", callback_data="delete_age_7_10")],
            [InlineKeyboardButton(text="🧒 11-14 yosh", callback_data="delete_age_11_14")]
        ]
    )
    
    await callback.message.edit_text(
        "❌ **Savol o'chirish**\n\n"
        "Qaysi yosh toifasidan savol o'chirmoqchisiz?",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("delete_age_"))
async def show_questions_for_delete(callback: CallbackQuery, state: FSMContext):
    age_group = "7-10" if "7_10" in callback.data else "11-14"
    
    # Load questions from database
    try:
        questions_list = DatabaseService.get_questions_by_age_group(age_group)
        if not questions_list:
            await callback.message.edit_text(
                f"❌ {age_group} yosh toifasi uchun savollar yo'q!",
                parse_mode="Markdown"
            )
            await callback.answer()
            return
    except Exception as e:
        logger.error(f"Error loading questions for deletion: {e}")
        await callback.message.edit_text("❌ Savollarni yuklashda xatolik!")
        await callback.answer()
        return
    
    text = f"❌ **Savol o'chirish** ({age_group} yosh)\n\n"
    text += "O'chirmoqchi bo'lgan savol raqamini kiriting:\n\n"
    
    for i, q in enumerate(questions_list[:10], 1):  # Birinchi 10 ta
        text += f"**{i}.** {q.question_text[:50]}...\n"
    
    if len(questions_list) > 10:
        text += f"\n... va yana {len(questions_list) - 10} ta savol"
    
    await state.update_data(delete_age_group=age_group)
    await callback.message.edit_text(text, parse_mode="Markdown")
    await state.set_state(AdminStates.waiting_for_delete_question_id)
    await callback.answer()

@dp.message(AdminStates.waiting_for_delete_question_id)
async def delete_question_by_id(message: Message, state: FSMContext):
    try:
        question_index = int(message.text) - 1
        data = await state.get_data()
        age_group = data["delete_age_group"]
        
        # Get questions from database
        questions_list = DatabaseService.get_questions_by_age_group(age_group)
        
        if question_index < 0 or question_index >= len(questions_list):
            await message.answer("❌ Noto'g'ri savol raqami!")
            return
        
        # Get the question to delete
        question_to_delete = questions_list[question_index]
        
        # Delete from database
        success = DatabaseService.delete_question(question_to_delete.id)
        
        if success:
            # Update global test_questions
            global test_questions
            test_questions = load_questions_from_db()
            
            remaining_questions = DatabaseService.get_questions_by_age_group(age_group)
            
            await message.answer(
                f"✅ **Savol o'chirildi!**\n\n"
                f"**O'chirilgan savol:** {question_to_delete.question_text}\n"
                f"**Qolgan savollar:** {len(remaining_questions)} ta",
                parse_mode="Markdown"
            )
        else:
            await message.answer("❌ Savolni o'chirishda xatolik yuz berdi!")
            
        await state.clear()
        
    except ValueError:
        await message.answer("❌ Iltimos, faqat raqam kiriting!")
    except Exception as e:
        logger.error(f"Error deleting question: {e}")
        await message.answer("❌ Savolni o'chirishda xatolik yuz berdi!")

@dp.callback_query(F.data == "view_questions")
async def view_questions(callback: CallbackQuery):
    text = "📊 **Barcha savollar:**\n\n"
    
    for age_group in ["7-10", "11-14"]:
        try:
            q_list = DatabaseService.get_questions_by_age_group(age_group)
            text += f"**{age_group} yosh toifasi:** {len(q_list)} ta savol\n"
            for i, q in enumerate(q_list[:3], 1):  # Birinchi 3 tani ko'rsatish
                text += f"  {i}. {q.question_text[:40]}...\n"
            if len(q_list) > 3:
                text += f"  ... va yana {len(q_list) - 3} ta\n"
            text += "\n"
        except Exception as e:
            logger.error(f"Error loading questions for {age_group}: {e}")
            text += f"**{age_group} yosh toifasi:** Xatolik\n\n"
    
    await callback.message.edit_text(text, parse_mode="Markdown")
    await callback.answer()

@dp.callback_query(F.data == "upload_questions")
async def upload_questions_start(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q!")
        return
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👶 7-10 yosh", callback_data="upload_age_7_10")],
            [InlineKeyboardButton(text="🧒 11-14 yosh", callback_data="upload_age_11_14")]
        ]
    )
    
    await callback.message.edit_text(
        "📁 **Fayl orqali savollar yuklash**\n\n"
        "Qaysi yosh toifasi uchun savollar yuklaysiz?\n\n"
        "**Qo'llab-quvvatlanadigan formatlar:**\n"
        "• Text fayl (.txt)\n"
        "• Word hujjat (.docx)\n"
        "• Excel jadval (.xlsx)\n\n"
        "**Talab qilinadigan format:**\n"
        "```\n"
        "1. Savol matni?\n"
        "A) Variant 1\n"
        "B) Variant 2\n"
        "C) Variant 3\n"
        "D) Variant 4\n"
        "Javob: C\n\n"
        "2. Keyingi savol...\n"
        "```",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.in_(["upload_age_7_10", "upload_age_11_14"]))
async def upload_age_selected(callback: CallbackQuery, state: FSMContext):
    age_group = "7-10" if "7_10" in callback.data else "11-14"
    await state.update_data(upload_age_group=age_group)
    
    await callback.message.edit_text(
        f"📁 **Fayl yuklash** ({age_group} yosh)\n\n"
        "Savollar fayli yuboring (txt, docx, xlsx):",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_question_file)
    await callback.answer()

@dp.message(AdminStates.waiting_for_question_file)
async def process_uploaded_file(message: Message, state: FSMContext):
    if not message.document:
        await message.answer("❌ Iltimos, fayl yuboring!")
        return
    
    file_name = message.document.file_name.lower()
    if not any(file_name.endswith(ext) for ext in ['.txt', '.docx', '.xlsx']):
        await message.answer("❌ Faqat .txt, .docx yoki .xlsx fayllarni yuklang!")
        return
    
    try:
        # Download file
        file = await bot.get_file(message.document.file_id)
        file_content = await bot.download_file(file.file_path)
        
        # Save temporarily
        temp_file = f"temp_{message.document.file_name}"
        with open(temp_file, 'wb') as f:
            f.write(file_content.read())
        
        # Parse file based on extension
        questions_text = ""
        
        if file_name.endswith('.txt'):
            with open(temp_file, 'r', encoding='utf-8') as f:
                questions_text = f.read()
                
        elif file_name.endswith('.docx'):
            try:
                import docx
                doc = docx.Document(temp_file)
                questions_text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
            except ImportError:
                await message.answer("❌ Word fayl o'qish uchun kutubxona o'rnatilmagan!")
                os.remove(temp_file)
                return
                
        elif file_name.endswith('.xlsx'):
            if not openpyxl:
                await message.answer("❌ Excel fayl o'qish uchun kutubxona o'rnatilmagan!")
                os.remove(temp_file)
                return
            
            wb = openpyxl.load_workbook(temp_file)
            ws = wb.active
            questions_text = ""
            for row in ws.iter_rows(values_only=True):
                if any(cell for cell in row):
                    questions_text += ' '.join([str(cell) if cell else '' for cell in row]) + '\n'
        
        # Parse questions from text
        parsed_questions = parse_questions_from_text(questions_text)
        
        if not parsed_questions:
            await message.answer(
                "❌ **Savollar topilmadi!**\n\n"
                "Fayl quyidagi formatda bo'lishi kerak:\n"
                "```\n"
                "1. Savol matni?\n"
                "A) Variant 1\n"
                "B) Variant 2\n"
                "C) Variant 3\n"
                "D) Variant 4\n"
                "Javob: C\n"
                "```",
                parse_mode="Markdown"
            )
            os.remove(temp_file)
            return
        
        # Save questions
        data = await state.get_data()
        age_group = data["upload_age_group"]
        
        questions = load_data(QUESTIONS_FILE, {})
        if age_group not in questions:
            questions[age_group] = []
        
        questions[age_group].extend(parsed_questions)
        save_data(QUESTIONS_FILE, questions)
        
        await message.answer(
            f"✅ **Savollar muvaffaqiyatli yuklandi!**\n\n"
            f"**Yosh toifasi:** {age_group}\n"
            f"**Yuklangan savollar:** {len(parsed_questions)} ta\n"
            f"**Jami savollar:** {len(questions[age_group])} ta",
            parse_mode="Markdown"
        )
        
        # Clean up
        os.remove(temp_file)
        await state.clear()
        
    except Exception as e:
        await message.answer(f"❌ Fayl ishlov berishda xatolik: {str(e)}")
        try:
            os.remove(temp_file)
        except:
            pass
        await state.clear()

def parse_questions_from_text(text):
    """Matndan savollarni ajratib olish"""
    questions = []
    lines = text.split('\n')
    
    i = 0
    while i < len(lines):
        # Find question start (number followed by dot)
        line = lines[i].strip()
        if not line:
            i += 1
            continue
            
        # Check if it's a question (starts with number)
        import re
        question_match = re.match(r'^\d+\.?\s*(.+)', line)
        if not question_match:
            i += 1
            continue
            
        question_text = question_match.group(1).strip()
        options = []
        correct_answer = None
        
        # Look for options A, B, C, D
        i += 1
        while i < len(lines):
            line = lines[i].strip()
            if line.startswith('A)'):
                options.append(line[2:].strip())
            elif line.startswith('B)'):
                options.append(line[2:].strip())
            elif line.startswith('C)'):
                options.append(line[2:].strip())
            elif line.startswith('D)'):
                options.append(line[2:].strip())
            elif line.startswith('Javob:'):
                answer_letter = line[6:].strip().upper()
                if answer_letter in ['A', 'B', 'C', 'D']:
                    correct_answer = ord(answer_letter) - ord('A')
                i += 1
                break
            elif re.match(r'^\d+\.', line):
                # Next question found, break
                break
            i += 1
        
        # Validate and add question
        if question_text and len(options) == 4 and correct_answer is not None:
            questions.append({
                "question": question_text,
                "options": options,
                "correct": correct_answer
            })
    
    return questions

@dp.callback_query(F.data == "bulk_add_questions")
async def bulk_add_questions_start(callback: CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q!")
        return
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👶 7-10 yosh", callback_data="bulk_age_7_10")],
            [InlineKeyboardButton(text="🧒 11-14 yosh", callback_data="bulk_age_11_14")]
        ]
    )
    
    await callback.message.edit_text(
        "📝 **Ko'p savol qo'shish**\n\n"
        "Qaysi yosh toifasi uchun savollar qo'shasiz?\n\n"
        "**Format:**\n"
        "```\n"
        "1. Birinchi savol matni?\n"
        "A) Variant 1\n"
        "B) Variant 2\n"
        "C) Variant 3\n"
        "D) Variant 4\n"
        "Javob: C\n\n"
        "2. Ikkinchi savol matni?\n"
        "A) Variant 1\n"
        "B) Variant 2\n"
        "C) Variant 3\n"
        "D) Variant 4\n"
        "Javob: A\n"
        "```",
        reply_markup=kb,
        parse_mode="Markdown"
    )
    await callback.answer()

@dp.callback_query(F.data.in_(["bulk_age_7_10", "bulk_age_11_14"]))
async def bulk_age_selected(callback: CallbackQuery, state: FSMContext):
    age_group = "7-10" if "7_10" in callback.data else "11-14"
    await state.update_data(bulk_age_group=age_group)
    
    await callback.message.edit_text(
        f"📝 **Ko'p savol qo'shish** ({age_group} yosh)\n\n"
        "Barcha savollarni quyidagi formatda yuboring:\n\n"
        "```\n"
        "1. Savol matni?\n"
        "A) Variant 1\n"
        "B) Variant 2\n"
        "C) Variant 3\n"
        "D) Variant 4\n"
        "Javob: C\n\n"
        "2. Keyingi savol?\n"
        "A) Variant A\n"
        "B) Variant B\n"
        "C) Variant C\n"
        "D) Variant D\n"
        "Javob: A\n"
        "```\n\n"
        "Har bir savol orasida bo'sh qator qoldiring.",
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_bulk_questions)
    await callback.answer()

@dp.message(AdminStates.waiting_for_bulk_questions)
async def process_bulk_questions(message: Message, state: FSMContext):
    data = await state.get_data()
    age_group = data["bulk_age_group"]
    
    # Parse multiple questions from text
    parsed_questions = parse_questions_from_text(message.text)
    
    if not parsed_questions:
        await message.answer(
            "❌ **Savollar topilmadi!**\n\n"
            "To'g'ri formatda kiriting:\n"
            "```\n"
            "1. Savol matni?\n"
            "A) Variant 1\n"
            "B) Variant 2\n"
            "C) Variant 3\n"
            "D) Variant 4\n"
            "Javob: C\n"
            "```",
            parse_mode="Markdown"
        )
        return
    
    # Save questions
    # Save questions to database
    for question_data in parsed_questions:
        question_to_save = {
            "question_text": question_data["question"],
            "age_group": age_group,
            "options": question_data["options"],
            "correct_answer": question_data["correct"]
        }
        save_question_to_db(question_to_save)
    
    # Update global test_questions
    global test_questions
    test_questions = load_questions_from_db()
    
    await message.answer(
        f"✅ **Savollar muvaffaqiyatli qo'shildi!**\n\n"
        f"**Yosh toifasi:** {age_group}\n"
        f"**Qo'shilgan savollar:** {len(parsed_questions)} ta\n"
        f"**Jami savollar:** {len(questions[age_group])} ta\n\n"
        f"**Qo'shilgan savollar:**\n" +
        "\n".join([f"{i+1}. {q['question'][:50]}..." for i, q in enumerate(parsed_questions[:5])]) +
        (f"\n... va yana {len(parsed_questions)-5} ta" if len(parsed_questions) > 5 else ""),
        parse_mode="Markdown"
    )
    await state.clear()

@dp.message(F.text == "📢 Xabar yuborish")
async def broadcast_start(message: Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    await message.answer(
        "📢 **Umumiy xabar yuborish**\n\n"
        "Barcha foydalanuvchilarga yubormoqchi bo'lgan xabaringizni yozing:",
        reply_markup=ReplyKeyboardRemove(),
        parse_mode="Markdown"
    )
    await state.set_state(AdminStates.waiting_for_broadcast)

@dp.message(AdminStates.waiting_for_broadcast)
async def send_broadcast(message: Message, state: FSMContext):
    user_id = message.from_user.id
    if not is_admin(user_id):
        return
    
    broadcast_text = message.text
    
    try:
        users = DatabaseService.get_all_users()
        
        if not users:
            await message.answer("❌ Foydalanuvchilar yo'q!")
            await state.clear()
            return
        
        success_count = 0
        fail_count = 0
        
        await message.answer(f"📤 Xabar yuborilmoqda... {len(users)} ta foydalanuvchiga")
        
        for user in users:
            try:
                await bot.send_message(int(user.user_id), f"📢 **E'lon:**\n\n{broadcast_text}", parse_mode="Markdown")
                success_count += 1
            except:
                fail_count += 1
            
            # Spamdan himoyalanish uchun
            await asyncio.sleep(0.1)
        
        kb = get_admin_keyboard()
        await message.answer(
            f"✅ Xabar yuborish yakunlandi!\n\n"
            f"📤 Muvaffaqiyatli: {success_count}\n"
            f"❌ Xatolik: {fail_count}",
            reply_markup=kb
        )
        await state.clear()
    except Exception as e:
        logger.error(f"Error in broadcast: {e}")
        await message.answer("❌ Xabar yuborishda xatolik yuz berdi!")
        await state.clear()

@dp.message(F.text == "📤 Ma'lumot eksport")
async def export_data(message: Message):
    user_id = message.from_user.id
    if not is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    try:
        users = DatabaseService.get_all_users()
        
        if not users:
            await message.answer("❌ Ma'lumotlar yo'q!")
            return
    except Exception as e:
        logger.error(f"Error loading users for export: {e}")
        await message.answer("❌ Foydalanuvchilar ma'lumotlarini yuklashda xatolik!")
        return
    
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📊 Excel", callback_data="export_excel")],
            [InlineKeyboardButton(text="📄 PDF", callback_data="export_pdf")]
        ]
    )
    
    await message.answer(
        "📤 **Ma'lumot eksport**\n\nQaysi formatda eksport qilmoqchisiz?",
        reply_markup=kb,
        parse_mode="Markdown"
    )

@dp.callback_query(F.data == "export_excel")
async def export_excel(callback: CallbackQuery):
    user_id = callback.from_user.id
    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q!")
        return
    
    if not openpyxl:
        await callback.answer("❌ Excel kutubxonasi o'rnatilmagan!")
        return
    
    await callback.answer("📊 Excel fayl tayyorlanmoqda...")
    
    try:
        users = DatabaseService.get_all_users()
        
        if not users:
            await callback.answer("❌ Ma'lumotlar yo'q!")
            return
    except Exception as e:
        logger.error(f"Error loading users for Excel export: {e}")
        await callback.answer("❌ Foydalanuvchilar ma'lumotlarini yuklashda xatolik!")
        return
    
    # Excel fayl yaratish
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    
    # Sarlavhalar
    headers = [
        "ID", "Bolaning ismi", "Ota-onaning ismi", "Telefon", "Yosh toifa",
        "Viloyat", "Tuman", "Mahalla", "Ro'yxatdan o'tgan sana", "Testlar soni"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Ma'lumotlar
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.user_id)
        ws.cell(row=row, column=2, value=user.child_name)
        ws.cell(row=row, column=3, value=user.parent_name)
        ws.cell(row=row, column=4, value=user.phone)
        ws.cell(row=row, column=5, value=user.age_group)
        ws.cell(row=row, column=6, value=user.region)
        ws.cell(row=row, column=7, value=user.district)
        ws.cell(row=row, column=8, value=user.mahalla)
        
        reg_date = ''
        if user.registration_date:
            reg_date = user.registration_date.strftime('%d.%m.%Y %H:%M')
        ws.cell(row=row, column=9, value=reg_date)
        
        ws.cell(row=row, column=10, value=len(user.test_results))
    
    # Ustunlar kengligini sozlash
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Faylni saqlash
    filename = f"kitobxon_kids_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    # Yuborish
    with open(filename, 'rb') as file:
        await callback.message.answer_document(
            BufferedInputFile(file.read(), filename),
            caption=f"📊 Foydalanuvchilar ma'lumotlari\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
    
    # Faylni o'chirish
    os.remove(filename)

@dp.callback_query(F.data == "export_pdf")
async def export_pdf(callback: CallbackQuery):
    user_id = callback.from_user.id
    if not is_admin(user_id):
        await callback.answer("❌ Ruxsat yo'q!")
        return
    
    if not SimpleDocTemplate:
        await callback.answer("❌ PDF kutubxonasi o'rnatilmagan!")
        return
    
    await callback.answer("📄 PDF fayl tayyorlanmoqda...")
    
    users = load_data(USERS_FILE, {})
    
    filename = f"kitobxon_kids_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=50, bottomMargin=50)
    
    # Style tanqislamasi
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontSize = 18
    title_style.textColor = colors.darkblue
    
    # Elementlar ro'yxati
    elements = []
    
    # Sarlavha
    title = Paragraph("📚 KITOBXON KIDS - FOYDALANUVCHILAR RO'YXATI", title_style)
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Sana va umumiy ma'lumotlar
    date_info = Paragraph(
        f"📅 Eksport sanasi: {datetime.now().strftime('%d.%m.%Y %H:%M')}<br/>"
        f"👥 Jami foydalanuvchilar: {len(users)} kishi<br/><br/>",
        styles['Normal']
    )
    elements.append(date_info)
    
    # Viloyat bo'yicha statistika
    region_stats = {}
    for user_info in users.values():
        region = user_info.get("region", "Noma'lum")
        region_stats[region] = region_stats.get(region, 0) + 1
    
    stats_text = "📊 VILOYATLAR BO'YICHA STATISTIKA:<br/>"
    for region, count in sorted(region_stats.items()):
        stats_text += f"• {region}: {count} kishi<br/>"
    
    stats_para = Paragraph(stats_text + "<br/>", styles['Normal'])
    elements.append(stats_para)
    
    # Load users from database
    try:
        users = DatabaseService.get_all_users()
        if not users:
            await callback.answer("❌ Ma'lumotlar yo'q!")
            return
    except Exception as e:
        logger.error(f"Error loading users for PDF export: {e}")
        await callback.answer("❌ Foydalanuvchilar ma'lumotlarini yuklashda xatolik!")
        return
    
    # Batafsil jadval uchun to'liq ma'lumotlar
    data = [["№", "Bolaning to'liq ismi", "Ota-onaning ismi", "Telefon raqami", "Yosh toifasi", "Viloyat", "Tuman/Shahar", "Mahalla", "Ro'yxatdan o'tgan sana", "Testlar soni"]]
    
    for i, user in enumerate(users, 1):
        reg_date = ''
        if user.registration_date:
            try:
                reg_date = user.registration_date.strftime('%d.%m.%Y')
            except:
                reg_date = 'Noma\'lum'
        
        data.append([
            str(i),
            user.child_name or 'Kiritilmagan',
            user.parent_name or 'Kiritilmagan',
            user.phone or 'Noma\'lum',
            user.age_group or 'Tanlanmagan',
            user.region or 'Kiritilmagan',
            user.district or 'Kiritilmagan',
            user.mahalla or 'Kiritilmagan',
            reg_date,
            str(len(user.test_results))
        ])
    
    # Jadval yaratish
    table = Table(data, colWidths=[25, 80, 80, 60, 50, 60, 60, 60, 55, 35])
    table.setStyle(TableStyle([
        # Sarlavha qatori
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Ma'lumotlar qatorlari
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        
        # Alternating row colors will be applied separately
    ]))
    
    # Juft va toq qatorlar uchun rang berish
    for i in range(2, len(data), 2):
        if i < len(data):
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, i), (-1, i), colors.white),
            ]))
    
    elements.append(table)
    
    # Oxirgi ma'lumot
    footer_text = f"<br/><br/>📊 Jami {len(users)} ta foydalanuvchi ro'yxati<br/>" \
                  f"🏆 Kitobxon Kids loyihasi - O'zbekiston bolalarining kitobxonlik tanlovchisi<br/>" \
                  f"📞 Qo'shimcha ma'lumot uchun adminlar bilan bog'laning"
    
    footer_para = Paragraph(footer_text, styles['Normal'])
    elements.append(footer_para)
    
    # PDF yaratish
    doc.build(elements)
    
    # Yuborish
    with open(filename, 'rb') as file:
        await callback.message.answer_document(
            BufferedInputFile(file.read(), filename),
            caption=f"📄 Foydalanuvchilar ma'lumotlari\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
    
    # Faylni o'chirish
    os.remove(filename)

# 📌 Noma'lum xabarlar
@dp.message()
async def unknown_message(message: Message):
    user_id = message.from_user.id
    user_info = get_user_info(user_id)
    
    if not user_info.get("registered"):
        await message.answer(
            "❓ Noma'lum buyruq. Avval ro'yxatdan o'ting!\n"
            "/start - Botni ishga tushirish"
        )
    else:
        kb = get_admin_keyboard() if is_admin(user_id) else get_main_keyboard()
        await message.answer(
            "❓ Noma'lum buyruq. Tugmalardan foydalaning:",
            reply_markup=kb
        )

# 🔧 Botni ishga tushirish
async def main():
    logger.info("Bot ishga tushirilmoqda...")
    
    # Ma'lumotlarni boshlang'ich yuklash
    global regions_data, test_questions
    
    # Load from database or fallback to JSON files
    try:
        regions_data = DatabaseService.get_regions()
        if not regions_data:
            regions_data = load_data(REGIONS_FILE, regions_data)
    except Exception as e:
        logger.warning(f"Could not load regions from database, using JSON: {e}")
        regions_data = load_data(REGIONS_FILE, regions_data)
    
    try:
        test_questions = load_questions_from_db()
        if not test_questions or not any(test_questions.values()):
            test_questions = load_data(QUESTIONS_FILE, test_questions)
    except Exception as e:
        logger.warning(f"Could not load questions from database, using JSON: {e}")
        test_questions = load_data(QUESTIONS_FILE, test_questions)
    
    logger.info(f"Yuklanganlar - Viloyatlar: {len(regions_data)}, Savollar: {sum(len(q) for q in test_questions.values())}")
    
    try:
        await dp.start_polling(bot, skip_updates=True)
    except Exception as e:
        logger.error(f"Bot ishga tushirishda xatolik: {e}")

if __name__ == "__main__":
    asyncio.run(main())
