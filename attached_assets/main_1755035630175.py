#!/usr/bin/env python3
"""
Professional Telegram Educational Testing Bot - Kitobxon Kids
A comprehensive bot for conducting reading comprehension tests for children aged 7-14.

Features:
- User registration with regional data validation
- Test management system with age-based categorization
- Role-based admin panel (Super Admin, Regular Admin)
- Statistics and analytics dashboard
- Broadcasting system for admin communications
- Document generation (Excel/PDF exports)
- Comprehensive error handling and logging

Author: Refactored for professional standards
Date: August 2025
"""

import asyncio
import json
import logging
import os
import random
import re
import tempfile
import uuid
from contextlib import asynccontextmanager
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union
from io import BytesIO

# Core Telegram bot imports
from aiogram import Bot, Dispatcher, F, types
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    ReplyKeyboardMarkup,
)

# Optional document generation libraries
try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("Excel library not available. Install with: pip install openpyxl")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logging.warning("PDF library not available. Install with: pip install reportlab")

# Configure professional logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bot.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════════
# CONFIGURATION AND CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════════

class BotConfig:
    """Centralized configuration management"""
    
    # Bot credentials and settings
    TOKEN = os.getenv("BOT_TOKEN", "7570796885:AAFkj7iY05fQUG21015viY7Gy8ifXXcnOpA").strip()
    SUPER_ADMIN_ID = int(os.getenv("SUPER_ADMIN_ID", "6578706277"))
    SPECIAL_ADMIN_IDS = [6578706277, 7853664401]
    CHANNEL_USERNAME = "@Kitobxon_Kids"
    
    # Timezone settings
    UZBEKISTAN_TZ = timezone(timedelta(hours=5))
    
    # File system configuration
    DATA_DIR = Path("bot_data")
    USERS_FILE = DATA_DIR / "users.json"
    ADMINS_FILE = DATA_DIR / "admins.json" 
    TESTS_FILE = DATA_DIR / "tests.json"
    RESULTS_FILE = DATA_DIR / "results.json"
    BROADCASTS_FILE = DATA_DIR / "broadcasts.json"
    STATISTICS_FILE = DATA_DIR / "statistics.json"
    BOT_USERS_FILE = DATA_DIR / "bot_users.json"  # All users who started the bot
    
    # Validation patterns
    PHONE_PATTERN = re.compile(r'^998[0-9]{9}$')
    NAME_PATTERN = re.compile(r'^[a-zA-ZА-Яа-яҚқҒғҲҳҰұЎўЎӯӨө\s\-\']{2,50}$')
    
    # Test configuration
    MIN_QUESTIONS_PER_TEST = 5
    MAX_QUESTIONS_PER_TEST = 50
    TEST_TIMEOUT_MINUTES = 30
    
    # Pagination settings
    USERS_PER_PAGE = 10
    TESTS_PER_PAGE = 5
    
    @classmethod
    def ensure_data_dir(cls) -> None:
        """Ensure data directory exists"""
        cls.DATA_DIR.mkdir(exist_ok=True)
        logger.info(f"Data directory ensured: {cls.DATA_DIR}")
    
    @classmethod
    def get_uzbekistan_time(cls) -> datetime:
        """Get current time in Uzbekistan timezone"""
        return datetime.now(cls.UZBEKISTAN_TZ)

# Regional data for comprehensive Uzbekistan coverage
REGIONS = {
    "Toshkent shahri": [
        "Bektemir", "Chilonzor", "Mirzo Ulug'bek", "Mirobod", "Olmazor", 
        "Shayxontohur", "Sergeli", "Uchtepa", "Yashnobod", "Yakkasaroy", "Yunusobod"
    ],
    "Toshkent viloyati": [
        "Bekabad", "Bo'ka", "Bo'stonliq", "Chinoz", "Chirchiq", "Ohangaron", 
        "Oqqo'rg'on", "Parkent", "Piskent", "Quyichirchiq", "O'rtachirchiq", 
        "Yangiyo'l", "Toshkent", "Yuqorichirchiq", "Zangiota", "Nurafshon", 
        "Olmaliq", "Angren"
    ],
    "Andijon": [
        "Andijon shahri", "Asaka", "Baliqchi", "Bo'ston", "Buloqboshi", "Izboskan", 
        "Jalaquduq", "Marhamat", "Oltinko'l", "Paxtaobod", "Paytug'", "Qo'rg'ontepa", 
        "Shahriston", "Xo'jaobod"
    ],
    "Farg'ona": [
        "Beshariq", "Buvayda", "Dang'ara", "Farg'ona shahri", "Ferghana tumani", 
        "Furqat", "Qo'qon", "Quva", "Rishton", "So'x", "Toshloq", "Uchko'prik", 
        "Yozyovon", "Oltiariq"
    ],
    "Namangan": [
        "Chortoq", "Chust", "Kosonsoy", "Namangan shahri", "Norin", "Pop", 
        "To'raqo'rg'on", "Uychi", "Uchqo'rg'on", "Yangiqo'rg'on", "Yangihayot"
    ],
    "Samarqand": [
        "Bulung'ur", "Ishtixon", "Jomboy", "Kattakurgan", "Oqdaryo", "Payariq", 
        "Pastdarg'om", "Qo'shrabot", "Samarqand shahri", "Toyloq", "Urgut"
    ],
    "Buxoro": [
        "Buxoro shahri", "Buxoro tumani", "G'ijduvon", "Jondor", "Kogon", "Olot", 
        "Peshku", "Qorako'l", "Qorovulbozor", "Romitan", "Shofirkon", "Vobkent"
    ],
    "Jizzax": [
        "Baxmal", "Chiroqchi", "Do'stlik", "Forish", "G'allaorol", "Zafarobod", 
        "Zarbdor", "Zomin", "Zafar", "Yangiobod", "Jizzax shahri", "Mirzacho'l"
    ],
    "Navoiy": [
        "Bespah", "Karmana", "Konimex", "Navbahor", "Nurota", "Tomdi", "Xatirchi", 
        "Uchquduq", "Navoiy shahri", "Zarafshon"
    ],
    "Qashqadaryo": [
        "Chiroqchi", "G'uzor", "Qarshi", "Kitob", "Koson", "Mirishkor", "Muborak", 
        "Nishon", "Shahrisabz", "Dehqonobod", "Yakkabog'"
    ],
    "Surxondaryo": [
        "Angor", "Bandixon", "Denov", "Jarqo'rg'on", "Muzrabot", "Oltinsoy", 
        "Sariosiyo", "Sherobod", "Sho'rchi", "Termiz", "Uzun", "Boysun"
    ],
    "Sirdaryo": [
        "Guliston", "Guliston tumani", "Mirzaobod", "Oqoltin", "Sardoba", 
        "Sayxunobod", "Sirdaryo tumani", "Xovos", "Boyovut", "Yangiyer"
    ],
    "Xorazm": [
        "Bog'ot", "Gurlan", "Hazorasp", "Khiva", "Qo'shko'pir", "Shovot", 
        "Urganch tumani", "Xonqa", "Yangiariq", "Yangibozor", "Tuproqqal'a", 
        "Urganch shahri"
    ],
    "Qoraqalpog'iston": [
        "Amudaryo", "Beruniy", "Chimboy", "Ellikqala", "Kegeyli", "Mo'ynoq", 
        "Nukus", "Qanliko'l", "Qo'ng'irot", "Taxiatosh", "To'rtko'l", "Xo'jayli"
    ]
}

# ═══════════════════════════════════════════════════════════════════════════════════
# CUSTOM EXCEPTIONS
# ═══════════════════════════════════════════════════════════════════════════════════

class BotError(Exception):
    """Base exception for bot-specific errors"""
    pass

class DataError(BotError):
    """Data handling and file operation errors"""
    pass

class ValidationError(BotError):
    """Input validation and format errors"""
    pass

class PermissionError(BotError):
    """Access control and permission errors"""
    pass

class TestError(BotError):
    """Test management and execution errors"""
    pass

# ═══════════════════════════════════════════════════════════════════════════════════
# DATA MANAGEMENT LAYER
# ═══════════════════════════════════════════════════════════════════════════════════

class DataManager:
    """Professional data management with atomic operations and error handling"""
    
    def __init__(self):
        self.config = BotConfig()
        self.config.ensure_data_dir()
        self._initialize_data_files()
        logger.info("DataManager initialized successfully")
    
    def _initialize_data_files(self) -> None:
        """Initialize all data files with proper default structures"""
        default_data_structures = {
            self.config.USERS_FILE: {},
            self.config.ADMINS_FILE: {
                str(self.config.SUPER_ADMIN_ID): {
                    "role": "super_admin",
                    "added_by": "system",
                    "added_date": self.config.get_uzbekistan_time().isoformat()
                }
            },
            self.config.TESTS_FILE: {"7-10": {}, "11-14": {}},
            self.config.RESULTS_FILE: [],
            self.config.BROADCASTS_FILE: [],
            self.config.STATISTICS_FILE: {},
            self.config.BOT_USERS_FILE: {}  # All users who started the bot
        }
        
        for file_path, default_content in default_data_structures.items():
            if not file_path.exists():
                self._save_json_data(file_path, default_content)
                logger.info(f"Initialized data file: {file_path}")
    
    @asynccontextmanager
    async def _safe_file_operation(self, file_path: Path, operation: str):
        """Context manager for safe file operations with comprehensive error handling"""
        try:
            yield
        except (FileNotFoundError, PermissionError) as e:
            logger.error(f"File operation '{operation}' failed for {file_path}: {e}")
            raise DataError(f"Failed to {operation} data file: {e}")
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error in {file_path}: {e}")
            raise DataError(f"Invalid JSON data in file: {e}")
        except Exception as e:
            logger.error(f"Unexpected error during {operation} of {file_path}: {e}")
            raise DataError(f"Unexpected error: {e}")
    
    def _load_json_data(self, file_path: Path, default_data: Any = None) -> Any:
        """Load data from JSON file with comprehensive error handling"""
        if not file_path.exists():
            return default_data or {}
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read().strip()
                if not content:
                    logger.warning(f"Empty file found: {file_path}")
                    return default_data or {}
                return json.loads(content)
        except (json.JSONDecodeError, FileNotFoundError, PermissionError) as e:
            logger.error(f"Error loading {file_path}: {e}")
            return default_data or {}
    
    def _save_json_data(self, file_path: Path, data: Any) -> None:
        """Save data to JSON file with atomic operation to prevent corruption"""
        temp_file = file_path.with_suffix('.tmp')
        
        try:
            # Write to temporary file first
            with open(temp_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            # Atomic operation: replace original file
            temp_file.replace(file_path)
            logger.debug(f"Successfully saved data to {file_path}")
            
        except Exception as e:
            # Cleanup temporary file on error
            if temp_file.exists():
                temp_file.unlink()
            logger.error(f"Error saving {file_path}: {e}")
            raise DataError(f"Failed to save data: {e}")
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # USER MANAGEMENT METHODS
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_users(self) -> Dict[str, Dict]:
        """Retrieve all registered users"""
        return self._load_json_data(self.config.USERS_FILE, {})
    
    def get_user(self, user_id: Union[int, str]) -> Optional[Dict]:
        """Get specific user data by ID"""
        users = self.get_users()
        return users.get(str(user_id))
    
    def save_user(self, user_id: Union[int, str], user_data: Dict) -> None:
        """Save user data with validation"""
        if not isinstance(user_data, dict):
            raise ValidationError("User data must be a dictionary")
        
        # Validate required fields
        required_fields = ['child_name', 'parent_name', 'region', 'district', 'age', 'phone']
        for field in required_fields:
            if field not in user_data:
                raise ValidationError(f"Required field missing: {field}")
        
        users = self.get_users()
        user_data['registration_date'] = self.config.get_uzbekistan_time().isoformat()
        users[str(user_id)] = user_data
        self._save_json_data(self.config.USERS_FILE, users)
        logger.info(f"User {user_id} saved successfully")
    
    def is_user_registered(self, user_id: Union[int, str]) -> bool:
        """Check if user is registered in the system"""
        return str(user_id) in self.get_users()
    
    def get_user_count(self) -> int:
        """Get total number of registered users"""
        return len(self.get_users())
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # BOT USERS TRACKING (ALL USERS WHO STARTED THE BOT)
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_bot_users(self) -> Dict[str, Dict]:
        """Retrieve all users who have started the bot"""
        return self._load_json_data(self.config.BOT_USERS_FILE, {})
    
    def add_bot_user(self, user_id: Union[int, str], user_info: Dict) -> None:
        """Add user to bot users list (for broadcasting)"""
        bot_users = self.get_bot_users()
        user_id_str = str(user_id)
        
        if user_id_str not in bot_users:
            bot_users[user_id_str] = {
                "first_name": user_info.get('first_name', 'Unknown'),
                "username": user_info.get('username'),
                "first_interaction": self.config.get_uzbekistan_time().isoformat(),
                "is_active": True
            }
        else:
            # Update activity status
            bot_users[user_id_str]['last_interaction'] = self.config.get_uzbekistan_time().isoformat()
            bot_users[user_id_str]['is_active'] = True
        
        self._save_json_data(self.config.BOT_USERS_FILE, bot_users)
    
    def get_all_bot_users_for_broadcast(self) -> Dict[str, Dict]:
        """Get all bot users for broadcasting (registered + non-registered)"""
        return self.get_bot_users()
    
    def get_bot_users_count(self) -> int:
        """Get total number of users who have started the bot"""
        return len(self.get_bot_users())
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # ADMIN MANAGEMENT METHODS
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_admins(self) -> Dict[str, Dict]:
        """Retrieve all admin users"""
        return self._load_json_data(self.config.ADMINS_FILE, {})
    
    def save_admin(self, admin_id: Union[int, str], admin_data: Dict) -> None:
        """Save admin data with validation"""
        if not isinstance(admin_data, dict):
            raise ValidationError("Admin data must be a dictionary")
        
        admins = self.get_admins()
        admin_data['added_date'] = self.config.get_uzbekistan_time().isoformat()
        admins[str(admin_id)] = admin_data
        self._save_json_data(self.config.ADMINS_FILE, admins)
        logger.info(f"Admin {admin_id} saved successfully")
    
    def remove_admin(self, admin_id: Union[int, str]) -> bool:
        """Remove admin by ID"""
        try:
            admins = self.get_admins()
            admin_id_str = str(admin_id)
            
            if admin_id_str in admins:
                del admins[admin_id_str]
                self._save_json_data(self.config.ADMINS_FILE, admins)
                logger.info(f"Admin {admin_id} removed successfully")
                return True
            return False
        except Exception as e:
            logger.error(f"Error removing admin {admin_id}: {e}")
            return False
    
    def is_admin(self, user_id: Union[int, str]) -> bool:
        """Check if user has admin privileges"""
        return str(user_id) in self.get_admins()
    
    def is_super_admin(self, user_id: Union[int, str]) -> bool:
        """Check if user has super admin privileges"""
        admins = self.get_admins()
        admin_data = admins.get(str(user_id))
        return admin_data and admin_data.get("role") == "super_admin"
    
    def has_special_privileges(self, user_id: Union[int, str]) -> bool:
        """Check if user has hardcoded special admin privileges"""
        return int(user_id) in self.config.SPECIAL_ADMIN_IDS
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # TEST MANAGEMENT METHODS
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_tests(self) -> Dict[str, Dict]:
        """Retrieve all tests organized by age group"""
        return self._load_json_data(self.config.TESTS_FILE, {"7-10": {}, "11-14": {}})
    
    def save_test(self, test_data: Dict) -> str:
        """Save test data and return generated test ID"""
        if not isinstance(test_data, dict):
            raise ValidationError("Test data must be a dictionary")
        
        # Validate required fields
        required_fields = ['age_group', 'book_name', 'questions']
        for field in required_fields:
            if field not in test_data:
                raise ValidationError(f"Required test field missing: {field}")
        
        age_group = test_data.get("age_group")
        if age_group not in ["7-10", "11-14"]:
            raise ValidationError("Invalid age group. Must be '7-10' or '11-14'")
        
        # Validate questions structure
        questions = test_data.get("questions", [])
        if len(questions) < self.config.MIN_QUESTIONS_PER_TEST:
            raise ValidationError(f"Test must have at least {self.config.MIN_QUESTIONS_PER_TEST} questions")
        
        tests = self.get_tests()
        test_id = str(uuid.uuid4())
        
        if age_group not in tests:
            tests[age_group] = {}
        
        test_data['created_date'] = self.config.get_uzbekistan_time().isoformat()
        tests[age_group][test_id] = test_data
        self._save_json_data(self.config.TESTS_FILE, tests)
        logger.info(f"Test {test_id} saved successfully for age group {age_group}")
        return test_id
    
    def delete_test(self, age_group: str, test_id: str) -> bool:
        """Delete a specific test"""
        try:
            tests = self.get_tests()
            if age_group in tests and test_id in tests[age_group]:
                del tests[age_group][test_id]
                self._save_json_data(self.config.TESTS_FILE, tests)
                logger.info(f"Test {test_id} deleted successfully from age group {age_group}")
                return True
            return False
        except Exception as e:
            logger.error(f"Error deleting test {test_id}: {e}")
            return False
    
    def get_test_by_id(self, age_group: str, test_id: str) -> Optional[Dict]:
        """Get a specific test by ID and age group"""
        tests = self.get_tests()
        return tests.get(age_group, {}).get(test_id)
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # RESULTS MANAGEMENT METHODS
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_results(self) -> List[Dict]:
        """Retrieve all test results"""
        return self._load_json_data(self.config.RESULTS_FILE, [])
    
    def save_result(self, result_data: Dict) -> None:
        """Save test result with validation"""
        if not isinstance(result_data, dict):
            raise ValidationError("Result data must be a dictionary")
        
        # Validate required fields
        required_fields = ['user_id', 'test_id', 'score', 'total_questions', 'correct_answers']
        for field in required_fields:
            if field not in result_data:
                raise ValidationError(f"Required result field missing: {field}")
        
        results = self.get_results()
        result_data['completion_date'] = self.config.get_uzbekistan_time().isoformat()
        results.append(result_data)
        self._save_json_data(self.config.RESULTS_FILE, results)
        logger.info(f"Result saved for user {result_data.get('user_id')}")
    
    def get_user_results(self, user_id: Union[int, str]) -> List[Dict]:
        """Get all results for a specific user"""
        results = self.get_results()
        return [r for r in results if str(r.get('user_id')) == str(user_id)]
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # BROADCAST MANAGEMENT METHODS
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_broadcasts(self) -> List[Dict]:
        """Retrieve broadcast history"""
        return self._load_json_data(self.config.BROADCASTS_FILE, [])
    
    def save_broadcast(self, broadcast_data: Dict) -> None:
        """Save broadcast information"""
        if not isinstance(broadcast_data, dict):
            raise ValidationError("Broadcast data must be a dictionary")
        
        broadcasts = self.get_broadcasts()
        broadcast_data['broadcast_date'] = self.config.get_uzbekistan_time().isoformat()
        broadcasts.append(broadcast_data)
        self._save_json_data(self.config.BROADCASTS_FILE, broadcasts)
        logger.info(f"Broadcast saved by admin {broadcast_data.get('admin_id')}")
    
    # ═══════════════════════════════════════════════════════════════════════════════════
    # STATISTICS MANAGEMENT METHODS
    # ═══════════════════════════════════════════════════════════════════════════════════
    
    def get_statistics(self) -> Dict:
        """Get current statistics data"""
        return self._load_json_data(self.config.STATISTICS_FILE, {})
    
    def update_statistics(self) -> None:
        """Update comprehensive system statistics"""
        try:
            users = self.get_users()
            results = self.get_results()
            
            # Regional statistics calculation
            regional_stats = {}
            for region, districts in REGIONS.items():
                regional_stats[region] = {
                    "total_users": 0,
                    "districts": {district: 0 for district in districts}
                }
            
            # Count users by region and district
            for user_data in users.values():
                region = user_data.get("region", "Unknown")
                district = user_data.get("district", "Unknown")
                
                if region in regional_stats:
                    regional_stats[region]["total_users"] += 1
                    if district in regional_stats[region]["districts"]:
                        regional_stats[region]["districts"][district] += 1
            
            # Test performance statistics
            test_stats = {
                "total_tests_taken": len(results),
                "average_score": 0,
                "high_scorers_70plus": 0,
                "age_group_stats": {
                    "7-10": {"count": 0, "avg_score": 0},
                    "11-14": {"count": 0, "avg_score": 0}
                }
            }
            
            if results:
                total_score = sum(result.get("score", 0) for result in results)
                test_stats["average_score"] = round(total_score / len(results), 2)
                test_stats["high_scorers_70plus"] = len([r for r in results if r.get("score", 0) >= 70])
                
                # Age group specific statistics
                for age_range in ["7-10", "11-14"]:
                    age_results = [r for r in results if r.get("age_group") == age_range]
                    if age_results:
                        test_stats["age_group_stats"][age_range]["count"] = len(age_results)
                        test_stats["age_group_stats"][age_range]["avg_score"] = round(
                            sum(r.get("score", 0) for r in age_results) / len(age_results), 2
                        )
            
            # Top performers ranking
            ranking_data = []
            for result in results:
                ranking_data.append({
                    "user_name": result.get("user_name", "Unknown"),
                    "score": result.get("score", 0),
                    "percentage": result.get("percentage", 0),
                    "age": result.get("age", "Unknown"),
                    "region": result.get("region", "Unknown"),
                    "date": result.get("completion_date", "Unknown")
                })
            
            # Sort by score (descending) and limit to top 100
            ranking_data.sort(key=lambda x: x["score"], reverse=True)
            
            # Compile final statistics
            statistics = {
                "last_updated": self.config.get_uzbekistan_time().isoformat(),
                "total_registered_users": len(users),
                "regional_statistics": regional_stats,
                "test_statistics": test_stats,
                "top_performers": ranking_data[:100]
            }
            
            self._save_json_data(self.config.STATISTICS_FILE, statistics)
            logger.info("Statistics updated successfully")
            
        except Exception as e:
            logger.error(f"Error updating statistics: {e}")
            raise DataError(f"Failed to update statistics: {e}")

# ═══════════════════════════════════════════════════════════════════════════════════
# INPUT VALIDATION AND SANITIZATION
# ═══════════════════════════════════════════════════════════════════════════════════

class InputValidator:
    """Comprehensive input validation and sanitization"""
    
    @staticmethod
    def sanitize_input(text: str) -> str:
        """Sanitize user input to prevent issues"""
        if not isinstance(text, str):
            return ""
        
        # Remove excessive whitespace and dangerous characters
        sanitized = re.sub(r'\s+', ' ', text.strip())
        sanitized = re.sub(r'[<>"]', '', sanitized)
        return sanitized[:500]  # Limit length
    
    @staticmethod
    def validate_name(name: str) -> bool:
        """Validate name format"""
        if not name or len(name.strip()) < 2:
            return False
        return BotConfig.NAME_PATTERN.match(name.strip()) is not None
    
    @staticmethod
    def validate_phone(phone: str) -> bool:
        """Validate Uzbekistan phone number format"""
        if not phone:
            return False
        phone_clean = re.sub(r'[^\d]', '', phone)
        return BotConfig.PHONE_PATTERN.match(phone_clean) is not None
    
    @staticmethod
    def validate_age(age: str) -> bool:
        """Validate age is within acceptable range"""
        try:
            age_int = int(age)
            return 7 <= age_int <= 14
        except (ValueError, TypeError):
            return False
    
    @staticmethod
    def validate_region_district(region: str, district: str) -> bool:
        """Validate region and district combination"""
        return region in REGIONS and district in REGIONS[region]
    
    @staticmethod
    def parse_question(question_text: str) -> Optional[Dict]:
        """Parse question format for test creation"""
        try:
            lines = [line.strip() for line in question_text.strip().split('\n') if line.strip()]
            
            if len(lines) < 6:  # Question + 4 options + correct answer
                return None
            
            question = lines[0]
            options = {}
            correct_answer = None
            
            # Parse options (A, B, C, D)
            for i, line in enumerate(lines[1:5], 1):
                if line.startswith(('A)', 'A.', 'A:')):
                    options['A'] = line[2:].strip()
                elif line.startswith(('B)', 'B.', 'B:')):
                    options['B'] = line[2:].strip()
                elif line.startswith(('C)', 'C.', 'C:')):
                    options['C'] = line[2:].strip()
                elif line.startswith(('D)', 'D.', 'D:')):
                    options['D'] = line[2:].strip()
            
            # Find correct answer
            for line in lines[5:]:
                if line.upper().startswith('JAVOB:') or line.upper().startswith('ANSWER:'):
                    answer_part = line.split(':', 1)[1].strip().upper()
                    if answer_part in ['A', 'B', 'C', 'D']:
                        correct_answer = answer_part
                        break
            
            if len(options) == 4 and correct_answer:
                return {
                    "question": question,
                    "options": options,
                    "correct_answer": correct_answer
                }
            
            return None
            
        except Exception as e:
            logger.error(f"Error parsing question: {e}")
            return None

# ═══════════════════════════════════════════════════════════════════════════════════
# KEYBOARD BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════════

class KeyboardBuilder:
    """Professional keyboard layout management"""
    
    @staticmethod
    def get_main_menu(is_admin: bool = False) -> ReplyKeyboardMarkup:
        """Main menu keyboard based on user role"""
        keyboard = [
            [KeyboardButton(text="📋 Ro'yxatdan o'tish")],
            [KeyboardButton(text="📝 Test topshirish")],
            [KeyboardButton(text="💬 Fikr va maslahatlar")],
            [KeyboardButton(text="📚 Loyiha haqida")]
        ]
        
        if is_admin:
            keyboard.append([KeyboardButton(text="👨‍💼 Admin panel")])
        
        return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    @staticmethod
    def get_admin_menu(is_super: bool = False) -> ReplyKeyboardMarkup:
        """Admin menu keyboard based on admin level"""
        keyboard = [
            [KeyboardButton(text="👥 Foydalanuvchilar ro'yxati")],
            [KeyboardButton(text="➕ Test qo'shish")]
        ]
        
        if is_super:
            keyboard.extend([
                [KeyboardButton(text="👨‍💼 Adminlar ro'yxati")],
                [KeyboardButton(text="➕ Admin qo'shish"), KeyboardButton(text="➖ Admin o'chirish")],
                [KeyboardButton(text="🗑 Test o'chirish")],
                [KeyboardButton(text="📊 Statistika"), KeyboardButton(text="🏆 Reyting")],
                [KeyboardButton(text="📢 Umumiy xabar"), KeyboardButton(text="📄 Hisobot")]
            ])
        
        keyboard.append([KeyboardButton(text="🔙 Asosiy menyu")])
        return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    @staticmethod
    def get_regions_keyboard() -> ReplyKeyboardMarkup:
        """Keyboard for region selection"""
        keyboard = []
        regions = list(REGIONS.keys())
        
        # Create rows of 2 regions each
        for i in range(0, len(regions), 2):
            row = [KeyboardButton(text=regions[i])]
            if i + 1 < len(regions):
                row.append(KeyboardButton(text=regions[i + 1]))
            keyboard.append(row)
        
        keyboard.append([KeyboardButton(text="🔙 Orqaga")])
        return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    @staticmethod
    def get_districts_keyboard(region: str) -> ReplyKeyboardMarkup:
        """Keyboard for district selection based on region"""
        if region not in REGIONS:
            return KeyboardBuilder.get_regions_keyboard()
        
        districts = REGIONS[region]
        keyboard = []
        
        # Create rows of 2 districts each
        for i in range(0, len(districts), 2):
            row = [KeyboardButton(text=districts[i])]
            if i + 1 < len(districts):
                row.append(KeyboardButton(text=districts[i + 1]))
            keyboard.append(row)
        
        keyboard.append([KeyboardButton(text="🔙 Orqaga")])
        return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    @staticmethod
    def get_age_keyboard() -> ReplyKeyboardMarkup:
        """Keyboard for age selection"""
        keyboard = []
        current_row = []
        
        for i in range(7, 15):
            current_row.append(KeyboardButton(text=str(i)))
            if len(current_row) == 4:  # 4 ages per row
                keyboard.append(current_row)
                current_row = []
        
        # Add any remaining ages
        if current_row:
            keyboard.append(current_row)
        
        keyboard.append([KeyboardButton(text="🔙 Orqaga")])
        return ReplyKeyboardMarkup(keyboard=keyboard, resize_keyboard=True)
    
    @staticmethod
    def get_test_selection_keyboard(age_group: str, tests: Dict) -> InlineKeyboardMarkup:
        """Keyboard for test selection"""
        keyboard = []
        group_tests = tests.get(age_group, {})
        
        for test_id, test_data in group_tests.items():
            book_name = test_data.get('book_name', 'Unknown Book')
            question_count = len(test_data.get('questions', []))
            keyboard.append([
                InlineKeyboardButton(
                    text=f"📖 {book_name} ({question_count} ta savol)",
                    callback_data=f"start_test:{age_group}:{test_id}"
                )
            ])
        
        if not keyboard:
            keyboard.append([
                InlineKeyboardButton(text="❌ Testlar mavjud emas", callback_data="no_tests")
            ])
        
        return InlineKeyboardMarkup(inline_keyboard=keyboard)
    
    @staticmethod
    def get_question_keyboard() -> InlineKeyboardMarkup:
        """Keyboard for answering questions"""
        keyboard = [
            [
                InlineKeyboardButton(text="A", callback_data="answer:A"),
                InlineKeyboardButton(text="B", callback_data="answer:B")
            ],
            [
                InlineKeyboardButton(text="C", callback_data="answer:C"),
                InlineKeyboardButton(text="D", callback_data="answer:D")
            ]
        ]
        return InlineKeyboardMarkup(inline_keyboard=keyboard)
    
    @staticmethod
    def get_pagination_keyboard(page: int, total_pages: int, prefix: str) -> InlineKeyboardMarkup:
        """Generic pagination keyboard"""
        keyboard = []
        
        if total_pages > 1:
            nav_buttons = []
            if page > 0:
                nav_buttons.append(InlineKeyboardButton(text="◀️", callback_data=f"{prefix}:prev:{page}"))
            
            nav_buttons.append(InlineKeyboardButton(text=f"{page + 1}/{total_pages}", callback_data="page_info"))
            
            if page < total_pages - 1:
                nav_buttons.append(InlineKeyboardButton(text="▶️", callback_data=f"{prefix}:next:{page}"))
            
            keyboard.append(nav_buttons)
        
        return InlineKeyboardMarkup(inline_keyboard=keyboard)

# ═══════════════════════════════════════════════════════════════════════════════════
# DOCUMENT GENERATION SERVICES
# ═══════════════════════════════════════════════════════════════════════════════════

class DocumentGenerator:
    """Professional document generation service"""
    
    @staticmethod
    def generate_excel_report(users: Dict, results: List) -> Optional[BytesIO]:
        """Generate comprehensive Excel report with complete user and test data"""
        if not EXCEL_AVAILABLE:
            logger.warning("Excel library not available")
            return None
        
        try:
            output = BytesIO()
            workbook = openpyxl.Workbook()
            
            # ═══════════════════════════════════════════════════════════════════════════════
            # USERS SHEET - Complete user information
            # ═══════════════════════════════════════════════════════════════════════════════
            ws_users = workbook.active
            ws_users.title = "Foydalanuvchilar"
            
            # Enhanced headers with more complete information
            headers = [
                "№", "Foydalanuvchi ID", "Bola ismi", "Ota-ona ismi", "Viloyat", "Tuman", 
                "Mahalla", "Yosh", "Telefon raqami", "Ro'yxatdan o'tish sanasi", 
                "Ro'yxatdan o'tish vaqti", "Holati", "Oxirgi faollik"
            ]
            ws_users.append(headers)
            
            # Style headers with better formatting
            for col, header in enumerate(headers, 1):
                cell = ws_users.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # Add user data with complete information
            row_num = 2
            for user_id, user_data in users.items():
                registration_date = user_data.get('registration_date', '')
                if registration_date:
                    try:
                        # Parse and format date properly
                        date_obj = datetime.fromisoformat(registration_date.replace('Z', '+00:00'))
                        formatted_date = date_obj.strftime('%d.%m.%Y')
                        formatted_time = date_obj.strftime('%H:%M')
                    except:
                        formatted_date = registration_date[:10] if len(registration_date) >= 10 else registration_date
                        formatted_time = registration_date[11:16] if len(registration_date) > 16 else ""
                else:
                    formatted_date = "Ma'lum emas"
                    formatted_time = ""
                
                row_data = [
                    row_num - 1,  # Sequential number
                    user_id or "Ma'lum emas",
                    user_data.get('child_name', 'Kiritilmagan'),
                    user_data.get('parent_name', 'Kiritilmagan'),
                    user_data.get('region', 'Kiritilmagan'),
                    user_data.get('district', 'Kiritilmagan'),
                    user_data.get('mahalla', 'Kiritilmagan'),
                    str(user_data.get('age', 'Kiritilmagan')),
                    user_data.get('phone', 'Kiritilmagan'),
                    formatted_date,
                    formatted_time,
                    "Ro'yxatdan o'tgan",
                    "Faol" if user_data.get('is_active', True) else "Nofaol"
                ]
                ws_users.append(row_data)
                
                # Style data rows
                for col in range(1, len(headers) + 1):
                    cell = ws_users.cell(row=row_num, column=col)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                row_num += 1
            
            # Auto-adjust column widths
            for col in ws_users.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws_users.column_dimensions[column].width = adjusted_width
            
            # ═══════════════════════════════════════════════════════════════════════════════
            # RESULTS SHEET - Complete test results
            # ═══════════════════════════════════════════════════════════════════════════════
            ws_results = workbook.create_sheet("Test Natijalari")
            result_headers = [
                "№", "Foydalanuvchi ID", "Bola ismi", "Test ID", "Kitob nomi", 
                "Yosh guruhi", "Jami savollar", "To'g'ri javoblar", "Noto'g'ri javoblar",
                "Ball", "Foiz", "Baho", "Test sanasi", "Test vaqti", "Davomiyligi"
            ]
            ws_results.append(result_headers)
            
            # Style result headers
            for col, header in enumerate(result_headers, 1):
                cell = ws_results.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
            
            # Add results data with complete information
            if results:
                result_row_num = 2
                for result in results:
                    user_id = result.get('user_id', '')
                    user_data = users.get(str(user_id), {})
                    
                    # Parse completion date
                    completion_date = result.get('completion_date', '')
                    if completion_date:
                        try:
                            date_obj = datetime.fromisoformat(completion_date.replace('Z', '+00:00'))
                            formatted_date = date_obj.strftime('%d.%m.%Y')
                            formatted_time = date_obj.strftime('%H:%M')
                        except:
                            formatted_date = completion_date[:10] if len(completion_date) >= 10 else completion_date
                            formatted_time = completion_date[11:16] if len(completion_date) > 16 else ""
                    else:
                        formatted_date = "Ma'lum emas"
                        formatted_time = ""
                    
                    total_questions = result.get('total_questions', 0)
                    correct_answers = result.get('correct_answers', 0)
                    incorrect_answers = max(0, total_questions - correct_answers)
                    percentage = result.get('percentage', 0)
                    
                    # Determine grade based on percentage
                    if percentage >= 90:
                        grade = "A'lo (90-100%)"
                    elif percentage >= 80:
                        grade = "Yaxshi (80-89%)"
                    elif percentage >= 70:
                        grade = "Qoniqarli (70-79%)"
                    elif percentage >= 60:
                        grade = "O'rtacha (60-69%)"
                    else:
                        grade = "Qoniqarsiz (<60%)"
                    
                    result_row_data = [
                        result_row_num - 1,  # Sequential number
                        str(user_id) or "Ma'lum emas",
                        user_data.get('child_name', 'Ma\'lum emas'),
                        result.get('test_id', 'Ma\'lum emas'),
                        result.get('book_name', 'Ma\'lum emas'),
                        result.get('age_group', 'Ma\'lum emas'),
                        str(total_questions) if total_questions else "0",
                        str(correct_answers) if correct_answers else "0",
                        str(incorrect_answers),
                        str(result.get('score', '0')),
                        f"{percentage}%" if percentage else "0%",
                        grade,
                        formatted_date,
                        formatted_time,
                        result.get('duration', 'Ma\'lum emas')
                    ]
                    ws_results.append(result_row_data)
                    
                    # Style result rows
                    for col in range(1, len(result_headers) + 1):
                        cell = ws_results.cell(row=result_row_num, column=col)
                        cell.border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Color coding based on performance
                        if col == 11:  # Percentage column
                            if percentage >= 80:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif percentage >= 60:
                                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        elif result_row_num % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    
                    result_row_num += 1
            
            # Auto-adjust column widths for results sheet
            for col in ws_results.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                ws_results.column_dimensions[column].width = adjusted_width
            
            # ═══════════════════════════════════════════════════════════════════════════════
            # STATISTICS SHEET - Summary statistics
            # ═══════════════════════════════════════════════════════════════════════════════
            ws_stats = workbook.create_sheet("Statistika")
            
            # Calculate comprehensive statistics
            total_users = len(users)
            total_tests = len(results) if results else 0
            
            if results:
                scores = [r.get('percentage', 0) for r in results]
                avg_score = sum(scores) / len(scores) if scores else 0
                high_scorers = len([s for s in scores if s >= 80])
                low_scorers = len([s for s in scores if s < 60])
            else:
                avg_score = 0
                high_scorers = 0
                low_scorers = 0
            
            # Statistics data
            stats_data = [
                ["Statistika turi", "Qiymat", "Izoh"],
                ["Jami ro'yxatdan o'tganlar", total_users, "Botda ro'yxatdan o'tgan foydalanuvchilar soni"],
                ["Jami topshirilgan testlar", total_tests, "Barcha foydalanuvchilar tomonidan topshirilgan testlar soni"],
                ["O'rtacha natija", f"{avg_score:.1f}%", "Barcha testlar bo'yicha o'rtacha foiz natija"],
                ["Yaxshi natija (80%+)", high_scorers, "80% va undan yuqori natija olganlar soni"],
                ["Past natija (<60%)", low_scorers, "60% dan past natija olganlar soni"],
                ["Hisobot yaratilgan sana", datetime.now().strftime('%d.%m.%Y'), "Ushbu hisobotning yaratilgan sanasi"],
                ["Hisobot yaratilgan vaqt", datetime.now().strftime('%H:%M'), "Ushbu hisobotning yaratilgan vaqti"]
            ]
            
            # Add statistics data
            for row_data in stats_data:
                ws_stats.append(row_data)
            
            # Style statistics sheet
            for row_num in range(1, len(stats_data) + 1):
                for col_num in range(1, 4):
                    cell = ws_stats.cell(row=row_num, column=col_num)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    
                    if row_num == 1:  # Header row
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left' if col_num == 3 else 'center', vertical='center')
                        if row_num % 2 == 0:
                            cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            # Auto-adjust column widths for statistics sheet
            for col in ws_stats.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_stats.column_dimensions[column].width = adjusted_width
            
            workbook.save(output)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error generating comprehensive Excel report: {e}")
            return None
    
    @staticmethod
    def generate_pdf_report(statistics: Dict) -> Optional[BytesIO]:
        """Generate PDF statistical report"""
        if not PDF_AVAILABLE:
            logger.warning("PDF library not available")
            return None
        
        try:
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            story = []
            
            styles = getSampleStyleSheet()
            title_style = styles['Title']
            normal_style = styles['Normal']
            
            # Title
            story.append(Paragraph("Kitobxon Kids - Statistik Hisobot", title_style))
            story.append(Paragraph(f"Yaratilgan: {BotConfig.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}", normal_style))
            
            # Statistics table
            test_stats = statistics.get('test_statistics', {})
            data = [
                ['Parametr', 'Qiymat'],
                ['Jami foydalanuvchilar', statistics.get('total_registered_users', 0)],
                ['Jami testlar', test_stats.get('total_tests_taken', 0)],
                ['O\'rtacha ball', f"{test_stats.get('average_score', 0)}%"],
                ['70+ ball olganlar', test_stats.get('high_scorers_70plus', 0)]
            ]
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            doc.build(story)
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            return None

# ═══════════════════════════════════════════════════════════════════════════════════
# FSM STATES DEFINITION
# ═══════════════════════════════════════════════════════════════════════════════════

class RegistrationStates(StatesGroup):
    """User registration flow states"""
    check_subscription = State()
    child_name = State()
    parent_name = State()
    region = State()
    district = State()
    mahalla = State()
    age = State()
    phone = State()
    feedback = State()

class AdminStates(StatesGroup):
    """Admin panel operation states"""
    add_admin = State()
    remove_admin = State()
    add_test_age = State()
    add_test_book = State()
    add_test_questions = State()
    delete_test_age = State()
    delete_test_select = State()
    broadcast_message = State()
    broadcast_confirm = State()

class TestStates(StatesGroup):
    """Test taking flow states"""
    select_age_group = State()
    taking_test = State()
    test_question = State()

# ═══════════════════════════════════════════════════════════════════════════════════
# GLOBAL INSTANCES
# ═══════════════════════════════════════════════════════════════════════════════════

# Initialize bot configuration and data manager
config = BotConfig()
data_manager = DataManager()
input_validator = InputValidator()

# Initialize bot and dispatcher
bot = Bot(
    token=config.TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher()

# ═══════════════════════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════════

async def check_subscription(user_id: int) -> bool:
    """Check if user is subscribed to the channel"""
    try:
        member = await bot.get_chat_member(config.CHANNEL_USERNAME, user_id)
        return member.status in ['member', 'administrator', 'creator']
    except Exception as e:
        logger.warning(f"Could not check subscription for user {user_id}: {e}")
        return True  # Allow access if check fails

def calculate_test_score(correct_answers: int, total_questions: int) -> Tuple[int, str]:
    """Calculate test score and performance level"""
    if total_questions == 0:
        return 0, "Noaniq"
    
    percentage = round((correct_answers / total_questions) * 100)
    
    if percentage >= 90:
        performance = "A'lo"
    elif percentage >= 80:
        performance = "Yaxshi"
    elif percentage >= 70:
        performance = "Qoniqarli"
    elif percentage >= 60:
        performance = "O'rtacha"
    else:
        performance = "Qoniqarsiz"
    
    return percentage, performance

async def send_admin_notification(message: str, exclude_id: Optional[int] = None):
    """Send notification to all admins"""
    try:
        admins = data_manager.get_admins()
        for admin_id in admins.keys():
            if exclude_id and int(admin_id) == exclude_id:
                continue
            try:
                await bot.send_message(int(admin_id), message)
            except Exception as e:
                logger.warning(f"Could not send notification to admin {admin_id}: {e}")
    except Exception as e:
        logger.error(f"Error sending admin notifications: {e}")

# ═══════════════════════════════════════════════════════════════════════════════════
# COMMAND HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════════

@dp.message(Command("start"))
async def cmd_start(message: types.Message, state: FSMContext):
    """Start command - welcome message and main menu"""
    await state.clear()
    user_id = message.from_user.id
    is_admin = data_manager.is_admin(user_id)
    
    # Track user who started the bot (for broadcasting)
    user_info = {
        'first_name': message.from_user.first_name,
        'username': message.from_user.username,
        'last_name': message.from_user.last_name
    }
    data_manager.add_bot_user(user_id, user_info)
    
    welcome_text = (
        "🌟 <b>Kitobxon Kids botiga xush kelibsiz!</b>\n\n"
        "📚 Bu bot 7-14 yosh oradagi bolalar uchun o'qish tushunchasi testlarini o'tkazish uchun yaratilgan.\n\n"
        "📋 <b>Quyidagi imkoniyatlar mavjud:</b>\n"
        "• Ro'yxatdan o'tish\n"
        "• Yosh guruhingizga mos testlar topshirish\n"
        "• Natijalarni kuzatish\n"
        "• Fikr va takliflar yuborish\n\n"
        "📝 Boshlash uchun kerakli tugmani tanlang!"
    )
    
    await message.answer(
        welcome_text,
        reply_markup=KeyboardBuilder.get_main_menu(is_admin)
    )

@dp.message(Command("help"))
async def cmd_help(message: types.Message):
    """Help command - detailed instructions"""
    help_text = (
        "📖 <b>Bot foydalanish qo'llanmasi</b>\n\n"
        "📋 <b>Ro'yxatdan o'tish:</b>\n"
        "Botdan foydalanish uchun avval ro'yxatdan o'ting\n\n"
        "📝 <b>Test topshirish:</b>\n"
        "Yosh guruhingizga mos testni tanlang va savollarga javob bering\n\n"
        "💬 <b>Fikr va maslahat:</b>\n"
        "Loyiha haqida o'z fikringizni bildiring\n\n"
        "📚 <b>Loyiha haqida:</b>\n"
        "Kitobxon Kids loyihasi haqida batafsil ma'lumot\n\n"
        "❓ Savollar bo'lsa, admin bilan bog'laning"
    )
    
    await message.answer(help_text)

@dp.message(Command("stats"))
async def cmd_stats(message: types.Message):
    """Statistics command for admins"""
    if not data_manager.is_admin(message.from_user.id):
        await message.answer("❌ Bu buyruq faqat adminlar uchun!")
        return
    
    try:
        data_manager.update_statistics()
        stats = data_manager.get_statistics()
        
        last_updated = stats.get('last_updated', "Noma'lum")[:19]
        stats_text = (
            "📊 <b>Bot statistikasi</b>\n\n"
            f"👥 Jami foydalanuvchilar: <b>{stats.get('total_registered_users', 0)}</b>\n"
            f"📝 Jami testlar: <b>{stats.get('test_statistics', {}).get('total_tests_taken', 0)}</b>\n"
            f"📈 O'rtacha ball: <b>{stats.get('test_statistics', {}).get('average_score', 0)}%</b>\n"
            f"🏆 70+ ball olganlar: <b>{stats.get('test_statistics', {}).get('high_scorers_70plus', 0)}</b>\n\n"
            f"📅 Oxirgi yangilanish: <b>{last_updated}</b>"
        )
        
        await message.answer(stats_text)
        
    except Exception as e:
        logger.error(f"Error in stats command: {e}")
        await message.answer("❌ Statistika olishda xatolik yuz berdi!")

# ═══════════════════════════════════════════════════════════════════════════════════
# MAIN MENU HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════════

@dp.message(F.text == "📋 Ro'yxatdan o'tish")
async def registration_start(message: types.Message, state: FSMContext):
    """Start registration process"""
    user_id = message.from_user.id
    
    # Check if already registered
    if data_manager.is_user_registered(user_id):
        await message.answer(
            "✅ Siz allaqachon ro'yxatdan o'tgansiz!\n\n"
            "📝 Test topshirish uchun tegishli tugmani bosing.",
            reply_markup=KeyboardBuilder.get_main_menu(data_manager.is_admin(user_id))
        )
        return
    
    # Check subscription
    if not await check_subscription(user_id):
        await message.answer(
            "❌ <b>Botdan foydalanish uchun kanalimizga obuna bo'ling!</b>\n\n"
            f"🔗 Kanal: {config.CHANNEL_USERNAME}\n\n"
            "Obuna bo'lgandan keyin qaytadan urinib ko'ring."
        )
        return
    
    await message.answer(
        "👶 <b>Bolangizning ismini kiriting:</b>\n\n"
        "Misol: Ahmad, Fatima, Shohrux\n\n"
        "📝 Ism va familiyani to'liq yozing.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Orqaga")]],
            resize_keyboard=True
        )
    )
    await state.set_state(RegistrationStates.child_name)

@dp.message(RegistrationStates.child_name)
async def process_child_name(message: types.Message, state: FSMContext):
    """Process child's name"""
    if message.text == "🔙 Orqaga":
        await state.clear()
        await message.answer(
            "🔙 Asosiy menyuga qaytdingiz.",
            reply_markup=KeyboardBuilder.get_main_menu(data_manager.is_admin(message.from_user.id))
        )
        return
    
    child_name = input_validator.sanitize_input(message.text or "")
    
    if not input_validator.validate_name(child_name):
        await message.answer(
            "❌ <b>Noto'g'ri ism formati!</b>\n\n"
            "Iltimos, bola ismini to'g'ri kiriting:\n"
            "• Kamida 2 ta harf\n"
            "• Faqat harflar va probel\n"
            "• Maxsus belgilar ishlatilmasin"
        )
        return
    
    await state.update_data(child_name=child_name)
    await message.answer(
        "👨‍👩‍👧‍👦 <b>Ota-ona ismini kiriting:</b>\n\n"
        "Misol: Oybek Karimov, Nodira Abdullayeva\n\n"
        "📝 To'liq ism va familiyani yozing."
    )
    await state.set_state(RegistrationStates.parent_name)

@dp.message(RegistrationStates.parent_name)
async def process_parent_name(message: types.Message, state: FSMContext):
    """Process parent's name"""
    if message.text == "🔙 Orqaga":
        await message.answer(
            "👶 <b>Bolangizning ismini kiriting:</b>\n\n"
            "Misol: Ahmad, Fatima, Shohrux\n\n"
            "📝 Ism va familiyani to'liq yozing."
        )
        await state.set_state(RegistrationStates.child_name)
        return
    
    parent_name = input_validator.sanitize_input(message.text or "")
    
    if not input_validator.validate_name(parent_name):
        await message.answer(
            "❌ <b>Noto'g'ri ism formati!</b>\n\n"
            "Iltimos, ota-ona ismini to'g'ri kiriting:\n"
            "• Kamida 2 ta harf\n"
            "• Faqat harflar va probel\n" 
            "• Maxsus belgilar ishlatilmasin"
        )
        return
    
    await state.update_data(parent_name=parent_name)
    await message.answer(
        "🏛 <b>Viloyatingizni tanlang:</b>",
        reply_markup=KeyboardBuilder.get_regions_keyboard()
    )
    await state.set_state(RegistrationStates.region)

@dp.message(RegistrationStates.region)
async def process_region(message: types.Message, state: FSMContext):
    """Process region selection"""
    if message.text == "🔙 Orqaga":
        await message.answer(
            "👨‍👩‍👧‍👦 <b>Ota-ona ismini kiriting:</b>\n\n"
            "Misol: Oybek Karimov, Nodira Abdullayeva\n\n"
            "📝 To'liq ism va familiyani yozing."
        )
        await state.set_state(RegistrationStates.parent_name)
        return
    
    region = message.text and message.text.strip()
    
    if not region or region not in REGIONS:
        await message.answer(
            "❌ <b>Noto'g'ri viloyat tanlandi!</b>\n\n"
            "Iltimos, ro'yxatdan viloyatni tanlang:",
            reply_markup=KeyboardBuilder.get_regions_keyboard()
        )
        return
    
    await state.update_data(region=region)
    await message.answer(
        f"🏘 <b>{region} viloyati uchun tumanni tanlang:</b>",
        reply_markup=KeyboardBuilder.get_districts_keyboard(region)
    )
    await state.set_state(RegistrationStates.district)

@dp.message(RegistrationStates.district)
async def process_district(message: types.Message, state: FSMContext):
    """Process district selection"""
    if message.text == "🔙 Orqaga":
        await message.answer(
            "🏛 <b>Viloyatingizni tanlang:</b>",
            reply_markup=KeyboardBuilder.get_regions_keyboard()
        )
        await state.set_state(RegistrationStates.region)
        return
    
    data = await state.get_data()
    region = data.get('region')
    district = message.text and message.text.strip()
    
    if not input_validator.validate_region_district(region or "", district or ""):
        await message.answer(
            f"❌ <b>Noto'g'ri tuman tanlandi!</b>\n\n"
            f"Iltimos, {region} viloyati uchun tumanni tanlang:",
            reply_markup=KeyboardBuilder.get_districts_keyboard(region or "")
        )
        return
    
    await state.update_data(district=district)
    await message.answer(
        "🏠 <b>Mahalla nomini kiriting:</b>\n\n"
        "Misol: Chilonzor, Bobur, Mustaqillik\n\n"
        "📝 Mahalla yoki MFY nomini yozing."
    )
    await state.set_state(RegistrationStates.mahalla)

@dp.message(RegistrationStates.mahalla)
async def process_mahalla(message: types.Message, state: FSMContext):
    """Process mahalla (neighborhood)"""
    if message.text == "🔙 Orqaga":
        data = await state.get_data()
        region = data.get('region', '')
        await message.answer(
            f"🏘 <b>{region} viloyati uchun tumanni tanlang:</b>",
            reply_markup=KeyboardBuilder.get_districts_keyboard(region)
        )
        await state.set_state(RegistrationStates.district)
        return
    
    mahalla = input_validator.sanitize_input(message.text or "")
    
    if not mahalla or len(mahalla.strip()) < 2:
        await message.answer(
            "❌ <b>Mahalla nomini to'g'ri kiriting!</b>\n\n"
            "Mahalla yoki MFY nomini kamida 2 ta harf bilan yozing."
        )
        return
    
    await state.update_data(mahalla=mahalla)
    await message.answer(
        "👶 <b>Bolangizning yoshini tanlang:</b>",
        reply_markup=KeyboardBuilder.get_age_keyboard()
    )
    await state.set_state(RegistrationStates.age)

@dp.message(RegistrationStates.age)
async def process_age(message: types.Message, state: FSMContext):
    """Process age selection"""
    if message.text == "🔙 Orqaga":
        await message.answer(
            "🏠 <b>Mahalla nomini kiriting:</b>\n\n"
            "Misol: Chilonzor, Bobur, Mustaqillik\n\n"
            "📝 Mahalla yoki MFY nomini yozing."
        )
        await state.set_state(RegistrationStates.mahalla)
        return
    
    age = message.text and message.text.strip()
    
    if not input_validator.validate_age(age or ""):
        await message.answer(
            "❌ <b>Noto'g'ri yosh!</b>\n\n"
            "Iltimos, 7-14 yosh oralig'ida tanlang:",
            reply_markup=KeyboardBuilder.get_age_keyboard()
        )
        return
    
    await state.update_data(age=age)
    await message.answer(
        "📱 <b>Telefon raqamingizni kiriting:</b>\n\n"
        "Format: 998901234567\n\n"
        "📞 O'zbekiston raqamlarini kiriting (998 bilan boshlangan).",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[
                [KeyboardButton(text="📱 Raqamni yuborish", request_contact=True)],
                [KeyboardButton(text="🔙 Orqaga")]
            ],
            resize_keyboard=True
        )
    )
    await state.set_state(RegistrationStates.phone)

@dp.message(RegistrationStates.phone)
async def process_phone(message: types.Message, state: FSMContext):
    """Process phone number"""
    if message.text == "🔙 Orqaga":
        await message.answer(
            "👶 <b>Bolangizning yoshini tanlang:</b>",
            reply_markup=KeyboardBuilder.get_age_keyboard()
        )
        await state.set_state(RegistrationStates.age)
        return
    
    # Handle contact sharing
    if message.contact:
        phone = str(message.contact.phone_number)
    else:
        phone = message.text or ""
    
    # Clean phone number
    phone_clean = re.sub(r'[^\d]', '', phone)
    if phone_clean.startswith('+998'):
        phone_clean = phone_clean[1:]
    elif phone_clean.startswith('8') and len(phone_clean) == 9:
        phone_clean = '998' + phone_clean[1:]
    elif not phone_clean.startswith('998') and len(phone_clean) == 9:
        phone_clean = '998' + phone_clean
    
    if not input_validator.validate_phone(phone_clean):
        await message.answer(
            "❌ <b>Noto'g'ri telefon raqam!</b>\n\n"
            "Format: 998901234567\n\n"
            "📞 O'zbekiston raqamlarini kiriting."
        )
        return
    
    # Save registration data
    data = await state.get_data()
    user_data = {
        "child_name": data.get('child_name'),
        "parent_name": data.get('parent_name'),
        "region": data.get('region'),
        "district": data.get('district'),
        "mahalla": data.get('mahalla'),
        "age": data.get('age'),
        "phone": phone_clean,
        "telegram_id": message.from_user.id,
        "username": message.from_user.username or "Noma'lum"
    }
    
    try:
        data_manager.save_user(message.from_user.id, user_data)
        data_manager.update_statistics()
        
        # Registration success message
        success_text = (
            "✅ <b>Tabriklaymiz! Ro'yxatdan muvaffaqiyatli o'tdingiz!</b>\n\n"
            "📋 <b>Ma'lumotlaringiz:</b>\n"
            f"👶 Bola: <b>{data.get('child_name')}</b>\n"
            f"👨‍👩‍👧‍👦 Ota-ona: <b>{data.get('parent_name')}</b>\n"
            f"🏛 Viloyat: <b>{data.get('region')}</b>\n"
            f"🏘 Tuman: <b>{data.get('district')}</b>\n"
            f"👶 Yosh: <b>{data.get('age')}</b>\n\n"
            "🎉 Endi testlar topshirishingiz mumkin!\n"
            "📝 'Test topshirish' tugmasini bosing."
        )
        
        await message.answer(
            success_text,
            reply_markup=KeyboardBuilder.get_main_menu(data_manager.is_admin(message.from_user.id))
        )
        
        # Notify admins
        admin_notification = (
            "👤 <b>Yangi foydalanuvchi ro'yxatdan o'tdi!</b>\n\n"
            f"👶 Bola: {data.get('child_name')}\n"
            f"👨‍👩‍👧‍👦 Ota-ona: {data.get('parent_name')}\n"
            f"📍 Manzil: {data.get('region')}, {data.get('district')}\n"
            f"👶 Yosh: {data.get('age')}\n"
            f"🆔 ID: {message.from_user.id}"
        )
        await send_admin_notification(admin_notification)
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error saving user registration: {e}")
        await message.answer(
            "❌ Ro'yxatdan o'tishda xatolik yuz berdi!\n"
            "Iltimos, qaytadan urinib ko'ring."
        )

@dp.message(F.text == "📝 Test topshirish")
async def test_start(message: types.Message, state: FSMContext):
    """Start test taking process"""
    user_id = message.from_user.id
    
    # Check registration
    if not data_manager.is_user_registered(user_id):
        await message.answer(
            "❌ Testlarni topshirish uchun avval ro'yxatdan o'ting!\n\n"
            "📋 'Ro'yxatdan o'tish' tugmasini bosing."
        )
        return
    
    user_data = data_manager.get_user(user_id)
    if not user_data:
        await message.answer(
            "❌ Foydalanuvchi ma'lumotlari topilmadi!\n"
            "Iltimos, qaytadan ro'yxatdan o'ting."
        )
        return
    
    user_age = int(user_data.get('age', 0))
    if user_age < 7 or user_age > 14:
        await message.answer(
            "❌ Yoshingiz test topshirish uchun mos emas!\n"
            "Testlar 7-14 yosh oradagi bolalar uchun mo'ljallangan."
        )
        return
    
    # Determine age group
    age_group = "7-10" if user_age <= 10 else "11-14"
    tests = data_manager.get_tests()
    
    if not tests.get(age_group):
        await message.answer(
            f"❌ {age_group} yosh guruhi uchun testlar mavjud emas!\n\n"
            "Admin bilan bog'laning yoki keyinroq qaytadan urinib ko'ring."
        )
        return
    
    # Show available tests
    test_text = (
        f"📝 <b>{age_group} yosh guruhi uchun mavjud testlar:</b>\n\n"
        "Quyidagi testlardan birini tanlang:"
    )
    
    keyboard = KeyboardBuilder.get_test_selection_keyboard(age_group, tests)
    await message.answer(test_text, reply_markup=keyboard)

@dp.callback_query(F.data.startswith("start_test:"))
async def start_test(callback: types.CallbackQuery, state: FSMContext):
    """Start a specific test"""
    try:
        _, age_group, test_id = callback.data.split(":")
        
        test_data = data_manager.get_test_by_id(age_group, test_id)
        if not test_data:
            await callback.answer("❌ Test topilmadi!")
            return
        
        questions = test_data.get('questions', [])
        if not questions:
            await callback.answer("❌ Test savollari mavjud emas!")
            return
        
        # Shuffle questions for variety
        shuffled_questions = questions.copy()
        random.shuffle(shuffled_questions)
        
        # Store test session data
        await state.update_data(
            test_id=test_id,
            age_group=age_group,
            book_name=test_data.get('book_name', 'Noma\'lum kitob'),
            questions=shuffled_questions,
            current_question=0,
            answers=[],
            start_time=config.get_uzbekistan_time().isoformat()
        )
        
        await show_question(callback.message, state)
        await state.set_state(TestStates.taking_test)
        
    except Exception as e:
        logger.error(f"Error starting test: {e}")
        await callback.answer("❌ Test boshlashda xatolik yuz berdi!")

async def show_question(message: types.Message, state: FSMContext):
    """Display current question to user"""
    data = await state.get_data()
    questions = data.get('questions', [])
    current_idx = data.get('current_question', 0)
    
    if current_idx >= len(questions):
        await complete_test(message, state)
        return
    
    question = questions[current_idx]
    question_num = current_idx + 1
    total_questions = len(questions)
    
    question_text = (
        f"📝 <b>Savol {question_num}/{total_questions}</b>\n\n"
        f"❓ {question['question']}\n\n"
        f"A) {question['options']['A']}\n"
        f"B) {question['options']['B']}\n"
        f"C) {question['options']['C']}\n"
        f"D) {question['options']['D']}"
    )
    
    keyboard = KeyboardBuilder.get_question_keyboard()
    
    try:
        if message.photo or message.document:
            await message.answer(question_text, reply_markup=keyboard)
        else:
            await message.edit_text(question_text, reply_markup=keyboard)
    except:
        await message.answer(question_text, reply_markup=keyboard)

@dp.callback_query(F.data.startswith("answer:"))
async def process_answer(callback: types.CallbackQuery, state: FSMContext):
    """Process test answer"""
    try:
        answer = callback.data.split(":")[1]
        
        data = await state.get_data()
        questions = data.get('questions', [])
        current_idx = data.get('current_question', 0)
        answers = data.get('answers', [])
        
        if current_idx >= len(questions):
            await callback.answer("❌ Test allaqachon yakunlangan!")
            return
        
        # Save answer
        answers.append(answer)
        current_idx += 1
        
        await state.update_data(
            answers=answers,
            current_question=current_idx
        )
        
        await callback.answer(f"✅ {answer} javob tanlandi!")
        
        # Show next question or complete test
        if current_idx < len(questions):
            await show_question(callback.message, state)
        else:
            await complete_test(callback.message, state)
            
    except Exception as e:
        logger.error(f"Error processing answer: {e}")
        await callback.answer("❌ Javob qayta ishlashda xatolik!")

async def complete_test(message: types.Message, state: FSMContext):
    """Complete test and show results"""
    try:
        data = await state.get_data()
        questions = data.get('questions', [])
        answers = data.get('answers', [])
        user_id = message.from_user.id
        
        # Calculate results
        correct_answers = 0
        for i, question in enumerate(questions):
            if i < len(answers) and answers[i] == question['correct_answer']:
                correct_answers += 1
        
        total_questions = len(questions)
        percentage, performance = calculate_test_score(correct_answers, total_questions)
        
        # Get user data
        user_data = data_manager.get_user(user_id) or {}
        
        # Save result
        result_data = {
            "user_id": user_id,
            "user_name": user_data.get('child_name', 'Unknown'),
            "test_id": data.get('test_id'),
            "book_name": data.get('book_name'),
            "age_group": data.get('age_group'),
            "age": user_data.get('age'),
            "region": user_data.get('region'),
            "district": user_data.get('district'),
            "total_questions": total_questions,
            "correct_answers": correct_answers,
            "score": percentage,
            "percentage": percentage,
            "performance": performance,
            "answers": answers,
            "start_time": data.get('start_time'),
            "completion_time": config.get_uzbekistan_time().isoformat()
        }
        
        data_manager.save_result(result_data)
        data_manager.update_statistics()
        
        # Result message
        emoji = "🎉" if percentage >= 70 else "📊"
        result_text = (
            f"{emoji} <b>Test yakunlandi!</b>\n\n"
            "📊 <b>Natijalar:</b>\n"
            f"✅ To'g'ri javoblar: <b>{correct_answers}/{total_questions}</b>\n"
            f"📈 Foiz: <b>{percentage}%</b>\n"
            f"🏅 Daraja: <b>{performance}</b>\n\n"
            f"📚 Kitob: <b>{result_data['book_name']}</b>\n"
            f"📅 Sana: <b>{config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}</b>\n\n"
            "🔄 Yana test topshirish uchun 'Test topshirish' tugmasini bosing!"
        )
        
        await message.answer(
            result_text,
            reply_markup=KeyboardBuilder.get_main_menu(data_manager.is_admin(user_id))
        )
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error completing test: {e}")
        await message.answer(
            "❌ Test yakunlashda xatolik yuz berdi!\n"
            "Iltimos, qaytadan urinib ko'ring."
        )

@dp.message(F.text == "💬 Fikr va maslahatlar")
async def feedback_start(message: types.Message, state: FSMContext):
    """Start feedback collection"""
    user_id = message.from_user.id
    
    if not data_manager.is_user_registered(user_id):
        await message.answer(
            "❌ Fikr bildirish uchun avval ro'yxatdan o'ting!\n\n"
            "📋 'Ro'yxatdan o'tish' tugmasini bosing."
        )
        return
    
    await message.answer(
        "💬 <b>Fikr va maslahatlaringizni yuboring!</b>\n\n"
        "📝 Loyiha haqida o'z fikringizni, taklif va tanqidlaringizni yozing.\n"
        "Sizning fikringiz biz uchun juda muhim!\n\n"
        "✍️ Xabaringizni yozing:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Orqaga")]],
            resize_keyboard=True
        )
    )
    await state.set_state(RegistrationStates.feedback)

@dp.message(RegistrationStates.feedback)
async def process_feedback(message: types.Message, state: FSMContext):
    """Process user feedback"""
    if message.text == "🔙 Orqaga":
        await state.clear()
        await message.answer(
            "🔙 Asosiy menyuga qaytdingiz.",
            reply_markup=KeyboardBuilder.get_main_menu(data_manager.is_admin(message.from_user.id))
        )
        return
    
    feedback_text = input_validator.sanitize_input(message.text or "")
    
    if not feedback_text or len(feedback_text.strip()) < 5:
        await message.answer(
            "❌ <b>Fikringiz juda qisqa!</b>\n\n"
            "Iltimos, kamida 5 ta harf bo'lgan matnni yozing."
        )
        return
    
    user_id = message.from_user.id
    user_data = data_manager.get_user(user_id) or {}
    
    # Send feedback to admins
    admin_notification = (
        "💬 <b>Yangi fikr-mulohaza!</b>\n\n"
        f"👤 Foydalanuvchi: {user_data.get('parent_name', 'Unknown')}\n"
        f"👶 Bola: {user_data.get('child_name', 'Unknown')}\n"
        f"🏛 Manzil: {user_data.get('region', 'Unknown')}, {user_data.get('district', 'Unknown')}\n"
        f"📱 Telefon: {user_data.get('phone', 'Unknown')}\n"
        f"🆔 ID: {user_id}\n\n"
        f"💭 <b>Fikr:</b>\n{feedback_text}\n\n"
        f"📅 Sana: {config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}"
    )
    
    await send_admin_notification(admin_notification)
    
    await message.answer(
        "✅ <b>Rahmat!</b>\n\n"
        "💬 Fikringiz adminlarga yuborildi.\n"
        "🔍 Tez orada ko'rib chiqiladi.\n\n"
        "📝 Davom etish uchun kerakli tugmani tanlang.",
        reply_markup=KeyboardBuilder.get_main_menu(data_manager.is_admin(user_id))
    )
    
    await state.clear()

@dp.message(F.text == "📚 Loyiha haqida")
async def project_info(message: types.Message):
    """Project information"""
    info_text = (
        "📚 <b>Kitobxon Kids loyihasi haqida</b>\n\n"
        "🎯 <b>Maqsad:</b>\n"
        "7-14 yosh oradagi bolalarning o'qish qobiliyatini rivojlantirish va baholash\n\n"
        "📖 <b>Imkoniyatlar:</b>\n"
        "• Yosh guruhlariga mos testlar\n"
        "• Natijalarni kuzatish\n"
        "• O'qish sifatini baholash\n"
        "• Regional statistika\n\n"
        "🏆 <b>Maqsad:</b>\n"
        "Bolalarning o'qish va tushunish qobiliyatini rivojlantirish\n\n"
        f"📢 Kanal: {config.CHANNEL_USERNAME}\n"
        "💬 Admin: @admin\n\n"
        "🎉 Biz bilan birga o'qish olamini kashf eting!"
    )
    
    await message.answer(info_text)

# ═══════════════════════════════════════════════════════════════════════════════════
# ADMIN PANEL HANDLERS
# ═══════════════════════════════════════════════════════════════════════════════════

@dp.message(F.text == "👨‍💼 Admin panel")
async def admin_panel(message: types.Message, state: FSMContext):
    """Admin panel access"""
    user_id = message.from_user.id
    
    if not data_manager.is_admin(user_id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    await state.clear()
    is_super = data_manager.is_super_admin(user_id)
    
    admin_text = (
        "👨‍💼 <b>Admin paneli</b>\n\n"
        f"👤 Admin: {message.from_user.first_name}\n"
        f"🔑 Daraja: {'Super Admin' if is_super else 'Admin'}\n\n"
        "⚙️ Kerakli amalni tanlang:"
    )
    
    await message.answer(
        admin_text,
        reply_markup=KeyboardBuilder.get_admin_menu(is_super)
    )

@dp.message(F.text == "👥 Foydalanuvchilar ro'yxati")
async def admin_users_list(message: types.Message):
    """Show users list for admins"""
    if not data_manager.is_admin(message.from_user.id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    users = data_manager.get_users()
    
    if not users:
        await message.answer("📭 Hozircha ro'yxatdan o'tgan foydalanuvchilar yo'q.")
        return
    
    # Show users with pagination
    user_list = list(users.items())
    page_size = config.USERS_PER_PAGE
    total_pages = (len(user_list) + page_size - 1) // page_size
    
    for i in range(0, len(user_list), page_size):
        page_users = user_list[i:i + page_size]
        text = f"👥 <b>Foydalanuvchilar ({i+1}-{min(i+page_size, len(user_list))} / {len(user_list)}):</b>\n\n"
        
        for user_id, user_data in page_users:
            text += (
                f"🆔 {user_id}\n"
                f"👶 {user_data.get('child_name', 'N/A')}\n"
                f"👨‍👩‍👧‍👦 {user_data.get('parent_name', 'N/A')}\n"
                f"📍 {user_data.get('region', 'N/A')}, {user_data.get('district', 'N/A')}\n"
                f"👶 {user_data.get('age', 'N/A')} yosh\n"
                f"📅 {user_data.get('registration_date', 'N/A')[:10]}\n"
                "➖➖➖➖➖➖➖➖➖➖\n"
            )
        
        await message.answer(text)
    
    # Send Excel report if available
    if EXCEL_AVAILABLE:
        try:
            results = data_manager.get_results()
            excel_data = DocumentGenerator.generate_excel_report(users, results)
            
            if excel_data:
                document = BufferedInputFile(
                    excel_data.read(),
                    filename=f"users_report_{config.get_uzbekistan_time().strftime('%Y%m%d')}.xlsx"
                )
                await message.answer_document(
                    document,
                    caption="📊 Foydalanuvchilar hisoboti (Excel format)"
                )
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")

@dp.message(F.text == "➕ Test qo'shish")
async def admin_add_test_start(message: types.Message, state: FSMContext):
    """Start test addition process"""
    if not data_manager.is_admin(message.from_user.id):
        await message.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    await message.answer(
        "📚 <b>Yangi test qo'shish</b>\n\n"
        "👶 Yosh guruhini tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="7-10 yosh", callback_data="add_test:7-10")],
            [InlineKeyboardButton(text="11-14 yosh", callback_data="add_test:11-14")]
        ])
    )

@dp.callback_query(F.data.startswith("add_test:"))
async def admin_add_test_age(callback: types.CallbackQuery, state: FSMContext):
    """Process age group selection for new test"""
    age_group = callback.data.split(":")[1]
    
    await state.update_data(age_group=age_group)
    await callback.message.edit_text(
        f"📚 <b>{age_group} yosh guruhi uchun test</b>\n\n"
        "📖 Kitob nomini yozing:"
    )
    await state.set_state(AdminStates.add_test_book)
    await callback.answer()

@dp.message(AdminStates.add_test_book)
async def admin_add_test_book(message: types.Message, state: FSMContext):
    """Process book name for new test"""
    book_name = input_validator.sanitize_input(message.text or "")
    
    if not book_name or len(book_name.strip()) < 2:
        await message.answer(
            "❌ Kitob nomini to'g'ri kiriting!\n"
            "Kamida 2 ta belgi bo'lishi kerak."
        )
        return
    
    await state.update_data(
        book_name=book_name,
        questions=[],
        current_question=1
    )
    
    data = await state.get_data()
    await message.answer(
        f"📖 <b>Kitob:</b> {book_name}\n\n"
        "❓ <b>1-savolni qo'shing:</b>\n\n"
        "<b>Format:</b>\n"
        "Savol matni?\n"
        "A) Birinchi variant\n"
        "B) Ikkinchi variant\n"
        "C) Uchinchi variant\n"
        "D) To'rtinchi variant\n"
        "Javob: A\n\n"
        f"✅ Jami qo'shilgan: 0 ta\n"
        f"📝 Minimum kerak: {config.MIN_QUESTIONS_PER_TEST} ta"
    )
    await state.set_state(AdminStates.add_test_questions)

@dp.message(AdminStates.add_test_questions)
async def admin_add_test_question(message: types.Message, state: FSMContext):
    """Process individual questions for test"""
    if message.text in ["✅ Testni saqlash", "/done", "/finish"]:
        await admin_finish_test(message, state)
        return
    
    question_data = input_validator.parse_question(message.text or "")
    
    if not question_data:
        await message.answer(
            "❌ <b>Savol formati noto'g'ri!</b>\n\n"
            "<b>To'g'ri format:</b>\n"
            "Savol matni?\n"
            "A) Birinchi variant\n"
            "B) Ikkinchi variant\n" 
            "C) Uchinchi variant\n"
            "D) To'rtinchi variant\n"
            "Javob: A\n\n"
            "🔄 Qaytadan urinib ko'ring."
        )
        return
    
    data = await state.get_data()
    questions = data.get('questions', [])
    questions.append(question_data)
    current_num = len(questions)
    
    await state.update_data(
        questions=questions,
        current_question=current_num + 1
    )
    
    # Create finish button after minimum questions
    keyboard = None
    if current_num >= config.MIN_QUESTIONS_PER_TEST:
        keyboard = ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="✅ Testni saqlash")]],
            resize_keyboard=True
        )
    
    response_text = (
        f"✅ <b>Savol {current_num} qo'shildi!</b>\n\n"
        f"📊 Jami savollar: <b>{len(questions)}</b>\n\n"
        f"❓ <b>{current_num + 1}-savolni qo'shing</b> yoki testni saqlang:"
    )
    
    if current_num < config.MIN_QUESTIONS_PER_TEST:
        response_text += f"\n(Minimum {config.MIN_QUESTIONS_PER_TEST} ta savol kerak)"
    
    await message.answer(response_text, reply_markup=keyboard)

async def admin_finish_test(message: types.Message, state: FSMContext):
    """Complete test creation process"""
    data = await state.get_data()
    questions = data.get('questions', [])
    
    if len(questions) < config.MIN_QUESTIONS_PER_TEST:
        await message.answer(
            f"❌ Test kamida {config.MIN_QUESTIONS_PER_TEST} ta savolga ega bo'lishi kerak!\n\n"
            f"📊 Hozirda: {len(questions)} ta savol\n"
            f"📝 Kerak: {config.MIN_QUESTIONS_PER_TEST} ta savol\n\n"
            "➕ Yana savollar qo'shing."
        )
        return
    
    try:
        test_data = {
            "age_group": data.get('age_group'),
            "book_name": data.get('book_name'),
            "questions": questions,
            "created_by": message.from_user.id
        }
        
        test_id = data_manager.save_test(test_data)
        
        success_text = (
            "✅ <b>Test muvaffaqiyatli saqlandi!</b>\n\n"
            f"📚 Kitob: <b>{test_data['book_name']}</b>\n"
            f"👥 Yosh guruhi: <b>{test_data['age_group']}</b>\n"
            f"❓ Savollar soni: <b>{len(questions)}</b>\n"
            f"🆔 Test ID: <code>{test_id}</code>\n\n"
            "🎉 Test foydalanuvchilar uchun tayyor!"
        )
        
        is_super = data_manager.is_super_admin(message.from_user.id)
        await message.answer(
            success_text,
            reply_markup=KeyboardBuilder.get_admin_menu(is_super)
        )
        
        # Notify other admins
        admin_name = message.from_user.first_name or "Admin"
        notification = (
            f"➕ <b>Yangi test qo'shildi!</b>\n\n"
            f"👤 Admin: {admin_name}\n"
            f"📚 Kitob: {test_data['book_name']}\n"
            f"👥 Yosh: {test_data['age_group']}\n"
            f"❓ Savollar: {len(questions)} ta"
        )
        await send_admin_notification(notification, exclude_id=message.from_user.id)
        
        await state.clear()
        
    except Exception as e:
        logger.error(f"Error saving test: {e}")
        await message.answer(
            "❌ Testni saqlashda xatolik yuz berdi!\n"
            "Iltimos, qaytadan urinib ko'ring."
        )

@dp.message(F.text == "➕ Admin qo'shish")
async def admin_add_admin_start(message: types.Message, state: FSMContext):
    """Start admin addition process (Super Admin only)"""
    if not data_manager.is_super_admin(message.from_user.id):
        await message.answer("❌ Bu funksiya faqat Super Adminlar uchun!")
        return
    
    await message.answer(
        "👤 <b>Yangi admin qo'shish</b>\n\n"
        "🆔 Yangi adminning Telegram ID raqamini yuboring:\n\n"
        "💡 ID ni qanday topish mumkin:\n"
        "• @userinfobot dan foydalaning\n"
        "• Yoki admindan o'zining ID sini so'rang",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Orqaga")]],
            resize_keyboard=True
        )
    )
    await state.set_state(AdminStates.add_admin)

@dp.message(AdminStates.add_admin)
async def admin_add_admin_process(message: types.Message, state: FSMContext):
    """Process admin addition"""
    if message.text == "🔙 Orqaga":
        await state.clear()
        is_super = data_manager.is_super_admin(message.from_user.id)
        await message.answer(
            "🔙 Admin panelga qaytdingiz.",
            reply_markup=KeyboardBuilder.get_admin_menu(is_super)
        )
        return
    
    try:
        new_admin_id = int(message.text.strip())
        
        if data_manager.is_admin(new_admin_id):
            await message.answer(
                "❌ Bu foydalanuvchi allaqachon admin!\n\n"
                "🔄 Boshqa ID kiriting yoki bekor qiling."
            )
            return
        
        # Try to get admin info
        try:
            admin_info = await bot.get_chat(new_admin_id)
            admin_name = admin_info.first_name or f"User_{new_admin_id}"
        except:
            admin_name = f"User_{new_admin_id}"
        
        # Save new admin
        admin_data = {
            "role": "admin",
            "added_by": message.from_user.id,
            "added_by_name": message.from_user.first_name or "Unknown"
        }
        
        data_manager.save_admin(new_admin_id, admin_data)
        
        success_text = (
            "✅ <b>Yangi admin qo'shildi!</b>\n\n"
            f"👤 Admin: {admin_name}\n"
            f"🆔 ID: <code>{new_admin_id}</code>\n"
            f"📅 {config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}\n\n"
            "🎉 Admin huquqlari faollashtirildi!"
        )
        
        is_super = data_manager.is_super_admin(message.from_user.id)
        await message.answer(
            success_text,
            reply_markup=KeyboardBuilder.get_admin_menu(is_super)
        )
        
        # Notify new admin
        try:
            welcome_admin_text = (
                f"🎉 <b>Tabriklaymiz!</b>\n\n"
                f"👑 Siz Kitobxon Kids botida admin etib tayinlandingiz!\n\n"
                f"👤 Tayinlagan: {message.from_user.first_name or 'Super Admin'}\n"
                f"📅 Sana: {config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}\n\n"
                f"⚙️ Admin paneliga kirish uchun /start bosing."
            )
            await bot.send_message(new_admin_id, welcome_admin_text)
        except Exception as e:
            logger.warning(f"Could not notify new admin {new_admin_id}: {e}")
        
        await state.clear()
        
    except ValueError:
        await message.answer(
            "❌ Noto'g'ri format!\n\n"
            "🆔 Faqat raqam kiriting.\n"
            "Misol: 123456789"
        )
    except Exception as e:
        logger.error(f"Error adding admin: {e}")
        await message.answer(
            "❌ Admin qo'shishda xatolik yuz berdi!\n"
            "Iltimos, qaytadan urinib ko'ring."
        )

@dp.message(F.text == "📊 Statistika")
async def admin_statistics(message: types.Message):
    """Show comprehensive statistics (Super Admin only)"""
    if not data_manager.is_super_admin(message.from_user.id):
        await message.answer("❌ Bu funksiya faqat Super Adminlar uchun!")
        return
    
    try:
        data_manager.update_statistics()
        stats = data_manager.get_statistics()
        
        total_users = stats.get('total_registered_users', 0)
        test_stats = stats.get('test_statistics', {})
        
        # Main statistics
        main_stats = (
            "📊 <b>Umumiy statistika</b>\n\n"
            f"👥 <b>Foydalanuvchilar:</b>\n"
            f"• Jami ro'yxatdan o'tganlar: <b>{total_users}</b>\n\n"
            f"📝 <b>Testlar:</b>\n"
            f"• Jami topshirilgan: <b>{test_stats.get('total_tests_taken', 0)}</b>\n"
            f"• O'rtacha ball: <b>{test_stats.get('average_score', 0)}%</b>\n"
            f"• 70+ ball olganlar: <b>{test_stats.get('high_scorers_70plus', 0)}</b>\n\n"
            f"👶 <b>Yosh guruhlari bo'yicha:</b>\n"
            f"• 7-10 yosh: <b>{test_stats.get('age_group_stats', {}).get('7-10', {}).get('count', 0)}</b> ta test "
            f"(o'rtacha: {test_stats.get('age_group_stats', {}).get('7-10', {}).get('avg_score', 0)}%)\n"
            f"• 11-14 yosh: <b>{test_stats.get('age_group_stats', {}).get('11-14', {}).get('count', 0)}</b> ta test "
            f"(o'rtacha: {test_stats.get('age_group_stats', {}).get('11-14', {}).get('avg_score', 0)}%)\n\n"
        )
        
        await message.answer(main_stats)
        
        # Regional statistics
        regional_stats = stats.get('regional_statistics', {})
        if regional_stats:
            region_text = "🗺 <b>Viloyatlar bo'yicha:</b>\n\n"
            
            for region, data in regional_stats.items():
                if data.get('total_users', 0) > 0:
                    region_text += f"• {region}: <b>{data['total_users']}</b>\n"
            
            await message.answer(region_text)
        
        # Send PDF report if available
        if PDF_AVAILABLE:
            try:
                pdf_data = DocumentGenerator.generate_pdf_report(stats)
                if pdf_data:
                    document = BufferedInputFile(
                        pdf_data.read(),
                        filename=f"statistics_report_{config.get_uzbekistan_time().strftime('%Y%m%d')}.pdf"
                    )
                    await message.answer_document(
                        document,
                        caption="📊 Statistik hisobot (PDF format)"
                    )
            except Exception as e:
                logger.error(f"Error generating PDF report: {e}")
        
    except Exception as e:
        logger.error(f"Error showing statistics: {e}")
        await message.answer("❌ Statistika olishda xatolik yuz berdi!")

@dp.message(F.text == "📢 Umumiy xabar")
async def admin_broadcast_start(message: types.Message, state: FSMContext):
    """Start broadcast process (Super Admin only)"""
    if not data_manager.is_super_admin(message.from_user.id):
        await message.answer("❌ Bu funksiya faqat Super Adminlar uchun!")
        return
    
    bot_users_count = data_manager.get_bot_users_count()
    registered_users_count = data_manager.get_user_count()
    
    if bot_users_count == 0:
        await message.answer("📭 Hozircha botga /start bosgan foydalanuvchilar yo'q!")
        return
    
    await message.answer(
        f"📤 <b>Umumiy xabar yuborish</b>\n\n"
        f"👥 Jami bot foydalanuvchilar: <b>{bot_users_count}</b>\n"
        f"✅ Ro'yxatdan o'tganlar: <b>{registered_users_count}</b>\n"
        f"👤 Ro'yxatsizlar: <b>{bot_users_count - registered_users_count}</b>\n\n"
        "✍️ Yubormoqchi bo'lgan xabaringizni yozing:\n\n"
        "💡 HTML formatdan foydalanishingiz mumkin:\n"
        "• <b>qalin</b> - qalin matn\n"
        "• <i>qiya</i> - qiya matn\n"
        "• <code>kod</code> - kod formati\n\n"
        "📝 Xabar barcha /start bosgan foydalanuvchilarga yuboriladi.",
        reply_markup=ReplyKeyboardMarkup(
            keyboard=[[KeyboardButton(text="🔙 Orqaga")]],
            resize_keyboard=True
        )
    )
    await state.set_state(AdminStates.broadcast_message)

@dp.message(AdminStates.broadcast_message)
async def admin_broadcast_message(message: types.Message, state: FSMContext):
    """Process broadcast message"""
    if message.text == "🔙 Orqaga":
        await state.clear()
        is_super = data_manager.is_super_admin(message.from_user.id)
        await message.answer(
            "🔙 Admin panelga qaytdingiz.",
            reply_markup=KeyboardBuilder.get_admin_menu(is_super)
        )
        return
    
    broadcast_text = input_validator.sanitize_input(message.text or "")
    
    if not broadcast_text or len(broadcast_text.strip()) < 5:
        await message.answer(
            "❌ <b>Xabar juda qisqa!</b>\n\n"
            "Kamida 5 ta belgi bo'lgan xabar yozing."
        )
        return
    
    await state.update_data(broadcast_text=broadcast_text)
    
    # Show preview
    preview_text = (
        f"📤 <b>Xabar ko'rinishi:</b>\n\n"
        f"{broadcast_text}\n\n"
        "❓ Xabarni yuborishni tasdiqlaysizmi?"
    )
    
    await message.answer(
        preview_text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Tasdiqlash", callback_data="broadcast_confirm"),
                InlineKeyboardButton(text="❌ Bekor qilish", callback_data="broadcast_cancel")
            ]
        ])
    )

@dp.callback_query(F.data == "broadcast_confirm")
async def admin_broadcast_confirm(callback: types.CallbackQuery, state: FSMContext):
    """Confirm and execute broadcast"""
    data = await state.get_data()
    broadcast_text = data.get('broadcast_text', '')
    
    # Get ALL bot users (registered + non-registered)
    bot_users = data_manager.get_all_bot_users_for_broadcast()
    
    if not bot_users:
        await callback.answer("📭 Bot foydalanuvchilar yo'q!")
        return
    
    # Start broadcasting
    await callback.message.edit_text(
        f"📤 <b>Xabar yuborilmoqda...</b>\n\n"
        f"👥 Jami bot foydalanuvchilar: {len(bot_users)}\n"
        "⏳ Iltimos, kuting..."
    )
    
    sent_count = 0
    failed_count = 0
    
    for user_id in bot_users.keys():
        try:
            await bot.send_message(int(user_id), broadcast_text)
            sent_count += 1
            await asyncio.sleep(0.05)  # Rate limiting
        except Exception as e:
            failed_count += 1
            logger.warning(f"Failed to send broadcast to {user_id}: {e}")
    
    # Save broadcast record
    broadcast_record = {
        "admin_id": callback.from_user.id,
        "admin_name": callback.from_user.first_name or "Unknown",
        "message": broadcast_text,
        "total_users": len(bot_users),
        "sent_count": sent_count,
        "failed_count": failed_count,
        "broadcast_type": "all_bot_users"
    }
    
    data_manager.save_broadcast(broadcast_record)
    
    # Show results
    result_text = (
        f"✅ <b>Xabar yuborish yakunlandi!</b>\n\n"
        f"📊 <b>Natijalar:</b>\n"
        f"• Muvaffaqiyatli: <b>{sent_count}</b>\n"
        f"• Xatolik: <b>{failed_count}</b>\n"
        f"• Jami: <b>{len(bot_users)}</b>\n\n"
        f"📝 Xabar barcha /start bosgan foydalanuvchilarga yuborildi\n"
        f"📅 {config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}"
    )
    
    is_super = data_manager.is_super_admin(callback.from_user.id)
    await callback.message.edit_text(
        result_text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Admin panelga", callback_data="back_to_admin")]
        ])
    )
    
    await state.clear()
    await callback.answer("✅ Xabar yuborildi!")

@dp.callback_query(F.data == "broadcast_cancel")
async def admin_broadcast_cancel(callback: types.CallbackQuery, state: FSMContext):
    """Cancel broadcast"""
    await state.clear()
    is_super = data_manager.is_super_admin(callback.from_user.id)
    
    await callback.message.edit_text(
        "❌ Xabar yuborish bekor qilindi.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔙 Admin panelga", callback_data="back_to_admin")]
        ])
    )
    await callback.answer("❌ Bekor qilindi")

@dp.callback_query(F.data == "back_to_admin")
async def back_to_admin_panel(callback: types.CallbackQuery):
    """Return to admin panel"""
    user_id = callback.from_user.id
    
    if not data_manager.is_admin(user_id):
        await callback.answer("❌ Sizda admin huquqlari yo'q!")
        return
    
    is_super = data_manager.is_super_admin(user_id)
    
    admin_text = (
        "👨‍💼 <b>Admin paneli</b>\n\n"
        f"👤 Admin: {callback.from_user.first_name}\n"
        f"🔑 Daraja: {'Super Admin' if is_super else 'Admin'}\n\n"
        "⚙️ Kerakli amalni tanlang:"
    )
    
    await callback.message.edit_text(admin_text)
    await callback.answer()

@dp.message(F.text == "📄 Hisobot")
async def admin_reports(message: types.Message):
    """Generate comprehensive reports (Super Admin only)"""
    if not data_manager.is_super_admin(message.from_user.id):
        await message.answer("❌ Bu funksiya faqat Super Adminlar uchun!")
        return
    
    try:
        users = data_manager.get_users()
        results = data_manager.get_results()
        
        if not users and not results:
            await message.answer("📭 Hisobot uchun ma'lumotlar yo'q!")
            return
        
        # Generate Excel report
        if EXCEL_AVAILABLE:
            excel_data = DocumentGenerator.generate_excel_report(users, results)
            if excel_data:
                document = BufferedInputFile(
                    excel_data.read(),
                    filename=f"full_report_{config.get_uzbekistan_time().strftime('%Y%m%d_%H%M')}.xlsx"
                )
                await message.answer_document(
                    document,
                    caption=(
                        f"📊 <b>Kitobxon Kids - To'liq hisobot</b>\n\n"
                        f"📅 Yaratilgan sana: {config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}\n"
                        f"👥 Foydalanuvchilar: {len(users)}\n"
                        f"📝 Test natijalari: {len(results)}\n\n"
                        "📊 Excel formatda to'liq ma'lumotlar"
                    )
                )
        
        # Generate PDF statistics report
        if PDF_AVAILABLE:
            data_manager.update_statistics()
            stats = data_manager.get_statistics()
            pdf_data = DocumentGenerator.generate_pdf_report(stats)
            
            if pdf_data:
                document = BufferedInputFile(
                    pdf_data.read(),
                    filename=f"statistics_{config.get_uzbekistan_time().strftime('%Y%m%d_%H%M')}.pdf"
                )
                await message.answer_document(
                    document,
                    caption=(
                        f"📄 <b>Kitobxon Kids - Statistik hisobot</b>\n\n"
                        f"📅 Yaratilgan sana: {config.get_uzbekistan_time().strftime('%Y-%m-%d %H:%M')}\n"
                        f"📊 PDF formatda statistik ma'lumotlar"
                    )
                )
        
        if not EXCEL_AVAILABLE and not PDF_AVAILABLE:
            await message.answer(
                "❌ Hisobot yaratish uchun kutubxonalar o'rnatilmagan!\n\n"
                "📝 Faqat matn formatda ma'lumotlar mavjud."
            )
        
    except Exception as e:
        logger.error(f"Error generating reports: {e}")
        await message.answer("❌ Hisobot yaratishda xatolik yuz berdi!")

@dp.message(F.text == "🔙 Asosiy menyu")
async def back_to_main_menu(message: types.Message, state: FSMContext):
    """Return to main menu"""
    await state.clear()
    is_admin = data_manager.is_admin(message.from_user.id)
    
    await message.answer(
        "🔙 Asosiy menyuga qaytdingiz.",
        reply_markup=KeyboardBuilder.get_main_menu(is_admin)
    )

# ═══════════════════════════════════════════════════════════════════════════════════
# ERROR HANDLERS AND MIDDLEWARE
# ═══════════════════════════════════════════════════════════════════════════════════

@dp.error()
async def error_handler(event: types.ErrorEvent):
    """Global error handler"""
    logger.error(f"Update {event.update} caused error: {event.exception}", exc_info=True)
    
    # Try to send error message to user if possible
    if event.update.message:
        try:
            await event.update.message.answer(
                "❌ Xatolik yuz berdi!\n"
                "Iltimos, qaytadan urinib ko'ring yoki admin bilan bog'laning."
            )
        except:
            pass
    
    return True

# ═══════════════════════════════════════════════════════════════════════════════════
# MAIN FUNCTION AND BOT STARTUP
# ═══════════════════════════════════════════════════════════════════════════════════

async def on_startup():
    """Bot startup initialization"""
    logger.info("Bot starting up...")
    
    # Validate bot token
    try:
        bot_info = await bot.get_me()
        logger.info(f"Bot started successfully: @{bot_info.username}")
    except Exception as e:
        logger.error(f"Failed to get bot info: {e}")
        raise
    
    # Initialize data manager and update statistics
    try:
        data_manager.update_statistics()
        logger.info("Data manager initialized and statistics updated")
    except Exception as e:
        logger.error(f"Error initializing data manager: {e}")
    
    logger.info("Bot startup completed successfully!")

async def on_shutdown():
    """Bot shutdown cleanup"""
    logger.info("Bot shutting down...")
    await bot.session.close()
    logger.info("Bot shutdown completed")

async def main():
    """Main function to run the bot"""
    try:
        # Startup procedures
        await on_startup()
        
        # Start polling
        logger.info("Starting bot polling...")
        await dp.start_polling(
            bot,
            skip_updates=True,
            close_bot_session=True
        )
        
    except KeyboardInterrupt:
        logger.info("Bot stopped by user")
    except Exception as e:
        logger.error(f"Critical error in main: {e}", exc_info=True)
        raise
    finally:
        await on_shutdown()

if __name__ == "__main__":
    """Entry point for the bot application"""
    try:
        # Set up asyncio event loop policy for Windows compatibility
        if os.name == 'nt':
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
        
        # Run the bot
        asyncio.run(main())
        
    except KeyboardInterrupt:
        logger.info("Bot application terminated by user")
    except Exception as e:
        logger.critical(f"Failed to start bot application: {e}", exc_info=True)
        raise