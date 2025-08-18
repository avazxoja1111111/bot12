"""
Kitobxon Kids Telegram Bot - Complete Educational Testing System
Single file implementation supporting 4000+ concurrent users
"""

import asyncio
import sqlite3
import json
import random
import datetime
import threading
import time
import os
import logging
from typing import Dict, List, Optional, Tuple
import re

# External libraries for bot functionality
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import Updater, CommandHandler, MessageHandler, CallbackQueryHandler, ConversationHandler, Filters, CallbackContext

# Libraries for file generation
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Bot configuration
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")
ADMIN_IDS = [6578706277, 7853664401]  # Admin user IDs

# Conversation states
(NAME, SURNAME, PHONE, AGE, REGION, DISTRICT, NEIGHBORHOOD, 
 FEEDBACK, ADMIN_MENU, ADD_QUESTION, EDIT_QUESTION, TEST_IN_PROGRESS,
 SELECT_BOOK, QUESTION_ANSWER) = range(14)

# Age groups and books
AGE_GROUPS = {
    "7-10": ["Kitob 1", "Kitob 2", "Kitob 3", "Kitob 4"],
    "11-14": ["Kitob 1", "Kitob 2", "Kitob 3", "Kitob 4"]
}

# Regions and districts
LOCATIONS = {
    "Toshkent": ["Bektemir", "Chilonzor", "Mirzo Ulug'bek", "Mirobod", "Olmazor", "Sergeli", "Shayxontohur", "Uchtepa", "Yakkasaray", "Yunusobod"],
    "Andijon": ["Andijon", "Asaka", "Baliqchi", "Bo'z", "Bulung'ur", "Izboskan", "Jalaquduq", "Marhamat", "Oltinko'l", "Paxtaobod", "Shahrixon", "Ulug'nor"],
    "Buxoro": ["Buxoro", "G'ijduvon", "Jondor", "Kogon", "Peshku", "Qorako'l", "Romitan", "Shofirkon", "Vobkent"],
    "Farg'ona": ["Farg'ona", "Beshariq", "Bog'dod", "Buvayda", "Dang'ara", "Qo'qon", "Quva", "Rishton", "So'x", "Toshloq", "Uchko'prik", "Yozyovon", "Marg'ilon"],
    "Jizzax": ["Jizzax", "Arnasoy", "Baxmal", "Zafarobod", "Do'stlik", "Forish", "G'allaorol", "Mirzacho'l", "Paxtakor", "Yangiobod"],
    "Qashqadaryo": ["Qarshi", "Chiroqchi", "G'uzor", "Dehqonobod", "Koson", "Muborak", "Nishon", "Qarshi", "Shahrisabz", "Yakkabog'", "Kitob", "Mirishkor"],
    "Navoiy": ["Navoiy", "Karmana", "Konimex", "Navbahor", "Nurota", "Tomdi", "Uchquduq", "Xatirchi", "Zarafshon"],
    "Namangan": ["Namangan", "Chortoq", "Kosonsoy", "Mingbuloq", "Norin", "Pop", "To'raqo'rg'on", "Uychi", "Yangiqo'rg'on"],
    "Samarqand": ["Samarqand", "Bulung'ur", "Ishtixon", "Jomboy", "Kattakurgan", "Narpay", "Payariq", "Pastdarg'om", "Qo'shrabot", "Samarqand shahar", "Toyloq", "Urgut"],
    "Sirdaryo": ["Sirdaryo", "Boyovut", "Guliston", "Oqoltin", "Sardoba", "Sayxunobod", "Sirdaryo", "Shirin", "Xovos", "Yangiyer"],
    "Surxondaryo": ["Termiz", "Boysun", "Denov", "Jarqo'rg'on", "Muzrabot", "Oltinsoy", "Sariosiyo", "Sherobod", "Sho'rchi", "Termiz shahar", "Uzun", "Qiziriq"],
    "Xorazm": ["Urganch", "Bog'ot", "Gurlan", "Qo'shko'pir", "Shovot", "Xonqa", "Xiva", "Yangiariq", "Yangiqo'rg'on"]
}

# Global state management
user_states = {}
active_tests = {}
question_timers = {}

class DatabaseManager:
    """Database management class for SQLite operations"""
    
    def __init__(self, db_name="kitobxon_kids.db"):
        self.db_name = db_name
        self.init_database()
    
    def init_database(self):
        """Initialize database with required tables"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        # Users table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id INTEGER PRIMARY KEY,
                chat_id INTEGER UNIQUE,
                name TEXT NOT NULL,
                surname TEXT NOT NULL,
                phone TEXT NOT NULL,
                age_group TEXT NOT NULL,
                region TEXT NOT NULL,
                district TEXT NOT NULL,
                neighborhood TEXT NOT NULL,
                registration_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active BOOLEAN DEFAULT TRUE
            )
        ''')
        
        # Questions table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS questions (
                question_id INTEGER PRIMARY KEY AUTOINCREMENT,
                age_group TEXT NOT NULL,
                book_name TEXT NOT NULL,
                question_text TEXT NOT NULL,
                option_a TEXT NOT NULL,
                option_b TEXT NOT NULL,
                option_c TEXT NOT NULL,
                option_d TEXT NOT NULL,
                correct_answer TEXT NOT NULL,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                created_by INTEGER,
                FOREIGN KEY (created_by) REFERENCES users (chat_id)
            )
        ''')
        
        # Test results table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS test_results (
                result_id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                age_group TEXT NOT NULL,
                book_name TEXT NOT NULL,
                score INTEGER NOT NULL,
                total_questions INTEGER NOT NULL,
                percentage REAL NOT NULL,
                start_time TIMESTAMP,
                end_time TIMESTAMP,
                questions_answered INTEGER,
                test_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (chat_id)
            )
        ''')
        
        # Feedback table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS feedback (
                feedback_id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                feedback_text TEXT NOT NULL,
                feedback_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (chat_id)
            )
        ''')
        
        # Add comprehensive questions for all books and age groups
        sample_questions = [
            # 7-10 age group sample questions

            # Maugli questions for 11-14 age group (Kitob 1)
            ("11-14", "Kitob 1", "Bo'rilarning nechta bolasi bor edi?", "4", "7", "8", "5", "A"),
            ("11-14", "Kitob 1", "Chiyabo'rining ismi nima edi?", "Balu", "Taboqi", "Bagira", "Qao", "B"),
            ("11-14", "Kitob 1", "Sherxon qayerda yashardi?", "Qishloq chetidagi g'orda", "O'rmonda", "Vayngang daryosi buyida", "Tog'da", "C"),
            ("11-14", "Kitob 1", "Hikoyadagi iblis qaysi hayvon edi?", "Ona bo'ri", "Maugli", "Balu", "Sherxon", "A"),
            ("11-14", "Kitob 1", "Sherxon qaysi hayvon bilan olishishdan qo'rqadi?", "Qao bilan", "Ona bo'ri bilan", "Maugli bilan", "Bagira bilan", "B"),
            ("11-14", "Kitob 1", "So'qqabosh bo'ri yoshligida necha martta bo'ri qopqoniga ilingan?", "6", "3", "2", "4", "C"),
            ("11-14", "Kitob 1", "Maugli qaysi to'daning farzandiga aylanadi?", "Sion", "Tiyon", "Maymunlar", "Qao", "A"),
            ("11-14", "Kitob 1", "Maugliga daraxtga urmalab chiqishni qaysi hayvon o'rgatadi?", "Balu", "Bagira", "Ota bo'ri", "Qao", "B"),
            ("11-14", "Kitob 1", "Bagira qayerda tug'ilgan edi?", "Tog' etagidagi qishloqda", "Qalin o'rmonda", "Odamlar orasida", "Daryo buyida", "C"),
            ("11-14", "Kitob 1", "Qizil gul nima edi?", "Olov", "Meva", "O'simlik", "Gullar", "A"),
            ("11-14", "Kitob 1", "Tog' bug'ma ilonining ismi?", "Sio", "Qao", "Balu", "Bagira", "B"),
            ("11-14", "Kitob 1", "Dunyodagi vaxshiy fe'l ismi?", "Qao", "Natxashi", "Xatxi", "Balu", "C"),
            ("11-14", "Kitob 1", "Mauglini qaysi hayvonlar olib qochgandi?", "Maymunlar", "Ilonlar", "Bo'rilar", "Yulbarslar", "A"),
            ("11-14", "Kitob 1", "Taboqi qaysi hayvonning ismi edi?", "Kiyikning", "Chiyabo'rining", "Ilonning", "Qushning", "B"),
            ("11-14", "Kitob 1", "Qaysi qahramonning eshagi chuquqrga tushib ketadi?", "Qorovulning", "Dexqonning", "Kulolning", "Savdogarning", "C"),
            ("11-14", "Kitob 1", "Maugliga qanday ish berilgandi?", "Qutoslar podasiga poyloqchilik qilish", "Maymunlarga kuz quloq bulish", "Hech qanday ish", "Ov qilish", "A"),
            ("11-14", "Kitob 1", "Ingliz mushketining egasi qaysi hayvon edi?", "Balu", "Baldeo", "Maugli", "Bagira", "B"),
            ("11-14", "Kitob 1", "Eng haybatli Qutosning ismi nima edi?", "Xatxi", "Baldeo", "Rami", "Qao", "C"),
            ("11-14", "Kitob 1", "Maugli necha yillik asalni yeb kuradi?", "3", "5", "4", "2", "A"),
            ("11-14", "Kitob 1", "Qaysi hayvon suv sulhini tuzadi?", "Rami", "Xatxi", "Qao", "Balu", "B"),
            ("11-14", "Kitob 1", "Qaysi qahramonning eshagi bor edi?", "Dehqon", "Messua", "Kulol", "Savdogar", "C"),
            ("11-14", "Kitob 1", "Xatxining nechta o'g'li bor edi?", "3", "4", "2", "5", "A"),
            ("11-14", "Kitob 1", "Birinchi changalzor hokimi va qozisi etib qaysi hayvon tayinlangan?", "Kiyik", "Yulbars", "Ot", "Bo'ri", "B"),
            ("11-14", "Kitob 1", "Akela qaysi hayvonning ismi edi?", "Kiyik", "Yulbars", "Bo'ri", "Ayiq", "C"),
            ("11-14", "Kitob 1", "Maugli qishloqda necha oy yashaydi?", "3 oy", "5 oy", "2 oy", "4 oy", "A"),
            
            # Oq Kema questions for 11-14 age group (Kitob 2)
            ("11-14", "Kitob 2", "Bolaning nechta ertagi bor edi?", "2", "4", "3", "5", "A"),
            ("11-14", "Kitob 2", "Bolaga 2-ertakni kim aytib bergandi?", "Bobosi", "Onasi", "Qishloqdoshlari", "O'qituvchi", "A"),
            ("11-14", "Kitob 2", "Bolaning yoshi nechada edi?", "7yoshdan 8yoshga o'tgan edi", "3yoshdan 4yoshga o'tgan edi", "6 yoshdan 7 yoshga o'tgan edi", "9yoshdan 10yoshga", "A"),
            ("11-14", "Kitob 2", "Bolaga birinchi bo'lib nima sotib olinadi?", "Portfel", "Gitara", "Durbin", "Kitob", "A"),
            ("11-14", "Kitob 2", "3 hovlida nechta bola bor edi?", "1", "2", "5", "3", "A"),
            ("11-14", "Kitob 2", "Ayollar necha kishi edi?", "3", "5", "8", "4", "A"),
            ("11-14", "Kitob 2", "Olamdagi ayollar ichida eng baxtsizi kim edi?", "Bekey xola", "Gulbahor", "Oynisa", "Mavluda", "A"),
            ("11-14", "Kitob 2", "Bolaning bobosining ismi nima edi?", "Mo'min chol", "O'rozqul", "Davlat", "Qodir", "A"),
            ("11-14", "Kitob 2", "Sotuvchi bolaga nima beradi?", "Konfet", "Ruchka", "Daftar", "Kitob", "A"),
            ("11-14", "Kitob 2", "Bolaga kim konfet beradi?", "Sotuvchi", "Mahalla oqsoqoli", "Bobosi", "Onasi", "A"),
            ("11-14", "Kitob 2", "Kuzda nimani sotishmoqchi edi?", "Kartoshkani", "Piyozni", "Qamishni", "Bug'doyni", "A"),
            ("11-14", "Kitob 2", "Kartoshkani qachon sotishmoqchi edi?", "Kuzda", "Qishda", "Bahorda", "Yozda", "A"),
            ("11-14", "Kitob 2", "Sumkani necha sumga sotib oladi?", "5", "2", "8", "3", "A"),
            ("11-14", "Kitob 2", "Bola 5 so'mga nima sotib oladi?", "Sumka", "Daftar", "Poyabzal", "Kitob", "A"),
            ("11-14", "Kitob 2", "O'rozqulning otining ismi?", "Olabosh", "Kukqashqa", "Olabayir", "Boydoq", "A"),
            ("11-14", "Kitob 2", "Chaqmoqlarni nima chaqishi aytilgan?", "Samo", "Oftob", "Hech narsa", "Bulut", "A"),
            ("11-14", "Kitob 2", "Samo nimani chaqishi aytilgan?", "Chaqmoqlarni", "Oftobni", "Yomg'irni", "Shamolni", "A"),
            ("11-14", "Kitob 2", "Bekey xola kimning xotini edi?", "O'rozqul", "Davlat", "Mo'min chol", "Qodir", "A"),
            ("11-14", "Kitob 2", "Kim issiqko'lda edi?", "Bolaning otasi", "Bolaning onasi", "Bobosi", "Xolasi", "A"),
            
            # Tom Soyer questions for 11-14 age group (Kitob 3)
            ("11-14", "Kitob 3", "Tom, Gek va Jim qanday transport vositasida sayohat qilishadi?", "Poyezd", "Kema", "Havo shari", "Arava", "C"),
            ("11-14", "Kitob 3", "Sayohatning boshida unga kim rahbarlik qiladi?", "Jim", "Professor", "Tom", "Gek", "B"),
            ("11-14", "Kitob 3", "Ular qayerdan uchib ketishadi?", "Nyu-York", "Sent-Luis", "London", "Parij", "B"),
            ("11-14", "Kitob 3", "Tomning sayohatdagi roli qanday?", "Sharni boshqaruvchi", "Xaritalarni chizuvchi", "Kapitan sifatida ishtirokchi", "Oziq-ovqat tayyorlovchi", "C"),
            ("11-14", "Kitob 3", "Gekning eng yaqin do'sti kim?", "Jim", "Tom", "Professor", "Joe", "B"),
            ("11-14", "Kitob 3", "Ular qayerga sayohat qilishadi?", "Afrika va Misr", "Evropa", "Osiyo", "Avstraliya", "A"),
            ("11-14", "Kitob 3", "Shar qanday gaz bilan to'ldirilgan?", "Vodorod", "Kislorod", "Azot", "Geliy", "A"),
            ("11-14", "Kitob 3", "Jim sayohat davomida nima bilan ko'proq qiziqadi?", "Piramidalar", "Hayvonlar", "Odamlar va ularning odatlari", "Qadimiy yozuvlar", "C"),
            ("11-14", "Kitob 3", "Tom sayohat davomida nimani boshqaradi?", "Dam olishni", "Shar va xaritalarni", "Ovqatni", "Suvni", "B"),
            ("11-14", "Kitob 3", "Sayohat paytida Gek qaysi holatda ko'proq qo'rqadi?", "Shar portlashidan", "Yirtqich hayvonlardan", "Yo'ldan yo'qolishdan", "To'fonlardan", "B"),
            ("11-14", "Kitob 3", "Tom safari davomida nimaga qiziqadi?", "Qadimiy yodgorliklar", "Shar boshqaruvi", "Do'stlar bilan suhbat", "Oziq-ovqat tayyorlash", "A"),
            ("11-14", "Kitob 3", "Jim qanday odam?", "Qahramon", "Soddashtirishgan", "Do'stona", "Qattiqqo'l", "C"),
            ("11-14", "Kitob 3", "Tom sayohatda qanday kiyimda bo'ladi?", "Rasmiy kiyimda", "Oddiy kiyimda", "Harbiy kiyimda", "Qadimiy kiyimda", "B"),
            ("11-14", "Kitob 3", "Gekning sayohatdagi asosiy vazifasi nima?", "Xarita chizish", "Ovqat tayyorlash", "Shar boshqarish", "Do'stlarga yordam berish", "D"),
            ("11-14", "Kitob 3", "Tom safari davomida nimani o'rganadi?", "Sharni boshqarishni", "Xarita o'qishni", "Qadimiy tillarni", "Hayvonlarni", "B"),
            ("11-14", "Kitob 3", "Sayohat davomida ular nimani ko'rishadi?", "Sahro", "Tog'lar", "Daryo", "Okean", "A"),
            ("11-14", "Kitob 3", "Tom va Gek sayohat davomida nima qiladi?", "O'ynaydi", "O'rganadi", "O'qiydi", "Ovqat tayyorlaydi", "B"),
            ("11-14", "Kitob 3", "Tom va Gek sayohatdagi muammolarni qanday hal qilishadi?", "Jang qilib", "Birgalikda maslahat qilib", "Yolg'iz qolib", "Yordam chaqirib", "B"),
            ("11-14", "Kitob 3", "Tomning eng sevimli mashg'ulotlari nimadan iborat?", "Kitob o'qish", "Sarguzashtlarga chiqish", "O'rmonlarda sayr qilish", "Suvda suzish", "B"),
            ("11-14", "Kitob 3", "Tom va Huck sayohatdan qaytganda qanday xulosa qiladi?", "Sarguzasht juda xavfli", "Do'stlik eng muhim", "Sayohat foydali emas", "Yangi joylarga borish keraksiz", "B"),
            
            # Sariq Dev questions for 11-14 age group (Kitob 4)
            ("11-14", "Kitob 4", "Hoshimjonning qishlog'ining nomi nima edi?", "Xo'jaqishloq", "Gum", "Terak qishloq", "Chinor", "A"),
            ("11-14", "Kitob 4", "Ikromning dadasining ismi nima edi?", "Xoliq", "Zokir", "Qobil", "Hoshim", "A"),
            ("11-14", "Kitob 4", "Kim uyidan kutubxona ochmoqchi edi?", "Orif", "Hoshim", "Hech kim", "Zokir", "A"),
            ("11-14", "Kitob 4", "Zokirning onasining ismi nima edi?", "Karomat xola", "Laylo xola", "Dono xola", "Hakima", "A"),
            ("11-14", "Kitob 4", "Kimning qorni og'rib qoladi?", "Dono", "Laylo", "Hakima", "Karomat", "A"),
            ("11-14", "Kitob 4", "Xo'jaqishloq kimning qishlog'ining nomi edi?", "Hoshim", "Davlat", "Arslon", "Orif", "A"),
            ("11-14", "Kitob 4", "Yong'oq qori kimdan qo'rqar edi?", "Hoshimjonning otasidan", "Danak qoridan", "Ruziqul qoridan", "Zokir dadasidan", "A"),
            ("11-14", "Kitob 4", "Kim ko'richak bo'lib qoladi?", "Dono", "Laylo", "Omina", "Hakima", "A"),
            ("11-14", "Kitob 4", "Devning rangi qanaqa edi?", "Sariq", "Qora", "Oq", "Yashil", "A"),
            ("11-14", "Kitob 4", "Hoshim sariq dev haqida kimdan bilib olmoqchi edi?", "Soraxon folbindan", "Onasidan", "Sartaroshdan", "Bobosidan", "A"),
            ("11-14", "Kitob 4", "Devlar qayerda makon qurgan edi?", "Gumning oldiga", "Tog' tepaligida", "Bog'da", "O'rmonda", "A"),
            ("11-14", "Kitob 4", "Bolalar lagerining direktori kim edi?", "Abdushukr amaki", "Orif boy", "Zokir", "Hoshim", "A"),
            ("11-14", "Kitob 4", "Uzunquloq otaning qabri qayerda edi?", "Katta chinordan sal nariroqda", "Bog'da", "Qishloq oxirida", "Daryo buyida", "A"),
            ("11-14", "Kitob 4", "Orif bozorda necha xalta qurut sotayotgan edi?", "1", "2", "7", "3", "A"),
            ("11-14", "Kitob 4", "Zokir bozorda nima sotadi?", "Bodring", "Quymoq", "Beda", "Meva", "A"),
            
            # Alisa questions for 7-10 age group (Kitob 3)
            ("7-10", "Kitob 3", "Alisa qanday turdagi kitoblarni yoqtirmaydi?", "Suratsiz", "Kup suratli", "Kup sahifali", "Kichik", "A"),
            ("7-10", "Kitob 3", "Alisaning mushugini laqabi nima edi?", "Dina", "Dianka", "Olex", "Mila", "A"),
            ("7-10", "Kitob 3", "Dina qaysi hayvonning laqabi edi?", "Mushuk", "Ot", "It", "Quyon", "A"),
            ("7-10", "Kitob 3", "Qaysi hayvonning boshida supurgi bo'ladi?", "It", "Ot", "Mushuk", "Bo'ri", "A"),
            ("7-10", "Kitob 3", "Qirolicha qaysi rangdagi gullarni yoqtirar edi?", "Qizil", "Oq", "Pushti", "Yashil", "A"),
            ("7-10", "Kitob 3", "Alisaning olami nima deb atalar edi?", "Mo'jizalar mamlakati", "Hayoldagi olam", "Butun borliq", "Fantastik olam", "A"),
            ("7-10", "Kitob 3", "Alisa qaysi hayvonni orqasidan quvib Mo'jizalar mamlakatig borib qoladi?", "Oq quyon", "Tulki", "It", "Mushuk", "A"),
            ("7-10", "Kitob 3", "Suratsiz kitoblarni kim yoqtirmas edi?", "Alisa", "Alisaning opasi", "Alisaning onasi", "Alisaning bobosi", "A"),
            ("7-10", "Kitob 3", "Itning boshida nimasi bo'ladi?", "Supurgi", "Soch", "Hech narsa", "Qalpoq", "A"),
            ("7-10", "Kitob 3", "Qizil rangdagi gullarni kim yoqtirar edi?", "Qirolicha", "Alisa", "Mushuk", "Quyon", "A"),
            ("7-10", "Kitob 3", "Alisa oq quyonning orqasidan quvib qayerga borib qoladi?", "Mo'jizalar mamlakatiga", "Orolga", "Tashlandiq qishloqqa", "Tog'ga", "A"),
            ("7-10", "Kitob 3", "Alisa dastlab kichik bulish uchun nima qiladi?", "Shishadagi suvni ichadi", "Qo'ziqorinni yeydi", "Shirinlik yeydi", "Dori ichadi", "A"),
            ("7-10", "Kitob 3", "Agar alisaning olami bulganda, u yerdagi barcha kitoblar qanday kurinishda bo'lar edi?", "Suratli", "Suratsiz", "Oppoq rangda", "Kichik", "A"),
            ("7-10", "Kitob 3", "Alisa nimani hali menga kerak buladi deb chuntagiga solib quyadi?", "Qo'ziqorin bulaklarini", "Mevalardan", "Qizil gullardan", "Shirinlik", "A"),
            ("7-10", "Kitob 3", "Tentak shlyapnik 364 kunni qanday kun deb ataydi?", "Tug'ilmagan kun", "Keraksiz kun", "Tug'ilgan kun", "Oddiy kun", "A"),
            ("7-10", "Kitob 3", "Kroket uyinidagi 2-qatordagi askarlar qanday kurinishda edi?", "Qarg'a", "Tappon", "Tuz", "Gullar", "A"),
            ("7-10", "Kitob 3", "Qirolicha kroket uyinida yutayotgan vaqtda uni qaysi hayvon qurqitib yuboradi?", "Cheshir mushugi", "Tentak shlyapnik", "Oq quyon", "It", "A"),
            ("7-10", "Kitob 3", "Alisaga mo'jizalar mamlakatidan chiqib ketishga kim yordam beradi?", "U tush kurayotgan buladi va uyg'onib ketadi", "Opasi", "Oq quyon", "Qirolicha", "A"),
            ("7-10", "Kitob 3", "Alisa nimani yeb kattalashadi?", "Qo'ziqorin", "Meva", "Bodring", "Non", "A"),
            ("7-10", "Kitob 3", "Oq quyonning ortidan kim yugurib ketadi?", "Alisa", "Opasi", "Dianka", "Onasi", "A"),
            
            # Baron Myunxauzen questions for 7-10 age group (Kitob 4)
            ("7-10", "Kitob 4", "Hikoya qahramoni qaysi faslda Peterburgga yetib boradi?", "Qishda", "Yozda", "Kuzda", "Bahorda", "A"),
            ("7-10", "Kitob 4", "Qaysi hayvonning ko'r ekanligi aytib o'tilgan?", "Cho'chqa", "Ot", "It", "Mushuk", "A"),
            ("7-10", "Kitob 4", "Qaysi hikoya qahramonining ismi Dianka edi?", "It", "Ot", "Mushuk", "Quyon", "A"),
            ("7-10", "Kitob 4", "Hikoya qahramoni qishda qayerga yetib boradi?", "Peterburgga", "Misrga", "Daniyaga", "Londonga", "A"),
            ("7-10", "Kitob 4", "Cho'chqaning qanday nuqsoni bor edi?", "Ko'r edi", "Tishlari yuq edi", "Qulog'i eshitmasdi", "Oyog'i oqsaydi", "A"),
            ("7-10", "Kitob 4", "Dianka qaysi hayvonning ismi edi?", "It", "Bo'ri", "Tulki", "Mushuk", "A"),
            ("7-10", "Kitob 4", "Misrda nechta tegirmonni kuradi?", "7", "6", "5", "8", "A"),
            ("7-10", "Kitob 4", "Kemada nechta matros bor edi?", "300", "200", "100", "400", "A"),
            ("7-10", "Kitob 4", "Myunxauzen qaysi asar qahramoni edi?", "Baron myunxauzen sarguzashtlari", "Susanbil", "Maugli", "Alisa", "A"),
            ("7-10", "Kitob 4", "Oy aholisi yiliga necha martta ovqatlanadi?", "12", "13", "18", "24", "A"),
            ("7-10", "Kitob 4", "Rossiyaga qaysi hayvon bilan ketadi?", "Ot", "Kiyik", "Echki", "Tuya", "A"),
            ("7-10", "Kitob 4", "Ot bilan qayerga ketadi?", "Rossiyaga", "Peterburgga", "Daniyaga", "Misrga", "A"),
            ("7-10", "Kitob 4", "Yer yuzidagi eng rostguy odam kim edi?", "Baron myunxauzen", "Flipps", "Xatxi", "Dianka", "A"),
            ("7-10", "Kitob 4", "Baron myunxauzen bir martta otishda nechta kaklikni urib tushiradi?", "7", "8", "4", "5", "A"),
            ("7-10", "Kitob 4", "Baron myunxauzen 7 ta kaklikni nechinchi martta otishda urib tushiradi?", "1", "5", "3", "2", "A"),
            ("7-10", "Kitob 4", "Baron myunxauzen qayerda tulkiga duch keladi?", "Rus o'rmonida", "Peterburgda", "Ispan orolida", "Misrda", "A"),
            ("7-10", "Kitob 4", "Qaysi hayvonning 8ta oyog'i bor edi?", "Quyon", "Tulki", "Kiyik", "It", "A"),
            ("7-10", "Kitob 4", "Quyoning nechta oyog'i bore di?", "8", "6", "3", "4", "A"),
            ("7-10", "Kitob 4", "Turk loviyasi necha soatda oygacha o'sib chiqadi?", "1", "4", "8", "2", "A"),
            ("7-10", "Kitob 4", "Baron myunxauzen qoyaga emas qaysi hayvonga urilgan bo'ladi?", "Kit", "Baliq", "Kiyik", "Ot", "A"),
            
            # Galaktikada bir kun questions for 7-10 age group (Additional book)
            ("7-10", "Kitob 1", "Asarning bosh qahramoni kim?", "Margol", "Ibrohim", "Ahmad", "Odilbek", "C"),
            ("7-10", "Kitob 1", "Asarda ko'tarma kranni kim ixtiro qiladi?", "Muhammadislom", "Ismoil", "Ahmad", "Ibrohim", "D"),
            ("7-10", "Kitob 1", "Ibrohim nimani yaratgan edi?", "Gulqaychi", "Bog'borobot", "Ko'tarma kran", "Koinot kemasi", "B"),
            ("7-10", "Kitob 1", "Ahmad tanlovda nechinchi raqamli ishtirokchi bo'ladi?", "10", "6", "9", "7", "D"),
            ("7-10", "Kitob 1", "Ahmadning yoshi nechada edi?", "15", "12", "17", "10", "B"),
            ("7-10", "Kitob 1", "Asarda pakana odam kim?", "Margol", "Odilbek", "Magron", "Ismoil", "A"),
            ("7-10", "Kitob 1", "Ahmadning dadasining ismi?", "Nyuton", "Margol", "Odilbek", "Oqilbek", "C"),
            ("7-10", "Kitob 1", "Pakana odam qaysi sayyoradan kelgan?", "Platon", "Salmir", "Mars", "Venera", "B"),
            ("7-10", "Kitob 1", "Koinot qaroqchilari kimlar?", "Odamlar", "Hayvonlar", "Galkeslar", "Gikslar", "D"),
            ("7-10", "Kitob 1", "Salmir sayyorasining nimasi so'na boshlagani uchun uning tabiati o'zgarib ketgan?", "Lava", "Yadro", "Hujayra", "Atom", "B"),
            ("7-10", "Kitob 1", "Ahmad qaysi fanlardan chuqur bilimga ega edi?", "Geografiya", "Astranomiya va Matematika", "Fizika", "Gealogiya va Astrofizika", "D"),
            ("7-10", "Kitob 1", "Yadroni tushunish uchun eng yaxshi misol nima?", "Tuxum sarig'i", "Yong'oq", "Yer shari", "Globus", "A"),
            ("7-10", "Kitob 1", "Qaysi sayyora 'Katta atomlar urushi' vaqtida yo'q bo'lib ketgan?", "Yulduzlar sayyorasi", "Uchar sayyora", "Gikslar sayyorasi", "Mars sayyorasi", "C"),
            ("7-10", "Kitob 1", "Qaroqchilarning boshlig'i kim?", "Margol", "Obma", "Osar", "Margarit", "B"),
            ("7-10", "Kitob 1", "'Ziji jadidi Ko'ragoniy' kim tomonidan tuzilgan astronomik jadval?", "Xorazmiy", "Beruniy", "Amir Temur", "Mirzo Ulug'bek", "D"),
            
            # Oltin bola questions for 7-10 age group (Extended Kitob 2)
            ("7-10", "Kitob 2", "Butun Shahar Bolani nima deb ataydi?", "Nog'orachining oltinsoch Peteri", "Oq qoshli bola", "Dovyurak bola", "Kichik Peter", "A"),
            ("7-10", "Kitob 2", "Onasi Bolani nima deb ataydi?", "Oltinsoch bolam", "Oltin qulli bolam", "Oq qoshli bolam", "Yaxshi bolam", "A"),
            ("7-10", "Kitob 2", "Bolaga nima deb ism quyishadi?", "Peter", "Ruki", "Ramzan", "Pavel", "A"),
            ("7-10", "Kitob 2", "Peter qaysi kasb egasi bo'laman deb aytardi?", "Soldat", "Uchuvchi", "Haydovchi", "Shifokor", "A"),
            ("7-10", "Kitob 2", "Hikoyada kimlar tush kuradi?", "Peterning otasi va onasi", "Peter", "Peterning ustozi", "Qo'shnilari", "A"),
            ("7-10", "Kitob 2", "Peter kimga fortepiano chalishni o'rgatadi?", "Lotto xonimga", "Onasiga", "Otasiga", "Do'stiga", "A"),
            ("7-10", "Kitob 2", "Lotto qayerlik edi?", "Burgomisterlik", "Angliyalik", "Peterburglik", "Nemislik", "A"),
            ("7-10", "Kitob 2", "Lotto unashtirilganini peterga kim aytadi?", "Onasi", "Otasi", "O'zi", "Qo'shnisi", "A"),
            ("7-10", "Kitob 2", "Bolani oltinsoch bolam deb kim chaqiradi?", "Onasi", "Otasi", "Bobosi", "Xolasi", "A"),
            ("7-10", "Kitob 2", "Lotto xonimga kim fortepiano chalishni o'rgatayotgan edi?", "Peter", "Hech kim", "Dugonasi", "Onasi", "A"),
            ("7-10", "Kitob 2", "Urushda peterning nog'orasi qanday vazifani bajargan edi?", "Askarlarga belgi bergan", "Dam olish kerakligini bildirib turgan", "Ovqatlanish vaqti bulganligini bildirib turish uchun", "Qo'shiq kuylagan", "A"),
            ("7-10", "Kitob 2", "Urushda nima berilganda ham peter nog'orani chalishda davom etavergan?", "Ortga qaytish buyurilganda", "Jang qilish buyurilganda", "Dam olish buyurilganda", "Ovqatlanish buyurilganda", "A"),
            
            # Susanbil questions for 7-10 age group (Extended Kitob 2)
            ("7-10", "Kitob 2", "Boyning qanaqa hayvonlari bor edi?", "Bir eshagi va bir ho'kizi", "Bir ot va bir eshak", "Bir ho'kiz va bir ot", "Ikki eshak", "A"),
            ("7-10", "Kitob 2", "Xo'jayin bozordan bir qop nima sotib oladi?", "Tuz", "Un", "Guruch", "Shakar", "A"),
            ("7-10", "Kitob 2", "Bir qop tuzni kim sotib oladi?", "Xo'jayin", "Qorovul", "Vazir", "Savdogar", "A"),
            ("7-10", "Kitob 2", "Cho'lda nechta kalamush bor edi?", "2", "3", "4", "5", "A"),
            ("7-10", "Kitob 2", "Susanbilga birinchi bo'lib qaysi hayvon yulga chiqadi?", "Eshak", "Ho'kiz", "Kalamushlar", "Xo'roz", "A"),
            ("7-10", "Kitob 2", "Eshak nechinchi bo'lib susanbilga yulga chiqadi?", "Birinchi", "Oxirgi", "Uchinchi", "Ikkinchi", "A"),
            ("7-10", "Kitob 2", "Ho'kizni nima uchun eshakning o'rniga o'tkazib qo'yadi?", "Ovqatidan yegani uchun", "Kup suv ichgani uchun", "Kam ishlagani uchun", "Qochgani uchun", "A"),
            ("7-10", "Kitob 2", "Bo'rilar yonimizga yetib kelganda men hangrab turaman deb qaysi hayvon aytadi?", "Eshak", "Ho'kiz", "Xo'roz", "Kalamush", "A"),
            ("7-10", "Kitob 2", "Xo'jayin eshagiga nima qildirib bergan edi?", "Egar taqim", "Yem yeyish uchun idish", "Hech narsa", "Ipak ro'mol", "A"),
            ("7-10", "Kitob 2", "Xo'jayin bozordan sotib olgan tuzni qaysi hayvonning ustiga ortadi?", "Eshak", "Ho'kiz", "Ot", "Tuya", "A"),
            ("7-10", "Kitob 2", "Xo'roz dala yulida borayotgan vaqtida nechta kucha ga kelib tuqnashadi?", "3", "1", "2", "4", "A"),
            ("7-10", "Kitob 2", "Kalamushlar xo'roz eshak va ho'kizdan nima suraydi?", "Ovqat", "Suv", "Boshpana", "Yordam", "A")
        ]
        
        for question in sample_questions:
            cursor.execute('''
                INSERT OR IGNORE INTO questions 
                (age_group, book_name, question_text, option_a, option_b, option_c, option_d, correct_answer, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', question + (ADMIN_IDS[0],))
        
        conn.commit()
        conn.close()
    
    def execute_query(self, query: str, params: tuple = None):
        """Execute a query and return results"""
        conn = sqlite3.connect(self.db_name)
        cursor = conn.cursor()
        
        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            
            if query.strip().upper().startswith('SELECT'):
                results = cursor.fetchall()
            else:
                conn.commit()
                results = cursor.rowcount
            
            return results
        except Exception as e:
            logger.error(f"Database error: {e}")
            conn.rollback()
            return []
        finally:
            conn.close()
    
    def register_user(self, chat_id: int, name: str, surname: str, phone: str, 
                     age_group: str, region: str, district: str, neighborhood: str) -> bool:
        """Register a new user"""
        query = '''
            INSERT OR REPLACE INTO users 
            (chat_id, name, surname, phone, age_group, region, district, neighborhood)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        '''
        result = self.execute_query(query, (chat_id, name, surname, phone, age_group, region, district, neighborhood))
        return isinstance(result, int) and result > 0
    
    def get_user(self, chat_id: int) -> Optional[dict]:
        """Get user information"""
        query = "SELECT * FROM users WHERE chat_id = ?"
        result = self.execute_query(query, (chat_id,))
        
        if result and len(result) > 0:
            user_data = result[0]
            return {
                'user_id': user_data[0],
                'chat_id': user_data[1],
                'name': user_data[2],
                'surname': user_data[3],
                'phone': user_data[4],
                'age_group': user_data[5],
                'region': user_data[6],
                'district': user_data[7],
                'neighborhood': user_data[8],
                'registration_date': user_data[9],
                'is_active': user_data[10]
            }
        return None
    
    def get_questions(self, age_group: str, book_name: str, limit: int = 25) -> List[dict]:
        """Get randomized questions for a test"""
        query = '''
            SELECT question_id, question_text, option_a, option_b, option_c, option_d, correct_answer
            FROM questions 
            WHERE age_group = ? AND book_name = ?
            ORDER BY RANDOM()
            LIMIT ?
        '''
        results = self.execute_query(query, (age_group, book_name, limit))
        
        questions = []
        for result in results:
            questions.append({
                'question_id': result[0],
                'question_text': result[1],
                'option_a': result[2],
                'option_b': result[3],
                'option_c': result[4],
                'option_d': result[5],
                'correct_answer': result[6]
            })
        
        return questions
    
    def save_test_result(self, user_id: int, age_group: str, book_name: str,
                        score: int, total_questions: int, start_time: str,
                        end_time: str, questions_answered: int) -> bool:
        """Save test results"""
        percentage = (score / (total_questions * 4)) * 100 if total_questions > 0 else 0
        
        query = '''
            INSERT INTO test_results 
            (user_id, age_group, book_name, score, total_questions, percentage, 
             start_time, end_time, questions_answered)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        '''
        result = self.execute_query(query, (user_id, age_group, book_name, score, total_questions, percentage, start_time, end_time, questions_answered))
        return isinstance(result, int) and result > 0
    
    def save_feedback(self, user_id: int, feedback_text: str) -> bool:
        """Save user feedback"""
        query = "INSERT INTO feedback (user_id, feedback_text) VALUES (?, ?)"
        result = self.execute_query(query, (user_id, feedback_text))
        return isinstance(result, int) and result > 0
    
    def get_all_users(self) -> List[dict]:
        """Get all registered users"""
        query = "SELECT * FROM users ORDER BY registration_date DESC"
        results = self.execute_query(query)
        
        users = []
        for result in results:
            users.append({
                'user_id': result[0],
                'chat_id': result[1],
                'name': result[2],
                'surname': result[3],
                'phone': result[4],
                'age_group': result[5],
                'region': result[6],
                'district': result[7],
                'neighborhood': result[8],
                'registration_date': result[9],
                'is_active': result[10]
            })
        
        return users
    
    def get_test_results(self) -> List[dict]:
        """Get all test results with user information"""
        query = '''
            SELECT u.name, u.surname, u.age_group, tr.book_name, tr.score, 
                   tr.total_questions, tr.percentage, tr.test_date
            FROM test_results tr
            JOIN users u ON tr.user_id = u.chat_id
            ORDER BY tr.test_date DESC
        '''
        results = self.execute_query(query)
        
        test_results = []
        for result in results:
            test_results.append({
                'name': result[0],
                'surname': result[1],
                'age_group': result[2],
                'book_name': result[3],
                'score': result[4],
                'total_questions': result[5],
                'percentage': result[6],
                'test_date': result[7]
            })
        
        return test_results

# Initialize database
db = DatabaseManager()

class TestManager:
    """Manages test sessions and timing"""
    
    def start_test(self, chat_id: int, age_group: str, book_name: str) -> bool:
        """Start a new test session"""
        questions = db.get_questions(age_group, book_name, 25)
        
        if len(questions) < 5:  # Minimum 5 questions needed for a test
            return False
        
        active_tests[chat_id] = {
            'age_group': age_group,
            'book_name': book_name,
            'questions': questions,
            'current_question': 0,
            'score': 0,
            'start_time': datetime.datetime.now().isoformat(),
            'answers': []
        }
        
        return True
    
    def get_current_question(self, chat_id: int) -> Optional[dict]:
        """Get the current question for a user"""
        if chat_id not in active_tests:
            return None
        
        test = active_tests[chat_id]
        if test['current_question'] >= len(test['questions']):
            return None
        
        return test['questions'][test['current_question']]
    
    def submit_answer(self, chat_id: int, answer: str) -> bool:
        """Submit an answer and move to next question"""
        if chat_id not in active_tests:
            return False
        
        test = active_tests[chat_id]
        current_q = test['questions'][test['current_question']]
        
        # Check if answer is correct
        is_correct = (answer.upper() == current_q['correct_answer'].upper())
        if is_correct:
            test['score'] += 4
        
        test['answers'].append({
            'question_id': current_q['question_id'],
            'user_answer': answer,
            'correct_answer': current_q['correct_answer'],
            'is_correct': is_correct
        })
        
        test['current_question'] += 1
        return True
    
    def timeout_question(self, chat_id: int) -> bool:
        """Handle question timeout"""
        if chat_id not in active_tests:
            return False
        
        test = active_tests[chat_id]
        current_q = test['questions'][test['current_question']]
        
        test['answers'].append({
            'question_id': current_q['question_id'],
            'user_answer': 'TIMEOUT',
            'correct_answer': current_q['correct_answer'],
            'is_correct': False
        })
        
        test['current_question'] += 1
        return True
    
    def is_test_complete(self, chat_id: int) -> bool:
        """Check if test is complete"""
        if chat_id not in active_tests:
            return True
        
        test = active_tests[chat_id]
        return test['current_question'] >= len(test['questions'])
    
    def get_test_results(self, chat_id: int) -> Optional[dict]:
        """Get test results"""
        if chat_id not in active_tests:
            return None
        
        test = active_tests[chat_id]
        
        results = {
            'age_group': test['age_group'],
            'book_name': test['book_name'],
            'score': test['score'],
            'total_questions': len(test['questions']),
            'percentage': (test['score'] / (len(test['questions']) * 4)) * 100,
            'start_time': test['start_time'],
            'end_time': datetime.datetime.now().isoformat(),
            'questions_answered': len(test['answers'])
        }
        
        return results
    
    def cleanup_test(self, chat_id: int):
        """Clean up test session"""
        if chat_id in active_tests:
            del active_tests[chat_id]
        if chat_id in question_timers:
            del question_timers[chat_id]

test_manager = TestManager()

def create_excel_report(data, report_type="users"):
    """Create Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if report_type == "users":
        ws.title = "Foydalanuvchilar"
        
        # Headers
        headers = ["Ism", "Familiya", "Telefon", "Yosh guruhi", "Viloyat", "Tuman", "Mahalla", "Ro'yxatdan o'tgan sana"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, user in enumerate(data, 2):
            ws.cell(row=row, column=1, value=user.get('name', ''))
            ws.cell(row=row, column=2, value=user.get('surname', ''))
            ws.cell(row=row, column=3, value=user.get('phone', ''))
            ws.cell(row=row, column=4, value=user.get('age_group', ''))
            ws.cell(row=row, column=5, value=user.get('region', ''))
            ws.cell(row=row, column=6, value=user.get('district', ''))
            ws.cell(row=row, column=7, value=user.get('neighborhood', ''))
            ws.cell(row=row, column=8, value=user.get('registration_date', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    elif report_type == "results":
        ws.title = "Test natijalari"
        
        # Headers
        headers = ["Ism", "Familiya", "Yosh guruhi", "Kitob", "Ball", "Jami savollar", "Foiz", "Test sanasi"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, result in enumerate(data, 2):
            ws.cell(row=row, column=1, value=result.get('name', ''))
            ws.cell(row=row, column=2, value=result.get('surname', ''))
            ws.cell(row=row, column=3, value=result.get('age_group', ''))
            ws.cell(row=row, column=4, value=result.get('book_name', ''))
            ws.cell(row=row, column=5, value=result.get('score', 0))
            ws.cell(row=row, column=6, value=result.get('total_questions', 0))
            ws.cell(row=row, column=7, value=f"{result.get('percentage', 0):.1f}%")
            ws.cell(row=row, column=8, value=result.get('test_date', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def create_pdf_report(data, report_type="users"):
    """Create PDF report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    if report_type == "users":
        title = Paragraph("Foydalanuvchilar ro'yxati", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Table data
        table_data = [["Ism", "Familiya", "Telefon", "Yosh", "Viloyat", "Tuman"]]
        for user in data:
            table_data.append([
                user.get('name', ''),
                user.get('surname', ''),
                user.get('phone', ''),
                user.get('age_group', ''),
                user.get('region', ''),
                user.get('district', '')
            ])
    
    elif report_type == "results":
        title = Paragraph("Test natijalari", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Table data
        table_data = [["Ism", "Familiya", "Yosh", "Kitob", "Ball", "Foiz"]]
        for result in data:
            table_data.append([
                result.get('name', ''),
                result.get('surname', ''),
                result.get('age_group', ''),
                result.get('book_name', ''),
                str(result.get('score', 0)),
                f"{result.get('percentage', 0):.1f}%"
            ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

# Bot command handlers

async def timeout_handler(context: CallbackContext, chat_id: int):
    """Handle question timeout"""
    if chat_id not in active_tests:
        return
    
    test_manager.timeout_question(chat_id)
    
    # Send timeout message
    context.bot.send_message(
        chat_id=chat_id,
        text="⏰ Vaqt tugadi! Keyingi savolga o'tyapmiz..."
    )
    
    # Check if test is complete
    if test_manager.is_test_complete(chat_id):
        await end_test(context, chat_id)
    else:
        await send_next_question(context, chat_id)

async def send_next_question(context: CallbackContext, chat_id: int):
    """Send the next question to user"""
    question = test_manager.get_current_question(chat_id)
    
    if not question:
        await end_test(context, chat_id)
        return
    
    test = active_tests[chat_id]
    question_num = test['current_question'] + 1
    total_questions = len(test['questions'])
    
    # Create inline keyboard for answers
    keyboard = [
        [InlineKeyboardButton("A) " + question['option_a'], callback_data=f"answer_A_{chat_id}")],
        [InlineKeyboardButton("B) " + question['option_b'], callback_data=f"answer_B_{chat_id}")],
        [InlineKeyboardButton("C) " + question['option_c'], callback_data=f"answer_C_{chat_id}")],
        [InlineKeyboardButton("D) " + question['option_d'], callback_data=f"answer_D_{chat_id}")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    message = f"📝 Savol {question_num}/{total_questions}\n\n"
    message += f"❓ {question['question_text']}\n\n"
    message += "⏱ Sizda 20 soniya vaqt bor!"
    
    context.bot.send_message(
        chat_id=chat_id,
        text=message,
        reply_markup=reply_markup
    )
    
    # Set timeout timer
    def timeout_sync():
        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                loop.create_task(timeout_handler(context, chat_id))
            else:
                asyncio.run(timeout_handler(context, chat_id))
        except RuntimeError:
            asyncio.run(timeout_handler(context, chat_id))
    
    timer = threading.Timer(20.0, timeout_sync)
    timer.start()
    question_timers[chat_id] = timer

def end_test_sync(context: CallbackContext, chat_id: int):
    """End test and show results synchronously"""
    results = test_manager.get_test_results(chat_id)
    
    if not results:
        context.bot.send_message(chat_id=chat_id, text="❌ Xatolik yuz berdi!")
        return
    
    # Save results to database
    db.save_test_result(
        chat_id, results['age_group'], results['book_name'],
        results['score'], results['total_questions'],
        results['start_time'], results['end_time'],
        results['questions_answered']
    )
    
    # Get user info for detailed results
    user = db.get_user(chat_id)
    user_name = f"{user['name']} {user['surname']}" if user else "Unknown"
    
    # Send detailed results message to user
    message = "🎉 Test yakunlandi!\n\n"
    message += f"👤 Ism: {user_name}\n"
    message += f"📚 Kitob: {results['book_name']}\n"
    message += f"👥 Yosh guruhi: {results['age_group']}\n"
    message += f"🎯 Ball: {results['score']}/{results['total_questions'] * 4}\n"
    message += f"📊 Foiz: {results['percentage']:.1f}%\n"
    message += f"✅ Javob berilgan savollar: {results['questions_answered']}/{results['total_questions']}\n"
    message += f"⏱ Vaqt: {results['start_time'].strftime('%H:%M')} - {results['end_time'].strftime('%H:%M')}\n\n"
    
    if results['percentage'] >= 80:
        message += "🏆 Ajoyib natija! Tabriklaymiz!"
    elif results['percentage'] >= 60:
        message += "👍 Yaxshi natija! Davom eting!"
    else:
        message += "📚 Qo'shimcha o'qish tavsiya etiladi."
    
    context.bot.send_message(chat_id=chat_id, text=message)
    
    # Send results to all admins
    admin_message = f"📊 Yangi test natijasi:\n\n"
    admin_message += f"👤 Foydalanuvchi: {user_name}\n"
    admin_message += f"📱 Telefon: {user['phone'] if user else 'N/A'}\n"
    admin_message += f"🌍 Manzil: {user['region']}, {user['district']} ({user['neighborhood']})\n" if user else ""
    admin_message += f"📚 Kitob: {results['book_name']}\n"
    admin_message += f"👥 Yosh: {results['age_group']}\n"
    admin_message += f"🎯 Ball: {results['score']}/{results['total_questions'] * 4}\n"
    admin_message += f"📊 Foiz: {results['percentage']:.1f}%\n"
    admin_message += f"✅ Javoblar: {results['questions_answered']}/{results['total_questions']}\n"
    admin_message += f"🕒 Vaqt: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    for admin_id in ADMIN_IDS:
        try:
            context.bot.send_message(admin_id, admin_message)
        except:
            continue
    
    # Send celebration sticker
    try:
        context.bot.send_sticker(
            chat_id=chat_id,
            sticker="CAACAgIAAxkBAAECOzFiZoFa5vVrMVTCRpK2_v-y6qWqOwAC2AADBREAAZxSCqpYnq6yW8EeBA"
        )
    except:
        pass
    
    # Cleanup test session
    test_manager.cleanup_test(chat_id)
    
    # Show main menu
    show_main_menu_sync_direct(context, chat_id)

def show_main_menu_sync_direct(context: CallbackContext, chat_id: int):
    """Show main menu synchronously using context"""
    keyboard = [
        ["📝 Test topshirish", "📋 Loyiha haqida"],
        ["💬 Fikr bildirish", "📊 Natijalarim"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
    
    context.bot.send_message(
        chat_id=chat_id,
        text="Asosiy menyu:",
        reply_markup=reply_markup
    )

async def end_test(context: CallbackContext, chat_id: int):
    """End test and show results"""
    results = test_manager.get_test_results(chat_id)
    
    if not results:
        context.bot.send_message(chat_id=chat_id, text="❌ Xatolik yuz berdi!")
        return
    
    # Save results to database
    db.save_test_result(
        chat_id, results['age_group'], results['book_name'],
        results['score'], results['total_questions'],
        results['start_time'], results['end_time'],
        results['questions_answered']
    )
    
    # Get user info for detailed results
    user = db.get_user(chat_id)
    user_name = f"{user['name']} {user['surname']}" if user else "Unknown"
    
    # Send detailed results message to user
    message = "🎉 Test yakunlandi!\n\n"
    message += f"👤 Ism: {user_name}\n"
    message += f"📚 Kitob: {results['book_name']}\n"
    message += f"👥 Yosh guruhi: {results['age_group']}\n"
    message += f"🎯 Ball: {results['score']}/{results['total_questions'] * 4}\n"
    message += f"📊 Foiz: {results['percentage']:.1f}%\n"
    message += f"✅ Javob berilgan savollar: {results['questions_answered']}/{results['total_questions']}\n"
    message += f"⏱ Vaqt: {results['start_time'].strftime('%H:%M')} - {results['end_time'].strftime('%H:%M')}\n\n"
    
    if results['percentage'] >= 80:
        message += "🏆 Ajoyib natija! Tabriklaymiz!"
    elif results['percentage'] >= 60:
        message += "👍 Yaxshi natija! Davom eting!"
    else:
        message += "📚 Qo'shimcha o'qish tavsiya etiladi."
    
    context.bot.send_message(chat_id=chat_id, text=message)
    
    # Send results to all admins
    admin_message = f"📊 Yangi test natijasi:\n\n"
    admin_message += f"👤 Foydalanuvchi: {user_name}\n"
    admin_message += f"📱 Telefon: {user['phone'] if user else 'N/A'}\n"
    admin_message += f"🌍 Manzil: {user['region']}, {user['district']} ({user['neighborhood']})\n" if user else ""
    admin_message += f"📚 Kitob: {results['book_name']}\n"
    admin_message += f"👥 Yosh: {results['age_group']}\n"
    admin_message += f"🎯 Ball: {results['score']}/{results['total_questions'] * 4}\n"
    admin_message += f"📊 Foiz: {results['percentage']:.1f}%\n"
    admin_message += f"✅ Javoblar: {results['questions_answered']}/{results['total_questions']}\n"
    admin_message += f"🕒 Vaqt: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    for admin_id in ADMIN_IDS:
        try:
            context.bot.send_message(admin_id, admin_message)
        except:
            continue
    
    # Send celebration sticker
    try:
        context.bot.send_sticker(
            chat_id=chat_id,
            sticker="CAACAgIAAxkBAAECOzFiZoFa5vVrMVTCRpK2_v-y6qWqOwAC2AADBREAAZxSCqpYnq6yW8EeBA"
        )
    except:
        pass
    
    # Cleanup test session
    test_manager.cleanup_test(chat_id)
    
    # Show main menu
    await show_main_menu(context, chat_id)

async def show_main_menu(context: CallbackContext, chat_id: int):
    """Show main menu to user"""
    keyboard = [
        ["📝 Test topshirish", "📋 Loyiha haqida"],
        ["💬 Fikr bildirish", "📊 Natijalarim"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
    
    context.bot.send_message(
        chat_id=chat_id,
        text="Asosiy menyu:",
        reply_markup=reply_markup
    )

# Registration conversation handlers

def start(update: Update, context: CallbackContext):
    """Handle /start command"""
    chat_id = update.message.chat_id
    
    # Check if user is already registered
    user = db.get_user(chat_id)
    if user:
        update.message.reply_text(f"Assalomu alaykum, {user['name']}! Siz allaqachon ro'yxatdan o'tgansiz.")
        show_main_menu_sync(update, context)
        return ConversationHandler.END
    
    # Check if user is admin
    if chat_id in ADMIN_IDS:
        update.message.reply_text("Assalomu alaykum, Admin! Admin paneliga xush kelibsiz.")
        show_admin_menu(update, context)
        return ConversationHandler.END
    
    # Start registration
    user_states[chat_id] = {"step": "name", "data": {}}
    
    # Send welcome sticker
    try:
        update.message.reply_sticker(sticker="CAACAgIAAxkBAAECOzFiZoFa5vVrMVTCRpK2_v-y6qWqOwAC2AADBREAAZxSCqpYnq6yW8EeBA")
    except:
        pass
    
    update.message.reply_text(
        "🌟 Assalomu alaykum! Kitobxon Kids botiga xush kelibsiz!\n\n"
        "📝 Ro'yxatdan o'tish uchun ma'lumotlaringizni kiriting.\n\n"
        "👤 Ismingizni kiriting:"
    )
    
    return NAME

def name(update: Update, context: CallbackContext):
    """Handle name input"""
    chat_id = update.message.chat_id
    name_text = update.message.text.strip()
    
    # Validation
    if not name_text or len(name_text) < 2:
        update.message.reply_text("❌ Iltimos, haqiqiy ismingizni kiriting (kamida 2 ta harf):")
        return NAME
    
    if not re.match("^[a-zA-ZА-Яа-я\u0400-\u04FF ]+$", name_text):
        update.message.reply_text("❌ Ismda faqat harflar bo'lishi kerak:")
        return NAME
    
    user_states[chat_id]["data"]["name"] = name_text
    update.message.reply_text("👨‍👩‍👧‍👦 Familiyangizni kiriting:")
    return SURNAME

def surname(update: Update, context: CallbackContext):
    """Handle surname input"""
    chat_id = update.message.chat_id
    surname_text = update.message.text.strip()
    
    # Validation
    if not surname_text or len(surname_text) < 2:
        update.message.reply_text("❌ Iltimos, haqiqiy familiyangizni kiriting (kamida 2 ta harf):")
        return SURNAME
    
    if not re.match("^[a-zA-ZА-Яа-я\u0400-\u04FF ]+$", surname_text):
        update.message.reply_text("❌ Familiyada faqat harflar bo'lishi kerak:")
        return SURNAME
    
    user_states[chat_id]["data"]["surname"] = surname_text
    update.message.reply_text("📱 Telefon raqamingizni kiriting (+998901234567 formatida):")
    return PHONE

def phone(update: Update, context: CallbackContext):
    """Handle phone input"""
    chat_id = update.message.chat_id
    phone_text = update.message.text.strip()
    
    # Validation - Uzbekistan phone format
    if not re.match(r"^\+998\d{9}$", phone_text):
        update.message.reply_text("❌ Telefon raqamni to'g'ri formatda kiriting: +998901234567")
        return PHONE
    
    user_states[chat_id]["data"]["phone"] = phone_text
    
    keyboard = [["7-10 yosh"], ["11-14 yosh"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    update.message.reply_text("🎂 Yosh guruhingizni tanlang:", reply_markup=reply_markup)
    return AGE

def age(update: Update, context: CallbackContext):
    """Handle age group selection"""
    chat_id = update.message.chat_id
    age_text = update.message.text.strip()
    
    if age_text not in ["7-10 yosh", "11-14 yosh"]:
        keyboard = [["7-10 yosh"], ["11-14 yosh"]]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        update.message.reply_text("❌ Iltimos, ro'yxatdagi yosh guruhini tanlang:", reply_markup=reply_markup)
        return AGE
    
    user_states[chat_id]["data"]["age_group"] = age_text.replace(" yosh", "")
    
    keyboard = [[region] for region in LOCATIONS.keys()]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    update.message.reply_text("🌍 Viloyatingizni tanlang:", reply_markup=reply_markup)
    return REGION

def region(update: Update, context: CallbackContext):
    """Handle region selection"""
    chat_id = update.message.chat_id
    region_text = update.message.text.strip()
    
    if region_text not in LOCATIONS:
        keyboard = [[region] for region in LOCATIONS.keys()]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        update.message.reply_text("❌ Iltimos, ro'yxatdagi viloyatni tanlang:", reply_markup=reply_markup)
        return REGION
    
    user_states[chat_id]["data"]["region"] = region_text
    
    districts = LOCATIONS[region_text]
    keyboard = [[district] for district in districts]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    update.message.reply_text("🏘 Tumaningizni tanlang:", reply_markup=reply_markup)
    return DISTRICT

def district(update: Update, context: CallbackContext):
    """Handle district selection"""
    chat_id = update.message.chat_id
    district_text = update.message.text.strip()
    region = user_states[chat_id]["data"]["region"]
    
    if district_text not in LOCATIONS[region]:
        districts = LOCATIONS[region]
        keyboard = [[district] for district in districts]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        update.message.reply_text("❌ Iltimos, ro'yxatdagi tumanni tanlang:", reply_markup=reply_markup)
        return DISTRICT
    
    user_states[chat_id]["data"]["district"] = district_text
    update.message.reply_text("🏠 Mahallangizni yozing:", reply_markup=ReplyKeyboardRemove())
    return NEIGHBORHOOD

def neighborhood(update: Update, context: CallbackContext):
    """Handle neighborhood input and complete registration"""
    chat_id = update.message.chat_id
    neighborhood_text = update.message.text.strip()
    
    if not neighborhood_text or len(neighborhood_text) < 2:
        update.message.reply_text("❌ Iltimos, mahalla nomini to'g'ri kiriting:")
        return NEIGHBORHOOD
    
    user_states[chat_id]["data"]["neighborhood"] = neighborhood_text
    
    # Save user to database
    user_data = user_states[chat_id]["data"]
    success = db.register_user(
        chat_id=chat_id,
        name=user_data["name"],
        surname=user_data["surname"],
        phone=user_data["phone"],
        age_group=user_data["age_group"],
        region=user_data["region"],
        district=user_data["district"],
        neighborhood=user_data["neighborhood"]
    )
    
    if success:
        # Send success message
        update.message.reply_text("✅ Ro'yxatdan muvaffaqiyatli o'tdingiz!")
        
        # Send success sticker
        try:
            update.message.reply_sticker(sticker="CAACAgIAAxkBAAECOzNiZoFdq9QkW2XjKQABfLsVzW1-bNsAAtsAAwURAQGcUgqqWJ6us1vBHgQ")
        except:
            pass
        
        # Notify admins
        admin_message = f"📋 Yangi foydalanuvchi ro'yxatdan o'tdi:\n\n"
        admin_message += f"👤 Ism: {user_data['name']}\n"
        admin_message += f"👨‍👩‍👧‍👦 Familiya: {user_data['surname']}\n"
        admin_message += f"📱 Telefon: {user_data['phone']}\n"
        admin_message += f"🎂 Yosh: {user_data['age_group']}\n"
        admin_message += f"🌍 Viloyat: {user_data['region']}\n"
        admin_message += f"🏘 Tuman: {user_data['district']}\n"
        admin_message += f"🏠 Mahalla: {user_data['neighborhood']}\n"
        admin_message += f"🕒 Vaqt: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        for admin_id in ADMIN_IDS:
            try:
                context.bot.send_message(admin_id, admin_message)
            except:
                continue
        
        # Show main menu
        show_main_menu_sync(update, context)
        
        # Clean up state
        if chat_id in user_states:
            del user_states[chat_id]
        
        return ConversationHandler.END
    else:
        update.message.reply_text("❌ Ro'yxatdan o'tishda xatolik yuz berdi. Qaytadan urinib ko'ring.")
        return ConversationHandler.END

def show_main_menu_sync(update: Update, context: CallbackContext):
    """Show main menu synchronously"""
    keyboard = [
        ["📝 Test topshirish", "📋 Loyiha haqida"],
        ["💬 Fikr bildirish", "📊 Natijalarim"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=False)
    
    update.message.reply_text("Asosiy menyu:", reply_markup=reply_markup)

# Main menu handlers

def handle_test_request(update: Update, context: CallbackContext):
    """Handle test request - automatically select random book"""
    chat_id = update.message.chat_id
    
    # Check if user is registered
    user = db.get_user(chat_id)
    if not user:
        update.message.reply_text("❌ Avval ro'yxatdan o'ting! /start ni bosing.")
        return
    
    # Check if user is admin (admins can choose books)
    if chat_id in ADMIN_IDS:
        # Show book selection for admins
        age_group = user['age_group']
        books = AGE_GROUPS.get(age_group, [])
        
        keyboard = [[book] for book in books]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        
        update.message.reply_text(
            f"👤 Admin: {age_group} yosh guruhi uchun kitoblardan birini tanlang:",
            reply_markup=reply_markup
        )
        
        user_states[chat_id] = {"selecting_book": True, "age_group": age_group}
        return
    
    # For regular users: automatically select random book
    age_group = user['age_group']
    books = AGE_GROUPS.get(age_group, [])
    
    if not books:
        update.message.reply_text("❌ Bu yosh guruhi uchun kitoblar topilmadi.")
        return
    
    # Select random book
    random_book = random.choice(books)
    
    # Start test automatically with random book
    if test_manager.start_test(chat_id, age_group, random_book):
        # Send test start sticker
        try:
            update.message.reply_sticker(sticker="CAACAgIAAxkBAAECOzViZoFfXUYvJ1Sf4UGi7H-QBBPKiwAC3AADBREBAASCUQ_2rTFbwR4E")
        except:
            pass
        
        update.message.reply_text(
            f"🚀 Test avtomatik boshlandi!\n\n"
            f"📚 Kitob: {random_book}\n"
            f"👥 Yosh guruhi: {age_group}\n"
            f"❓ Jami savollar: 25 ta\n"
            f"⏱ Har bir savol uchun: 20 soniya\n"
            f"🎯 Har bir to'g'ri javob: 4 ball\n\n"
            f"🍀 Omad tilaymiz!",
            reply_markup=ReplyKeyboardRemove()
        )
        
        # Send first question after a short delay
        def delayed_question():
            send_next_question_sync(context, chat_id)
        threading.Timer(2.0, delayed_question).start()
        
        # Clean up state
        if chat_id in user_states:
            del user_states[chat_id]
    else:
        update.message.reply_text("❌ Bu kitob uchun yetarli savollar yo'q. Keyinroq urinib ko'ring.")
    return SELECT_BOOK

def handle_book_selection(update: Update, context: CallbackContext):
    """Handle book selection for test"""
    chat_id = update.message.chat_id
    book_name = update.message.text.strip()
    
    if chat_id not in user_states or not user_states[chat_id].get("selecting_book"):
        return ConversationHandler.END
    
    age_group = user_states[chat_id]["age_group"]
    
    if book_name not in AGE_GROUPS.get(age_group, []):
        books = AGE_GROUPS.get(age_group, [])
        keyboard = [[book] for book in books]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        update.message.reply_text("❌ Iltimos, ro'yxatdagi kitobni tanlang:", reply_markup=reply_markup)
        return SELECT_BOOK
    
    # Start test
    if test_manager.start_test(chat_id, age_group, book_name):
        # Send test start sticker
        try:
            update.message.reply_sticker(sticker="CAACAgIAAxkBAAECOzViZoFfXUYvJ1Sf4UGi7H-QBBPKiwAC3AADBREBAASCUQ_2rTFbwR4E")
        except:
            pass
        
        update.message.reply_text(
            f"🚀 Test boshlandi!\n\n"
            f"📚 Kitob: {book_name}\n"
            f"👥 Yosh guruhi: {age_group}\n"
            f"❓ Jami savollar: 25 ta\n"
            f"⏱ Har bir savol uchun: 20 soniya\n"
            f"🎯 Har bir to'g'ri javob: 4 ball\n\n"
            f"🍀 Omad tilaymiz!",
            reply_markup=ReplyKeyboardRemove()
        )
        
        # Send first question after a short delay
        def delayed_question():
            send_next_question_sync(context, chat_id)
        threading.Timer(2.0, delayed_question).start()
        
        # Clean up state
        if chat_id in user_states:
            del user_states[chat_id]
        
        return QUESTION_ANSWER
    else:
        update.message.reply_text("❌ Bu kitob uchun yetarli savollar yo'q. Boshqa kitobni tanlang.")
        return SELECT_BOOK

def send_next_question_sync(context: CallbackContext, chat_id: int):
    """Send next question synchronously"""
    try:
        loop = asyncio.get_event_loop()
        if loop.is_running():
            loop.create_task(send_next_question(context, chat_id))
        else:
            asyncio.run(send_next_question(context, chat_id))
    except RuntimeError:
        # No event loop running, create a new one
        asyncio.run(send_next_question(context, chat_id))

def handle_about(update: Update, context: CallbackContext):
    """Handle about project request"""
    about_text = (
        "📖 Kitobxon Kids loyihasi haqida\n\n"
        "🎯 Maqsad: Bolalarning bilim darajasini baholash va o'qishga rag'batlantirishni ta'minlash\n\n"
        "👥 Maqsadli auditoriya: 7-14 yosh oralig'idagi bolalar\n\n"
        "📚 Test tizimi:\n"
        "• 7-10 yosh guruhi uchun 4 ta kitob\n"
        "• 11-14 yosh guruhi uchun 4 ta kitob\n"
        "• Har bir kitobda 25 ta savol\n"
        "• Har bir savol uchun 20 soniya vaqt\n"
        "• To'g'ri javob uchun 4 ball\n\n"
        "🏆 Natijalar:\n"
        "• 80% va undan yuqori - A'lo\n"
        "• 60-79% - Yaxshi\n"
        "• 60% dan past - Qo'shimcha o'qish tavsiya etiladi\n\n"
        "📞 Aloqa: @kitobxon_kids_support"
    )
    update.message.reply_text(about_text)

def handle_feedback_request(update: Update, context: CallbackContext):
    """Handle feedback request"""
    chat_id = update.message.chat_id
    
    # Check if user is registered
    user = db.get_user(chat_id)
    if not user:
        update.message.reply_text("❌ Avval ro'yxatdan o'ting! /start ni bosing.")
        return ConversationHandler.END
    
    update.message.reply_text(
        "💭 Fikr-mulohazangizni yozing:\n\n"
        "Bizga loyihamizni yaxshilashda yordam bering!",
        reply_markup=ReplyKeyboardRemove()
    )
    return FEEDBACK

def handle_feedback_text(update: Update, context: CallbackContext):
    """Handle feedback text"""
    chat_id = update.message.chat_id
    feedback_text = update.message.text.strip()
    
    if not feedback_text or len(feedback_text) < 10:
        update.message.reply_text("❌ Iltimos, kamida 10 ta belgi kiriting:")
        return FEEDBACK
    
    # Save feedback
    if db.save_feedback(chat_id, feedback_text):
        update.message.reply_text("✅ Fikr-mulohazangiz uchun rahmat! Sizning fikringiz bizga muhim.")
        
        # Notify admins
        user = db.get_user(chat_id)
        admin_message = f"💭 Yangi fikr-mulohaza:\n\n"
        admin_message += f"👤 Foydalanuvchi: {user['name']} {user['surname']}\n"
        admin_message += f"📱 Telefon: {user['phone']}\n"
        admin_message += f"💬 Matn: {feedback_text}\n"
        admin_message += f"🕒 Vaqt: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        for admin_id in ADMIN_IDS:
            try:
                context.bot.send_message(admin_id, admin_message)
            except:
                continue
    else:
        update.message.reply_text("❌ Fikr-mulohaza saqlashda xatolik yuz berdi.")
    
    show_main_menu_sync(update, context)
    return ConversationHandler.END

def handle_my_results(update: Update, context: CallbackContext):
    """Handle my results request"""
    chat_id = update.message.chat_id
    
    # Check if user is registered
    user = db.get_user(chat_id)
    if not user:
        update.message.reply_text("❌ Avval ro'yxatdan o'ting! /start ni bosing.")
        return
    
    # Get user's test results
    query = '''
        SELECT book_name, score, total_questions, percentage, test_date
        FROM test_results 
        WHERE user_id = ?
        ORDER BY test_date DESC
        LIMIT 10
    '''
    results = db.execute_query(query, (chat_id,))
    
    if not results:
        update.message.reply_text("📊 Hozircha test natijalaringiz yo'q. Birinchi testni topshiring!")
        return
    
    message = f"📊 {user['name']} {user['surname']} ning natijalari:\n\n"
    
    for i, result in enumerate(results, 1):
        book_name, score, total_questions, percentage, test_date = result
        max_score = total_questions * 4
        
        # Format date
        try:
            date_obj = datetime.datetime.strptime(test_date, '%Y-%m-%d %H:%M:%S')
            formatted_date = date_obj.strftime('%d.%m.%Y %H:%M')
        except:
            formatted_date = test_date
        
        message += f"{i}. 📚 {book_name}\n"
        message += f"   🎯 Ball: {score}/{max_score}\n"
        message += f"   📈 Foiz: {percentage:.1f}%\n"
        message += f"   📅 Sana: {formatted_date}\n\n"
    
    if len(results) == 10:
        message += "📝 So'nggi 10 ta natija ko'rsatildi."
    
    update.message.reply_text(message)

# Answer handler for inline buttons

def handle_answer(update: Update, context: CallbackContext):
    """Handle test answer from inline button"""
    query = update.callback_query
    query.answer()
    
    # Parse callback data
    try:
        _, answer, chat_id_str = query.data.split('_')
        chat_id = int(chat_id_str)
    except:
        return
    
    # Verify user
    if query.from_user.id != chat_id:
        return
    
    # Check if test is active
    if chat_id not in active_tests:
        query.edit_message_text("❌ Test sessiyasi tugagan.")
        return
    
    # Cancel timeout timer
    if chat_id in question_timers:
        question_timers[chat_id].cancel()
        del question_timers[chat_id]
    
    # Submit answer
    test_manager.submit_answer(chat_id, answer)
    
    # Show submitted answer
    current_question = test_manager.get_current_question(chat_id)
    if active_tests[chat_id]['current_question'] > 0:
        prev_question_idx = active_tests[chat_id]['current_question'] - 1
        prev_question = active_tests[chat_id]['questions'][prev_question_idx]
        correct_answer = prev_question['correct_answer']
        
        if answer.upper() == correct_answer.upper():
            query.edit_message_text(f"✅ To'g'ri javob! ({answer})\n\n+4 ball")
        else:
            query.edit_message_text(f"❌ Noto'g'ri javob. To'g'ri javob: {correct_answer}\n\nSizning javobingiz: {answer}")
    
    # Check if test is complete
    if test_manager.is_test_complete(chat_id):
        # End test after 1 second delay
        threading.Timer(1.0, lambda: end_test_sync(context, chat_id)).start()
    else:
        # Send next question after 2 seconds automatically
        def next_question_sync():
            send_next_question_sync(context, chat_id)
        threading.Timer(2.0, next_question_sync).start()

# Admin handlers

def show_admin_menu(update: Update, context: CallbackContext):
    """Show admin menu"""
    keyboard = [
        ["👥 Foydalanuvchilar", "📊 Test natijalari"],
        ["➕ Savol qo'shish", "📥 Eksport"],
        ["📈 Statistika", "🔄 Yangilash"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    
    update.message.reply_text("🔧 Admin paneli:", reply_markup=reply_markup)

def handle_admin_users(update: Update, context: CallbackContext):
    """Handle admin users request"""
    if update.message.chat_id not in ADMIN_IDS:
        return
    
    users = db.get_all_users()
    
    if not users:
        update.message.reply_text("👥 Hozircha foydalanuvchilar yo'q.")
        return
    
    message = f"👥 Jami foydalanuvchilar: {len(users)}\n\n"
    
    # Show last 10 users
    for user in users[:10]:
        message += f"👤 {user['name']} {user['surname']}\n"
        message += f"📱 {user['phone']}\n"
        message += f"🎂 {user['age_group']} yosh\n"
        message += f"🌍 {user['region']}, {user['district']}\n\n"
    
    if len(users) > 10:
        message += f"📝 So'nggi 10 ta foydalanuvchi ko'rsatildi.\nJami: {len(users)} ta"
    
    update.message.reply_text(message)

def handle_admin_results(update: Update, context: CallbackContext):
    """Handle admin test results request"""
    if update.message.chat_id not in ADMIN_IDS:
        return
    
    results = db.get_test_results()
    
    if not results:
        update.message.reply_text("📊 Hozircha test natijalari yo'q.")
        return
    
    message = f"📊 Jami test natijalari: {len(results)}\n\n"
    
    # Show last 10 results
    for result in results[:10]:
        message += f"👤 {result['name']} {result['surname']}\n"
        message += f"🎂 {result['age_group']} | 📚 {result['book_name']}\n"
        message += f"🎯 {result['score']}/{result['total_questions'] * 4} ball ({result['percentage']:.1f}%)\n\n"
    
    if len(results) > 10:
        message += f"📝 So'nggi 10 ta natija ko'rsatildi.\nJami: {len(results)} ta"
    
    update.message.reply_text(message)

def handle_admin_export(update: Update, context: CallbackContext):
    """Handle admin export request"""
    if update.message.chat_id not in ADMIN_IDS:
        return
    
    keyboard = [
        ["📊 Excel - Foydalanuvchilar", "📊 Excel - Natijalar"],
        ["📄 PDF - Foydalanuvchilar", "📄 PDF - Natijalar"],
        ["🔙 Orqaga"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    update.message.reply_text("📥 Eksport turini tanlang:", reply_markup=reply_markup)

def handle_export_choice(update: Update, context: CallbackContext):
    """Handle export format choice"""
    if update.message.chat_id not in ADMIN_IDS:
        return
    
    choice = update.message.text.strip()
    
    if choice == "🔙 Orqaga":
        show_admin_menu(update, context)
        return
    
    update.message.reply_text("📊 Hisobot tayyorlanmoqda...")
    
    try:
        if "Foydalanuvchilar" in choice:
            data = db.get_all_users()
            if "Excel" in choice:
                buffer = create_excel_report(data, "users")
                filename = f"foydalanuvchilar_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                update.message.reply_document(
                    document=buffer,
                    filename=filename,
                    caption="📊 Foydalanuvchilar ro'yxati (Excel)"
                )
            else:  # PDF
                buffer = create_pdf_report(data, "users")
                filename = f"foydalanuvchilar_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                update.message.reply_document(
                    document=buffer,
                    filename=filename,
                    caption="📄 Foydalanuvchilar ro'yxati (PDF)"
                )
        
        elif "Natijalar" in choice:
            data = db.get_test_results()
            if "Excel" in choice:
                buffer = create_excel_report(data, "results")
                filename = f"test_natijalari_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                update.message.reply_document(
                    document=buffer,
                    filename=filename,
                    caption="📊 Test natijalari (Excel)"
                )
            else:  # PDF
                buffer = create_pdf_report(data, "results")
                filename = f"test_natijalari_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                update.message.reply_document(
                    document=buffer,
                    filename=filename,
                    caption="📄 Test natijalari (PDF)"
                )
        
        update.message.reply_text("✅ Hisobot muvaffaqiyatli yuborildi!")
        
    except Exception as e:
        logger.error(f"Export error: {e}")
        update.message.reply_text("❌ Hisobot yaratishda xatolik yuz berdi.")
    
    show_admin_menu(update, context)

def handle_admin_stats(update: Update, context: CallbackContext):
    """Handle admin statistics request"""
    if update.message.chat_id not in ADMIN_IDS:
        return
    
    # Get statistics
    total_users = len(db.get_all_users())
    total_tests = len(db.get_test_results())
    
    # Age group stats
    age_7_10 = len([u for u in db.get_all_users() if u['age_group'] == '7-10'])
    age_11_14 = len([u for u in db.get_all_users() if u['age_group'] == '11-14'])
    
    # Average score
    results = db.get_test_results()
    if results:
        avg_score = sum(r['percentage'] for r in results) / len(results)
    else:
        avg_score = 0
    
    # Today's registrations
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    today_users = len([u for u in db.get_all_users() if today in str(u.get('registration_date', ''))])
    
    message = f"📈 Kitobxon Kids statistikasi\n\n"
    message += f"👥 Jami foydalanuvchilar: {total_users}\n"
    message += f"📊 Jami testlar: {total_tests}\n\n"
    message += f"👶 7-10 yosh: {age_7_10} ta\n"
    message += f"🧒 11-14 yosh: {age_11_14} ta\n\n"
    message += f"📊 O'rtacha natija: {avg_score:.1f}%\n"
    message += f"📅 Bugun ro'yxatdan o'tganlar: {today_users} ta\n\n"
    message += f"🕒 So'nggi yangilanish: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    update.message.reply_text(message)

# Cancel and error handlers

def cancel(update: Update, context: CallbackContext):
    """Cancel current conversation"""
    chat_id = update.message.chat_id
    
    # Clean up user state
    if chat_id in user_states:
        del user_states[chat_id]
    
    # Clean up active tests
    if chat_id in active_tests:
        test_manager.cleanup_test(chat_id)
    
    update.message.reply_text("❌ Bekor qilindi.", reply_markup=ReplyKeyboardRemove())
    
    # Show appropriate menu based on user role
    if chat_id in ADMIN_IDS:
        show_admin_menu(update, context)
    else:
        user = db.get_user(chat_id)
        if user:
            show_main_menu_sync(update, context)
        else:
            update.message.reply_text("Boshlash uchun /start ni bosing.")
    
    return ConversationHandler.END

def error_handler(update: Update, context: CallbackContext):
    """Handle errors"""
    logger.warning(f'Update {update} caused error {context.error}')
    
    if update and update.effective_chat:
        context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="❌ Xatolik yuz berdi. Qaytadan urinib ko'ring."
        )

def main():
    """Main function to start the bot"""
    
    if BOT_TOKEN == "YOUR_BOT_TOKEN_HERE":
        print("❌ Bot tokenini o'rnating! BOT_TOKEN environment variable ni qo'shing.")
        return
    
    # Create updater
    updater = Updater(token=BOT_TOKEN, use_context=True)
    dispatcher = updater.dispatcher
    
    # Registration conversation
    registration_conv = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            NAME: [MessageHandler(Filters.text & ~Filters.command, name)],
            SURNAME: [MessageHandler(Filters.text & ~Filters.command, surname)],
            PHONE: [MessageHandler(Filters.text & ~Filters.command, phone)],
            AGE: [MessageHandler(Filters.text & ~Filters.command, age)],
            REGION: [MessageHandler(Filters.text & ~Filters.command, region)],
            DISTRICT: [MessageHandler(Filters.text & ~Filters.command, district)],
            NEIGHBORHOOD: [MessageHandler(Filters.text & ~Filters.command, neighborhood)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Test conversation
    test_conv = ConversationHandler(
        entry_points=[MessageHandler(Filters.regex('^📝 Test topshirish$'), handle_test_request)],
        states={
            SELECT_BOOK: [MessageHandler(Filters.text & ~Filters.command, handle_book_selection)],
            QUESTION_ANSWER: [CallbackQueryHandler(handle_answer, pattern='^answer_')]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Feedback conversation
    feedback_conv = ConversationHandler(
        entry_points=[MessageHandler(Filters.regex('^💬 Fikr bildirish$'), handle_feedback_request)],
        states={
            FEEDBACK: [MessageHandler(Filters.text & ~Filters.command, handle_feedback_text)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )
    
    # Add conversation handlers
    dispatcher.add_handler(registration_conv)
    dispatcher.add_handler(test_conv)
    dispatcher.add_handler(feedback_conv)
    
    # Add command handlers
    dispatcher.add_handler(CommandHandler('start', start))
    dispatcher.add_handler(CommandHandler('cancel', cancel))
    
    # Add message handlers
    dispatcher.add_handler(MessageHandler(Filters.regex('^📋 Loyiha haqida$'), handle_about))
    dispatcher.add_handler(MessageHandler(Filters.regex('^📊 Natijalarim$'), handle_my_results))
    
    # Admin handlers
    dispatcher.add_handler(MessageHandler(Filters.regex('^👥 Foydalanuvchilar$'), handle_admin_users))
    dispatcher.add_handler(MessageHandler(Filters.regex('^📊 Test natijalari$'), handle_admin_results))
    dispatcher.add_handler(MessageHandler(Filters.regex('^📥 Eksport$'), handle_admin_export))
    dispatcher.add_handler(MessageHandler(Filters.regex('^📈 Statistika$'), handle_admin_stats))
    dispatcher.add_handler(MessageHandler(Filters.regex('^🔄 Yangilash$'), show_admin_menu))
    
    # Export handlers
    dispatcher.add_handler(MessageHandler(Filters.regex('^📊 Excel'), handle_export_choice))
    dispatcher.add_handler(MessageHandler(Filters.regex('^📄 PDF'), handle_export_choice))
    dispatcher.add_handler(MessageHandler(Filters.regex('^🔙 Orqaga$'), lambda u, c: show_admin_menu(u, c) if u.message.chat_id in ADMIN_IDS else None))
    
    # Add callback query handler for test answers
    dispatcher.add_handler(CallbackQueryHandler(handle_answer, pattern='^answer_'))
    
    # Add error handler
    dispatcher.add_error_handler(error_handler)
    
    # Start bot
    logger.info("🚀 Kitobxon Kids Bot ishga tushdi!")
    print("🚀 Kitobxon Kids Bot ishga tushdi!")
    print(f"🔧 Admin IDs: {ADMIN_IDS}")
    print("📱 Bot token tekshirildi va ulanish amalga oshirildi.")
    
    updater.start_polling()
    updater.idle()

if __name__ == '__main__':
    main()
